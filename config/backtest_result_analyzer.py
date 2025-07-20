"""
Module: Backtest Result Analyzer
File: backtest_result_analyzer.py
Description: 
  4-2-2「複合戦略バックテスト機能実装」- Result Analyzer
  バックテスト結果の分析とExcel/可視化レポート生成

Author: imega
Created: 2025-07-20
Modified: 2025-07-20

Functions:
  - バックテスト結果分析
  - Excel詳細レポート生成
  - 可視化レポート作成
  - 戦略比較分析
"""

import os
import sys
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, field, asdict
from enum import Enum
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
import warnings
warnings.filterwarnings('ignore')

# Excel処理用
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.charts import ScatterChart, LineChart, BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.getLogger(__name__).warning("openpyxl not available, Excel reports disabled")

# プロジェクトパスの追加
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

# 既存システムインポート
try:
    from config.enhanced_performance_calculator import EnhancedPerformanceCalculator, PerformanceAnalysis, CompositeStrategyPerformance
    from config.backtest_scenario_generator import TestScenario, ScenarioGenerationResult
    PERFORMANCE_CALCULATOR_AVAILABLE = True
except ImportError:
    PERFORMANCE_CALCULATOR_AVAILABLE = False
    logging.getLogger(__name__).warning("Performance calculator not available")

# ロガーの設定
logger = logging.getLogger(__name__)

class AnalysisType(Enum):
    """分析タイプ"""
    PERFORMANCE_ANALYSIS = "performance_analysis"
    RISK_ANALYSIS = "risk_analysis"
    SCENARIO_ANALYSIS = "scenario_analysis"
    COMPARISON_ANALYSIS = "comparison_analysis"
    ATTRIBUTION_ANALYSIS = "attribution_analysis"

class ReportFormat(Enum):
    """レポート形式"""
    EXCEL_DETAILED = "excel_detailed"
    HTML_VISUALIZATION = "html_visualization"
    PDF_SUMMARY = "pdf_summary"
    JSON_DATA = "json_data"

@dataclass
class AnalysisResult:
    """分析結果"""
    analysis_id: str
    analysis_type: AnalysisType
    analysis_date: datetime
    summary_metrics: Dict[str, Any]
    detailed_results: Dict[str, Any]
    charts_data: Dict[str, Any]
    recommendations: List[str]
    warnings: List[str]
    data_quality_score: float

@dataclass
class ComparativeAnalysis:
    """比較分析結果"""
    comparison_id: str
    strategies_compared: List[str]
    performance_ranking: Dict[str, int]
    statistical_significance: Dict[str, float]
    correlation_analysis: pd.DataFrame
    risk_return_scatter: Dict[str, Any]
    drawdown_comparison: Dict[str, Any]
    regime_performance: Dict[str, Dict[str, Any]]

class BacktestResultAnalyzer:
    """バックテスト結果分析器"""
    
    def __init__(self, config_path: Optional[str] = None, output_dir: str = "output"):
        """分析器の初期化"""
        self.logger = logging.getLogger(__name__)
        self.output_dir = output_dir
        
        # 出力ディレクトリの作成
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(os.path.join(output_dir, "charts"), exist_ok=True)
        os.makedirs(os.path.join(output_dir, "excel"), exist_ok=True)
        os.makedirs(os.path.join(output_dir, "html"), exist_ok=True)
        
        # 設定の読み込み
        self.config = self._load_analysis_config(config_path)
        self.chart_settings = self.config.get('chart_settings', {})
        self.excel_settings = self.config.get('excel_settings', {})
        
        # パフォーマンス計算器の初期化
        self.performance_calculator = None
        if PERFORMANCE_CALCULATOR_AVAILABLE:
            try:
                self.performance_calculator = EnhancedPerformanceCalculator()
            except Exception as e:
                self.logger.warning(f"Failed to initialize performance calculator: {e}")
        
        # 分析統計
        self.analysis_stats = {
            "total_analyses": 0,
            "successful_reports": 0,
            "failed_reports": 0,
            "last_analysis": None
        }
        
        # カラーパレット
        self.colors = {
            'primary': '#1f77b4',
            'secondary': '#ff7f0e', 
            'success': '#2ca02c',
            'danger': '#d62728',
            'warning': '#ff9500',
            'info': '#17a2b8'
        }
        
        # 日本語フォント設定
        plt.rcParams['font.family'] = 'DejaVu Sans'
        sns.set_style("whitegrid")
        
        self.logger.info("BacktestResultAnalyzer initialized")
    
    def _load_analysis_config(self, config_path: Optional[str] = None) -> Dict[str, Any]:
        """分析設定の読み込み"""
        
        if config_path is None:
            config_path = os.path.join(
                os.path.dirname(__file__), 
                "backtest", 
                "composite_backtest_config.json"
            )
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            analysis_config = config.get('analysis_settings', {})
            self.logger.info(f"Analysis config loaded from {config_path}")
            return analysis_config
        except Exception as e:
            self.logger.warning(f"Failed to load analysis config: {e}")
            return self._get_default_analysis_config()
    
    def _get_default_analysis_config(self) -> Dict[str, Any]:
        """デフォルト分析設定"""
        
        return {
            "chart_settings": {
                "figure_size": [12, 8],
                "dpi": 300,
                "style": "whitegrid",
                "color_palette": "husl"
            },
            "excel_settings": {
                "include_charts": True,
                "freeze_panes": True,
                "auto_filter": True,
                "conditional_formatting": True
            },
            "analysis_thresholds": {
                "min_sharpe_ratio": 1.0,
                "max_drawdown_threshold": 0.20,
                "min_win_rate": 0.45,
                "statistical_significance": 0.05
            }
        }
    
    def analyze_backtest_results(self, 
                               backtest_results: Dict[str, Any],
                               scenarios: Optional[List[TestScenario]] = None) -> AnalysisResult:
        """バックテスト結果の分析"""
        
        try:
            analysis_id = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # 基本メトリクスの抽出
            summary_metrics = self._extract_summary_metrics(backtest_results)
            
            # 詳細分析の実行
            detailed_results = {}
            
            # パフォーマンス分析
            if 'performance_data' in backtest_results:
                detailed_results['performance'] = self._analyze_performance_detailed(
                    backtest_results['performance_data']
                )
            
            # リスク分析
            if 'returns_data' in backtest_results:
                detailed_results['risk'] = self._analyze_risk_detailed(
                    backtest_results['returns_data']
                )
            
            # シナリオ分析
            if scenarios:
                detailed_results['scenarios'] = self._analyze_scenarios(
                    backtest_results, scenarios
                )
            
            # チャートデータの準備
            charts_data = self._prepare_charts_data(backtest_results, detailed_results)
            
            # 推奨事項の生成
            recommendations = self._generate_recommendations(summary_metrics, detailed_results)
            
            # 警告の生成
            warnings = self._generate_warnings(summary_metrics, detailed_results)
            
            # データ品質スコア
            data_quality_score = self._calculate_data_quality_score(backtest_results)
            
            # 分析結果の作成
            analysis_result = AnalysisResult(
                analysis_id=analysis_id,
                analysis_type=AnalysisType.PERFORMANCE_ANALYSIS,
                analysis_date=datetime.now(),
                summary_metrics=summary_metrics,
                detailed_results=detailed_results,
                charts_data=charts_data,
                recommendations=recommendations,
                warnings=warnings,
                data_quality_score=data_quality_score
            )
            
            # 統計更新
            self.analysis_stats["total_analyses"] += 1
            self.analysis_stats["last_analysis"] = datetime.now()
            
            self.logger.info(f"Analysis completed: {analysis_id}")
            return analysis_result
            
        except Exception as e:
            self.logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_summary_metrics(self, backtest_results: Dict[str, Any]) -> Dict[str, Any]:
        """サマリーメトリクスの抽出"""
        
        summary = {}
        
        # 基本パフォーマンス
        if 'total_return' in backtest_results:
            summary['total_return'] = backtest_results['total_return']
        if 'sharpe_ratio' in backtest_results:
            summary['sharpe_ratio'] = backtest_results['sharpe_ratio']
        if 'max_drawdown' in backtest_results:
            summary['max_drawdown'] = backtest_results['max_drawdown']
        if 'win_rate' in backtest_results:
            summary['win_rate'] = backtest_results['win_rate']
        
        # 期間情報
        if 'start_date' in backtest_results and 'end_date' in backtest_results:
            start_date = pd.to_datetime(backtest_results['start_date'])
            end_date = pd.to_datetime(backtest_results['end_date'])
            summary['test_period_days'] = (end_date - start_date).days
            summary['start_date'] = start_date
            summary['end_date'] = end_date
        
        # 戦略情報
        if 'strategy_combinations' in backtest_results:
            summary['strategies_tested'] = len(backtest_results['strategy_combinations'])
        
        return summary
    
    def _analyze_performance_detailed(self, performance_data: Dict[str, Any]) -> Dict[str, Any]:
        """詳細パフォーマンス分析"""
        
        analysis = {}
        
        try:
            # 月次パフォーマンス
            if 'monthly_returns' in performance_data:
                monthly_returns = pd.Series(performance_data['monthly_returns'])
                analysis['monthly_stats'] = {
                    'best_month': monthly_returns.max(),
                    'worst_month': monthly_returns.min(),
                    'positive_months': (monthly_returns > 0).sum(),
                    'negative_months': (monthly_returns < 0).sum(),
                    'monthly_volatility': monthly_returns.std()
                }
            
            # ローリング分析
            if 'daily_returns' in performance_data:
                daily_returns = pd.Series(performance_data['daily_returns'])
                analysis['rolling_analysis'] = {
                    'rolling_sharpe_3m': self._calculate_rolling_sharpe(daily_returns, 63),
                    'rolling_sharpe_6m': self._calculate_rolling_sharpe(daily_returns, 126),
                    'rolling_volatility': self._calculate_rolling_volatility(daily_returns, 21)
                }
            
            # リターン分布分析
            if 'daily_returns' in performance_data:
                returns = pd.Series(performance_data['daily_returns'])
                analysis['return_distribution'] = {
                    'skewness': returns.skew(),
                    'kurtosis': returns.kurtosis(),
                    'jarque_bera_test': stats.jarque_bera(returns.dropna()),
                    'percentiles': {
                        '5%': returns.quantile(0.05),
                        '25%': returns.quantile(0.25),
                        '75%': returns.quantile(0.75),
                        '95%': returns.quantile(0.95)
                    }
                }
        
        except Exception as e:
            self.logger.warning(f"Detailed performance analysis failed: {e}")
        
        return analysis
    
    def _analyze_risk_detailed(self, returns_data: Union[pd.Series, List, Dict]) -> Dict[str, Any]:
        """詳細リスク分析"""
        
        analysis = {}
        
        try:
            # データの正規化
            if isinstance(returns_data, dict):
                returns = pd.Series(returns_data)
            elif isinstance(returns_data, list):
                returns = pd.Series(returns_data)
            else:
                returns = returns_data
            
            returns = returns.dropna()
            
            # VaRとCVaR分析
            confidence_levels = [0.95, 0.99]
            var_analysis = {}
            
            for confidence in confidence_levels:
                var = returns.quantile(1 - confidence)
                cvar = returns[returns <= var].mean()
                var_analysis[f'VaR_{int(confidence*100)}'] = var
                var_analysis[f'CVaR_{int(confidence*100)}'] = cvar
            
            analysis['var_analysis'] = var_analysis
            
            # テールリスク分析
            analysis['tail_risk'] = {
                'tail_ratio': self._calculate_tail_ratio(returns),
                'extreme_losses': (returns < returns.quantile(0.01)).sum(),
                'extreme_gains': (returns > returns.quantile(0.99)).sum()
            }
            
            # ドローダウン分析
            cumulative_returns = (1 + returns).cumprod()
            running_max = cumulative_returns.expanding().max()
            drawdown = (cumulative_returns - running_max) / running_max
            
            analysis['drawdown_analysis'] = {
                'max_drawdown': drawdown.min(),
                'avg_drawdown': drawdown[drawdown < 0].mean(),
                'drawdown_duration': self._calculate_drawdown_duration(drawdown),
                'recovery_time': self._calculate_recovery_time(drawdown)
            }
            
        except Exception as e:
            self.logger.warning(f"Detailed risk analysis failed: {e}")
        
        return analysis
    
    def _analyze_scenarios(self, 
                         backtest_results: Dict[str, Any], 
                         scenarios: List[TestScenario]) -> Dict[str, Any]:
        """シナリオ分析"""
        
        scenario_analysis = {}
        
        try:
            for scenario in scenarios:
                scenario_id = scenario.scenario_id
                
                # シナリオ固有の結果を抽出
                if f'scenario_{scenario_id}' in backtest_results:
                    scenario_results = backtest_results[f'scenario_{scenario_id}']
                    
                    scenario_analysis[scenario_id] = {
                        'scenario_name': scenario.name,
                        'market_regime': [c.regime.value for c in scenario.market_conditions],
                        'performance': self._extract_scenario_performance(scenario_results),
                        'success_criteria_met': self._check_success_criteria(
                            scenario_results, scenario.success_criteria
                        ),
                        'challenges_identified': scenario.expected_challenges
                    }
            
        except Exception as e:
            self.logger.warning(f"Scenario analysis failed: {e}")
        
        return scenario_analysis
    
    def _prepare_charts_data(self, 
                           backtest_results: Dict[str, Any], 
                           detailed_results: Dict[str, Any]) -> Dict[str, Any]:
        """チャートデータの準備"""
        
        charts_data = {}
        
        try:
            # 累積リターンチャート
            if 'daily_returns' in backtest_results:
                returns = pd.Series(backtest_results['daily_returns'])
                cumulative_returns = (1 + returns).cumprod()
                charts_data['cumulative_returns'] = cumulative_returns.to_dict()
            
            # ドローダウンチャート
            if 'daily_returns' in backtest_results:
                returns = pd.Series(backtest_results['daily_returns'])
                cumulative_returns = (1 + returns).cumprod()
                running_max = cumulative_returns.expanding().max()
                drawdown = (cumulative_returns - running_max) / running_max
                charts_data['drawdown'] = drawdown.to_dict()
            
            # リターン分布ヒストグラム
            if 'daily_returns' in backtest_results:
                returns = pd.Series(backtest_results['daily_returns'])
                hist_data, bins = np.histogram(returns.dropna(), bins=50)
                charts_data['return_histogram'] = {
                    'counts': hist_data.tolist(),
                    'bins': bins.tolist()
                }
            
            # 月次リターン
            if 'monthly_returns' in backtest_results:
                charts_data['monthly_returns'] = backtest_results['monthly_returns']
            
            # ローリングメトリクス
            if 'performance' in detailed_results and 'rolling_analysis' in detailed_results['performance']:
                rolling_data = detailed_results['performance']['rolling_analysis']
                for key, value in rolling_data.items():
                    if hasattr(value, 'to_dict'):
                        charts_data[f'rolling_{key}'] = value.to_dict()
        
        except Exception as e:
            self.logger.warning(f"Chart data preparation failed: {e}")
        
        return charts_data
    
    def _generate_recommendations(self, 
                                summary_metrics: Dict[str, Any], 
                                detailed_results: Dict[str, Any]) -> List[str]:
        """推奨事項の生成"""
        
        recommendations = []
        
        try:
            # シャープレシオベースの推奨
            sharpe_ratio = summary_metrics.get('sharpe_ratio', 0)
            if sharpe_ratio < 1.0:
                recommendations.append("シャープレシオが1.0を下回っています。リスク調整リターンの改善を検討してください。")
            elif sharpe_ratio > 2.0:
                recommendations.append("優秀なシャープレシオです。現在の戦略の維持を推奨します。")
            
            # ドローダウンベースの推奨
            max_drawdown = abs(summary_metrics.get('max_drawdown', 0))
            if max_drawdown > 0.20:
                recommendations.append("最大ドローダウンが20%を超えています。リスク管理の強化を検討してください。")
            elif max_drawdown < 0.05:
                recommendations.append("ドローダウンが非常に低く抑えられています。")
            
            # 勝率ベースの推奨
            win_rate = summary_metrics.get('win_rate', 0)
            if win_rate < 0.40:
                recommendations.append("勝率が40%を下回っています。戦略の見直しやフィルターの追加を検討してください。")
            
            # リターン分布ベースの推奨
            if 'performance' in detailed_results and 'return_distribution' in detailed_results['performance']:
                dist = detailed_results['performance']['return_distribution']
                skewness = dist.get('skewness', 0)
                if skewness < -0.5:
                    recommendations.append("リターン分布が負に偏っています。テールリスクの管理を強化してください。")
            
            # リスク分析ベースの推奨
            if 'risk' in detailed_results:
                risk_analysis = detailed_results['risk']
                if 'var_analysis' in risk_analysis:
                    var_95 = risk_analysis['var_analysis'].get('VaR_95', 0)
                    if var_95 < -0.05:  # 日次5%以上の損失リスク
                        recommendations.append("VaR(95%)が高いレベルにあります。ポジションサイズの調整を検討してください。")
        
        except Exception as e:
            self.logger.warning(f"Recommendation generation failed: {e}")
        
        if not recommendations:
            recommendations.append("分析結果は概ね良好です。継続的なモニタリングを推奨します。")
        
        return recommendations
    
    def _generate_warnings(self, 
                         summary_metrics: Dict[str, Any], 
                         detailed_results: Dict[str, Any]) -> List[str]:
        """警告の生成"""
        
        warnings = []
        
        try:
            # データ不足の警告
            test_period_days = summary_metrics.get('test_period_days', 0)
            if test_period_days < 180:
                warnings.append("テスト期間が6ヶ月未満です。統計的信頼性が限定的です。")
            
            # 極端な値の警告
            sharpe_ratio = summary_metrics.get('sharpe_ratio', 0)
            if sharpe_ratio > 5.0:
                warnings.append("シャープレシオが異常に高い値です。データの確認を推奨します。")
            
            max_drawdown = abs(summary_metrics.get('max_drawdown', 0))
            if max_drawdown > 0.50:
                warnings.append("最大ドローダウンが50%を超えています。戦略の実用性に懸念があります。")
            
            # 分布の異常値警告
            if 'performance' in detailed_results and 'return_distribution' in detailed_results['performance']:
                dist = detailed_results['performance']['return_distribution']
                kurtosis = dist.get('kurtosis', 0)
                if kurtosis > 10:
                    warnings.append("リターン分布の尖度が異常に高く、極端な値が含まれている可能性があります。")
        
        except Exception as e:
            self.logger.warning(f"Warning generation failed: {e}")
        
        return warnings
    
    def generate_excel_report(self, 
                            analysis_result: AnalysisResult,
                            filename: Optional[str] = None) -> str:
        """Excel詳細レポートの生成"""
        
        if not EXCEL_AVAILABLE:
            self.logger.error("Excel report generation requires openpyxl")
            return ""
        
        try:
            if filename is None:
                filename = f"backtest_report_{analysis_result.analysis_id}.xlsx"
            
            filepath = os.path.join(self.output_dir, "excel", filename)
            
            # Excelワークブックの作成
            wb = openpyxl.Workbook()
            
            # サマリーシートの作成
            self._create_summary_sheet(wb, analysis_result)
            
            # 詳細分析シート
            self._create_detailed_analysis_sheet(wb, analysis_result)
            
            # リスク分析シート
            self._create_risk_analysis_sheet(wb, analysis_result)
            
            # 推奨事項シート
            self._create_recommendations_sheet(wb, analysis_result)
            
            # チャートシート
            if self.excel_settings.get('include_charts', True):
                self._create_charts_sheet(wb, analysis_result)
            
            # ワークブックの保存
            wb.save(filepath)
            
            self.analysis_stats["successful_reports"] += 1
            self.logger.info(f"Excel report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            self.logger.error(f"Excel report generation failed: {e}")
            self.analysis_stats["failed_reports"] += 1
            return ""
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, analysis_result: AnalysisResult):
        """サマリーシートの作成"""
        
        ws = wb.active
        ws.title = "サマリー"
        
        # ヘッダー
        ws['A1'] = "バックテスト結果サマリー"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 基本情報
        row = 3
        ws[f'A{row}'] = "分析ID"
        ws[f'B{row}'] = analysis_result.analysis_id
        row += 1
        
        ws[f'A{row}'] = "分析日時"
        ws[f'B{row}'] = analysis_result.analysis_date.strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        ws[f'A{row}'] = "データ品質スコア"
        ws[f'B{row}'] = f"{analysis_result.data_quality_score:.2f}"
        row += 2
        
        # パフォーマンスメトリクス
        ws[f'A{row}'] = "主要パフォーマンス指標"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        for key, value in analysis_result.summary_metrics.items():
            if isinstance(value, (int, float)):
                ws[f'A{row}'] = key
                if isinstance(value, float) and abs(value) < 1:
                    ws[f'B{row}'] = f"{value:.4f}"
                else:
                    ws[f'B{row}'] = f"{value:.2f}"
                row += 1
        
        # 推奨事項
        row += 1
        ws[f'A{row}'] = "推奨事項"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        for i, recommendation in enumerate(analysis_result.recommendations, 1):
            ws[f'A{row}'] = f"{i}. {recommendation}"
            row += 1
        
        # 警告
        if analysis_result.warnings:
            row += 1
            ws[f'A{row}'] = "警告"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            row += 1
            
            for i, warning in enumerate(analysis_result.warnings, 1):
                ws[f'A{row}'] = f"{i}. {warning}"
                ws[f'A{row}'].fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
                row += 1
    
    def _create_detailed_analysis_sheet(self, wb: openpyxl.Workbook, analysis_result: AnalysisResult):
        """詳細分析シートの作成"""
        
        ws = wb.create_sheet(title="詳細分析")
        
        # 詳細結果の出力
        row = 1
        for category, data in analysis_result.detailed_results.items():
            ws[f'A{row}'] = category.upper()
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 2
            
            if isinstance(data, dict):
                for key, value in data.items():
                    ws[f'A{row}'] = key
                    if isinstance(value, dict):
                        ws[f'B{row}'] = str(value)
                    else:
                        ws[f'B{row}'] = value
                    row += 1
            
            row += 1
    
    def _create_risk_analysis_sheet(self, wb: openpyxl.Workbook, analysis_result: AnalysisResult):
        """リスク分析シートの作成"""
        
        ws = wb.create_sheet(title="リスク分析")
        
        # リスク分析データ
        if 'risk' in analysis_result.detailed_results:
            risk_data = analysis_result.detailed_results['risk']
            
            row = 1
            ws[f'A{row}'] = "リスク分析結果"
            ws[f'A{row}'].font = Font(size=16, bold=True)
            row += 2
            
            for risk_category, metrics in risk_data.items():
                ws[f'A{row}'] = risk_category
                ws[f'A{row}'].font = Font(size=14, bold=True)
                row += 1
                
                if isinstance(metrics, dict):
                    for metric, value in metrics.items():
                        ws[f'B{row}'] = metric
                        ws[f'C{row}'] = value
                        row += 1
                
                row += 1
    
    def _create_recommendations_sheet(self, wb: openpyxl.Workbook, analysis_result: AnalysisResult):
        """推奨事項シートの作成"""
        
        ws = wb.create_sheet(title="推奨事項")
        
        # 推奨事項
        row = 1
        ws[f'A{row}'] = "推奨事項と改善点"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        row += 2
        
        for i, recommendation in enumerate(analysis_result.recommendations, 1):
            ws[f'A{row}'] = f"推奨 {i}"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = recommendation
            row += 2
        
        # 警告
        if analysis_result.warnings:
            row += 1
            ws[f'A{row}'] = "注意事項"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            row += 2
            
            for i, warning in enumerate(analysis_result.warnings, 1):
                ws[f'A{row}'] = f"警告 {i}"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
                ws[f'B{row}'] = warning
                row += 2
    
    def _create_charts_sheet(self, wb: openpyxl.Workbook, analysis_result: AnalysisResult):
        """チャートシートの作成"""
        
        ws = wb.create_sheet(title="チャート")
        
        # チャートデータの出力
        charts_data = analysis_result.charts_data
        
        # 累積リターンデータ
        if 'cumulative_returns' in charts_data:
            ws['A1'] = "累積リターン推移"
            ws['A1'].font = Font(size=14, bold=True)
            
            cum_returns = charts_data['cumulative_returns']
            row = 3
            ws['A2'] = "日付"
            ws['B2'] = "累積リターン"
            
            for date_str, return_val in cum_returns.items():
                ws[f'A{row}'] = date_str
                ws[f'B{row}'] = return_val
                row += 1
    
    def generate_html_visualization(self, 
                                  analysis_result: AnalysisResult,
                                  filename: Optional[str] = None) -> str:
        """HTML可視化レポートの生成"""
        
        try:
            if filename is None:
                filename = f"visualization_report_{analysis_result.analysis_id}.html"
            
            filepath = os.path.join(self.output_dir, "html", filename)
            
            # チャートの生成
            chart_paths = self._generate_visualization_charts(analysis_result)
            
            # HTMLの作成
            html_content = self._create_html_template(analysis_result, chart_paths)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"HTML visualization generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"HTML visualization generation failed: {e}")
            return ""
    
    def _generate_visualization_charts(self, analysis_result: AnalysisResult) -> Dict[str, str]:
        """可視化チャートの生成"""
        
        chart_paths = {}
        charts_data = analysis_result.charts_data
        
        try:
            # 累積リターンチャート
            if 'cumulative_returns' in charts_data:
                fig, ax = plt.subplots(figsize=(12, 6))
                
                cum_returns = pd.Series(charts_data['cumulative_returns'])
                cum_returns.plot(ax=ax, color=self.colors['primary'], linewidth=2)
                
                ax.set_title('累積リターン推移', fontsize=16, fontweight='bold')
                ax.set_xlabel('日付')
                ax.set_ylabel('累積リターン')
                ax.grid(True, alpha=0.3)
                
                chart_path = os.path.join(self.output_dir, "charts", f"cumulative_returns_{analysis_result.analysis_id}.png")
                plt.tight_layout()
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                
                chart_paths['cumulative_returns'] = os.path.relpath(chart_path, self.output_dir)
            
            # ドローダウンチャート
            if 'drawdown' in charts_data:
                fig, ax = plt.subplots(figsize=(12, 6))
                
                drawdown = pd.Series(charts_data['drawdown'])
                ax.fill_between(drawdown.index, drawdown, 0, alpha=0.7, color=self.colors['danger'])
                ax.plot(drawdown.index, drawdown, color='darkred', linewidth=1)
                
                ax.set_title('ドローダウン推移', fontsize=16, fontweight='bold')
                ax.set_xlabel('日付')
                ax.set_ylabel('ドローダウン (%)')
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1%}'))
                ax.grid(True, alpha=0.3)
                
                chart_path = os.path.join(self.output_dir, "charts", f"drawdown_{analysis_result.analysis_id}.png")
                plt.tight_layout()
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                
                chart_paths['drawdown'] = os.path.relpath(chart_path, self.output_dir)
            
            # リターン分布ヒストグラム
            if 'return_histogram' in charts_data:
                fig, ax = plt.subplots(figsize=(10, 6))
                
                hist_data = charts_data['return_histogram']
                bins = hist_data['bins']
                counts = hist_data['counts']
                
                ax.hist(bins[:-1], bins=bins, weights=counts, alpha=0.7, color=self.colors['info'])
                ax.axvline(0, color='red', linestyle='--', alpha=0.8, label='ゼロライン')
                
                ax.set_title('リターン分布', fontsize=16, fontweight='bold')
                ax.set_xlabel('日次リターン')
                ax.set_ylabel('頻度')
                ax.legend()
                ax.grid(True, alpha=0.3)
                
                chart_path = os.path.join(self.output_dir, "charts", f"return_distribution_{analysis_result.analysis_id}.png")
                plt.tight_layout()
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                
                chart_paths['return_distribution'] = os.path.relpath(chart_path, self.output_dir)
        
        except Exception as e:
            self.logger.warning(f"Chart generation failed: {e}")
        
        return chart_paths
    
    def _create_html_template(self, analysis_result: AnalysisResult, chart_paths: Dict[str, str]) -> str:
        """HTMLテンプレートの作成"""
        
        html = f"""
        <!DOCTYPE html>
        <html lang="ja">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>バックテスト結果レポート - {analysis_result.analysis_id}</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }}
                h1 {{ color: #333; border-bottom: 2px solid #007acc; padding-bottom: 10px; }}
                h2 {{ color: #007acc; margin-top: 30px; }}
                .metric {{ background-color: #f8f9fa; padding: 10px; margin: 5px 0; border-left: 4px solid #007acc; }}
                .warning {{ background-color: #fff3cd; border-left: 4px solid #ffc107; padding: 10px; margin: 10px 0; }}
                .recommendation {{ background-color: #d1ecf1; border-left: 4px solid #17a2b8; padding: 10px; margin: 10px 0; }}
                .chart {{ text-align: center; margin: 20px 0; }}
                .chart img {{ max-width: 100%; height: auto; border: 1px solid #ddd; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f8f9fa; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>バックテスト結果レポート</h1>
                
                <div class="metric">
                    <strong>分析ID:</strong> {analysis_result.analysis_id}
                </div>
                <div class="metric">
                    <strong>分析日時:</strong> {analysis_result.analysis_date.strftime('%Y-%m-%d %H:%M:%S')}
                </div>
                <div class="metric">
                    <strong>データ品質スコア:</strong> {analysis_result.data_quality_score:.2f}
                </div>
                
                <h2>主要パフォーマンス指標</h2>
        """
        
        # サマリーメトリクス
        for key, value in analysis_result.summary_metrics.items():
            if isinstance(value, (int, float)):
                if isinstance(value, float) and abs(value) < 1:
                    value_str = f"{value:.4f}"
                else:
                    value_str = f"{value:.2f}"
                html += f'<div class="metric"><strong>{key}:</strong> {value_str}</div>\n'
        
        # チャート
        if chart_paths:
            html += "<h2>パフォーマンスチャート</h2>\n"
            for chart_name, chart_path in chart_paths.items():
                html += f'''
                <div class="chart">
                    <h3>{chart_name.replace('_', ' ').title()}</h3>
                    <img src="{chart_path}" alt="{chart_name}">
                </div>
                '''
        
        # 推奨事項
        if analysis_result.recommendations:
            html += "<h2>推奨事項</h2>\n"
            for i, recommendation in enumerate(analysis_result.recommendations, 1):
                html += f'<div class="recommendation"><strong>{i}.</strong> {recommendation}</div>\n'
        
        # 警告
        if analysis_result.warnings:
            html += "<h2>注意事項</h2>\n"
            for i, warning in enumerate(analysis_result.warnings, 1):
                html += f'<div class="warning"><strong>警告 {i}:</strong> {warning}</div>\n'
        
        html += """
            </div>
        </body>
        </html>
        """
        
        return html
    
    # ヘルパーメソッド
    def _calculate_rolling_sharpe(self, returns: pd.Series, window: int) -> pd.Series:
        """ローリングシャープレシオの計算"""
        return returns.rolling(window).mean() / returns.rolling(window).std() * np.sqrt(252)
    
    def _calculate_rolling_volatility(self, returns: pd.Series, window: int) -> pd.Series:
        """ローリングボラティリティの計算"""
        return returns.rolling(window).std() * np.sqrt(252)
    
    def _calculate_tail_ratio(self, returns: pd.Series) -> float:
        """テールレシオの計算"""
        return abs(returns.quantile(0.95)) / abs(returns.quantile(0.05))
    
    def _calculate_drawdown_duration(self, drawdown: pd.Series) -> int:
        """ドローダウン期間の計算"""
        in_drawdown = drawdown < 0
        if not in_drawdown.any():
            return 0
        
        # 最長ドローダウン期間を計算
        drawdown_periods = []
        current_period = 0
        
        for is_dd in in_drawdown:
            if is_dd:
                current_period += 1
            else:
                if current_period > 0:
                    drawdown_periods.append(current_period)
                    current_period = 0
        
        if current_period > 0:
            drawdown_periods.append(current_period)
        
        return max(drawdown_periods) if drawdown_periods else 0
    
    def _calculate_recovery_time(self, drawdown: pd.Series) -> int:
        """回復時間の計算"""
        max_dd_idx = drawdown.idxmin()
        recovery_series = drawdown[max_dd_idx:]
        
        recovery_idx = recovery_series[recovery_series >= 0].first_valid_index()
        if recovery_idx is None:
            return len(recovery_series)  # まだ回復していない
        
        return recovery_series.index.get_loc(recovery_idx)
    
    def _extract_scenario_performance(self, scenario_results: Dict[str, Any]) -> Dict[str, float]:
        """シナリオパフォーマンスの抽出"""
        return {
            'return': scenario_results.get('total_return', 0.0),
            'sharpe': scenario_results.get('sharpe_ratio', 0.0),
            'max_dd': scenario_results.get('max_drawdown', 0.0)
        }
    
    def _check_success_criteria(self, scenario_results: Dict[str, Any], success_criteria: Dict[str, float]) -> bool:
        """成功基準のチェック"""
        for criterion, threshold in success_criteria.items():
            result_value = scenario_results.get(criterion, 0)
            
            if criterion in ['min_sharpe_ratio', 'min_win_rate']:
                if result_value < threshold:
                    return False
            elif criterion in ['max_drawdown']:
                if abs(result_value) > threshold:
                    return False
        
        return True
    
    def _calculate_data_quality_score(self, backtest_results: Dict[str, Any]) -> float:
        """データ品質スコアの計算"""
        score = 1.0
        
        # 基本的なデータの存在チェック
        required_fields = ['daily_returns', 'start_date', 'end_date']
        missing_fields = [field for field in required_fields if field not in backtest_results]
        
        if missing_fields:
            score -= 0.3 * len(missing_fields) / len(required_fields)
        
        # データの長さチェック
        if 'daily_returns' in backtest_results:
            returns_length = len(backtest_results['daily_returns'])
            if returns_length < 60:  # 2ヶ月未満
                score -= 0.4
            elif returns_length < 180:  # 6ヶ月未満
                score -= 0.2
        
        return max(0.0, score)
    
    def get_analysis_stats(self) -> Dict[str, Any]:
        """分析統計の取得"""
        return self.analysis_stats.copy()

# テスト関数
def test_backtest_result_analyzer():
    """テスト関数"""
    logger.info("Testing BacktestResultAnalyzer")
    
    # サンプルバックテスト結果
    np.random.seed(42)
    dates = pd.date_range(start='2023-01-01', end='2023-12-31', freq='D')
    returns = np.random.normal(0.001, 0.02, len(dates))
    
    backtest_results = {
        'daily_returns': returns.tolist(),
        'total_return': (1 + pd.Series(returns)).prod() - 1,
        'sharpe_ratio': np.mean(returns) / np.std(returns) * np.sqrt(252),
        'max_drawdown': -0.15,
        'win_rate': (pd.Series(returns) > 0).mean(),
        'start_date': dates[0],
        'end_date': dates[-1],
        'monthly_returns': pd.Series(returns).resample('M').apply(lambda x: (1 + x).prod() - 1).to_dict()
    }
    
    # 分析器の初期化
    analyzer = BacktestResultAnalyzer()
    
    # 分析実行
    analysis_result = analyzer.analyze_backtest_results(backtest_results)
    
    logger.info(f"Analysis ID: {analysis_result.analysis_id}")
    logger.info(f"Data quality score: {analysis_result.data_quality_score:.2f}")
    logger.info(f"Recommendations: {len(analysis_result.recommendations)}")
    logger.info(f"Warnings: {len(analysis_result.warnings)}")
    
    # Excelレポート生成
    excel_path = analyzer.generate_excel_report(analysis_result)
    if excel_path:
        logger.info(f"Excel report generated: {excel_path}")
    
    # HTML可視化レポート生成
    html_path = analyzer.generate_html_visualization(analysis_result)
    if html_path:
        logger.info(f"HTML visualization generated: {html_path}")
    
    # 統計表示
    stats = analyzer.get_analysis_stats()
    logger.info(f"Analysis stats: {stats}")
    
    return analysis_result

if __name__ == "__main__":
    # テスト実行
    result = test_backtest_result_analyzer()
    print(f"Test completed: Analysis ID {result.analysis_id}")
