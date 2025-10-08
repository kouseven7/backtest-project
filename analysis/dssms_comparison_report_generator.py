"""
DSSMS Phase 3 Task 3.2: 比較分析機能向上
Excel拡張レポート生成器
File: dssms_comparison_report_generator.py

DSSMS比較分析結果をExcel形式で出力する拡張レポート生成器
既存のSimpleExcelExporterと統合し、比較分析専用の詳細レポートを生成

Author: imega (Agent Mode Implementation)
Created: 2025-01-22
Based on: Previous conversation design specifications
"""

import os
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Union
from datetime import datetime
import logging
from pathlib import Path

# Excel操作
try:
    # openpyxl遅延インポート (TODO-PERF-001: Stage 3)
from src.utils.openpyxl_lazy_wrapper import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 既存システム統合
try:
    from output.simple_excel_exporter import SimpleExcelExporter
    from analysis.dssms_comparison_analysis_engine import ComparisonResult, AnalysisConfiguration
    HAS_CORE_MODULES = True
except ImportError as e:
    logging.warning(f"Core modules not available: {e}")
    HAS_CORE_MODULES = False

# ロギング設定
logger = logging.getLogger(__name__)

class DSSMSComparisonReportGenerator:
    """DSSMS比較分析Excel拡張レポート生成器"""
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        初期化
        
        Parameters:
            output_dir: 出力ディレクトリ
        """
        self.output_dir = output_dir or os.path.join(os.getcwd(), "output", "comparison_reports")
        self.ensure_output_directory()
        
        # 既存システム統合
        self.excel_exporter = None
        if HAS_CORE_MODULES:
            try:
                self.excel_exporter = SimpleExcelExporter()
                logger.info("SimpleExcelExporter統合完了")
            except Exception as e:
                logger.warning(f"SimpleExcelExporter統合失敗: {e}")
        
        # レポート履歴
        self.report_history = []
        
        # スタイル設定
        self._setup_excel_styles()
        
        logger.info(f"DSSMS比較レポート生成器初期化完了: {self.output_dir}")

    def ensure_output_directory(self):
        """出力ディレクトリ確保"""
        try:
            Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.error(f"出力ディレクトリ作成エラー: {e}")
            self.output_dir = os.getcwd()

    def _setup_excel_styles(self):
        """Excelスタイル設定"""
        if not HAS_OPENPYXL:
            self.styles = {}
            return
        
        self.styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=12),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'subheader': {
                'font': Font(bold=True, color='000000', size=11),
                'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'highlight_good': {
                'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            },
            'highlight_bad': {
                'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            },
            'number': {
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
        }

    def generate_comprehensive_report(
        self, 
        comparison_result: 'ComparisonResult',
        include_charts: bool = True,
        report_name: Optional[str] = None
    ) -> str:
        """
        包括的比較レポート生成
        
        Parameters:
            comparison_result: 比較分析結果
            include_charts: チャート含む
            report_name: レポート名
            
        Returns:
            生成されたファイルパス
        """
        try:
            # ファイル名生成
            if report_name is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                report_name = f"dssms_comparison_comprehensive_{timestamp}.xlsx"
            
            if not report_name.endswith('.xlsx'):
                report_name += '.xlsx'
            
            output_path = os.path.join(self.output_dir, report_name)
            
            logger.info(f"包括的比較レポート生成開始: {output_path}")
            
            if not HAS_OPENPYXL:
                return self._generate_simple_excel_report(comparison_result, output_path)
            
            # Excelワークブック作成
            wb = Workbook()
            
            # デフォルトシート削除
            wb.remove(wb.active)
            
            # シート生成
            self._create_executive_summary_sheet(wb, comparison_result)
            self._create_strategy_performance_sheet(wb, comparison_result)
            self._create_risk_analysis_sheet(wb, comparison_result)
            self._create_regime_analysis_sheet(wb, comparison_result)
            self._create_correlation_analysis_sheet(wb, comparison_result)
            self._create_recommendations_sheet(wb, comparison_result)
            self._create_detailed_data_sheet(wb, comparison_result)
            
            # チャート追加
            if include_charts:
                self._add_performance_charts(wb, comparison_result)
            
            # メタデータシート
            self._create_metadata_sheet(wb, comparison_result)
            
            # ファイル保存
            wb.save(output_path)
            
            # 履歴更新
            self.report_history.append({
                'timestamp': datetime.now(),
                'file_path': output_path,
                'analysis_id': comparison_result.analysis_id,
                'report_type': 'comprehensive'
            })
            
            logger.info(f"包括的比較レポート生成完了: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"包括的比較レポート生成エラー: {e}")
            import traceback
            traceback.print_exc()
            return ""

    def _generate_simple_excel_report(self, comparison_result: 'ComparisonResult', output_path: str) -> str:
        """シンプルExcelレポート生成（openpyxl非使用時）"""
        try:
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: # pandasのExcelWriterを使用
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # エグゼクティブサマリー
                summary_data = self._prepare_summary_data(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: summary_data.to_excel(writer, sheet_name='エグゼクティブサマリー', index=False)
                
                # 戦略パフォーマンス
                perf_data = self._prepare_performance_data(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: perf_data.to_excel(writer, sheet_name='戦略パフォーマンス', index=False)
                
                # 推奨事項
                rec_data = self._prepare_recommendations_data(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: rec_data.to_excel(writer, sheet_name='推奨事項', index=False)
                
                # メタデータ
                meta_data = self._prepare_metadata(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: meta_data.to_excel(writer, sheet_name='メタデータ', index=False)
            
            logger.info(f"シンプルExcelレポート生成完了: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"シンプルExcelレポート生成エラー: {e}")
            return ""

    def _create_executive_summary_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """エグゼクティブサマリーシート作成"""
        try:
            ws = wb.create_sheet("エグゼクティブサマリー")
            
            # タイトル
            ws['A1'] = 'DSSMS 戦略比較分析 - エグゼクティブサマリー'
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:F1')
            
            # 分析概要
            row = 3
            ws[f'A{row}'] = '分析概要'
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws[f'A{row}'].fill = self.styles['subheader']['fill']
            
            row += 1
            summary_items = [
                ('分析ID', result.analysis_id),
                ('分析日時', result.timestamp.strftime('%Y-%m-%d %H:%M:%S')),
                ('分析期間', f"{result.analysis_period[0].strftime('%Y-%m-%d')} ～ {result.analysis_period[1].strftime('%Y-%m-%d')}"),
                ('分析モード', result.analysis_mode),
                ('信頼度レベル', f"{result.confidence_level:.1%}"),
                ('データ品質スコア', f"{result.data_quality_score:.1%}")
            ]
            
            for item, value in summary_items:
                ws[f'A{row}'] = item
                ws[f'B{row}'] = value
                row += 1
            
            # 主要結果
            row += 2
            ws[f'A{row}'] = '主要結果'
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws[f'A{row}'].fill = self.styles['subheader']['fill']
            
            row += 1
            
            # 戦略ランキング表示
            if result.strategy_rankings:
                for metric, ranking in result.strategy_rankings.items():
                    if ranking:
                        ws[f'A{row}'] = f'{metric}ランキング'
                        ws[f'B{row}'] = f"1位: {ranking[0]}"
                        if len(ranking) > 1:
                            ws[f'C{row}'] = f"2位: {ranking[1]}"
                        if len(ranking) > 2:
                            ws[f'D{row}'] = f"3位: {ranking[2]}"
                        row += 1
            
            # 推奨事項サマリー
            row += 2
            ws[f'A{row}'] = '主要推奨事項'
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws[f'A{row}'].fill = self.styles['subheader']['fill']
            
            row += 1
            for i, recommendation in enumerate(result.overall_recommendations[:5], 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = recommendation
                row += 1
            
            # 列幅調整
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            
        except Exception as e:
            logger.error(f"エグゼクティブサマリーシート作成エラー: {e}")

    def _create_strategy_performance_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """戦略パフォーマンスシート作成"""
        try:
            ws = wb.create_sheet("戦略パフォーマンス")
            
            # タイトル
            ws['A1'] = '戦略パフォーマンス比較'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:H1')
            
            # パフォーマンステーブル
            if result.strategy_performance:
                row = 3
                
                # ヘッダー
                headers = ['戦略名', '総リターン', '年率リターン', 'ボラティリティ', 'シャープレシオ', '最大ドローダウン', 'Calmarレシオ']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles['header']['font']
                    cell.fill = self.styles['header']['fill']
                    cell.alignment = self.styles['header']['alignment']
                    cell.border = self.styles['header']['border']
                
                row += 1
                
                # データ行
                for strategy, metrics in result.strategy_performance.items():
                    ws.cell(row=row, column=1, value=strategy)
                    ws.cell(row=row, column=2, value=f"{metrics.get('total_return', 0):.2%}")
                    ws.cell(row=row, column=3, value=f"{metrics.get('annual_return', 0):.2%}")
                    ws.cell(row=row, column=4, value=f"{metrics.get('volatility', 0):.2%}")
                    ws.cell(row=row, column=5, value=f"{metrics.get('sharpe_ratio', 0):.3f}")
                    ws.cell(row=row, column=6, value=f"{metrics.get('max_drawdown', 0):.2%}")
                    ws.cell(row=row, column=7, value=f"{metrics.get('calmar_ratio', 0):.3f}")
                    
                    # 条件付き書式
                    sharpe_cell = ws.cell(row=row, column=5)
                    if metrics.get('sharpe_ratio', 0) > 1.0:
                        sharpe_cell.fill = self.styles['highlight_good']['fill']
                    elif metrics.get('sharpe_ratio', 0) < 0.5:
                        sharpe_cell.fill = self.styles['highlight_bad']['fill']
                    
                    row += 1
                
                # 統計サマリー
                row += 2
                ws[f'A{row}'] = 'パフォーマンス統計'
                ws[f'A{row}'].font = self.styles['subheader']['font']
                ws[f'A{row}'].fill = self.styles['subheader']['fill']
                
                row += 1
                if result.performance_differences:
                    for metric, value in result.performance_differences.items():
                        ws[f'A{row}'] = metric.replace('_', ' ').title()
                        ws[f'B{row}'] = f"{value:.4f}"
                        row += 1
            
            # 列幅調整
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 15
            
        except Exception as e:
            logger.error(f"戦略パフォーマンスシート作成エラー: {e}")

    def _create_risk_analysis_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """リスク分析シート作成"""
        try:
            ws = wb.create_sheet("リスク分析")
            
            # タイトル
            ws['A1'] = 'リスク分析'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            row = 3
            
            # リスク調整ランキング
            if result.risk_adjusted_rankings:
                ws[f'A{row}'] = 'リスク調整ランキング'
                ws[f'A{row}'].font = self.styles['subheader']['font']
                ws[f'A{row}'].fill = self.styles['subheader']['fill']
                
                row += 1
                for rank_type, strategies in result.risk_adjusted_rankings.items():
                    ws[f'A{row}'] = rank_type.replace('_', ' ').title()
                    for i, strategy in enumerate(strategies[:5], 1):
                        ws[f'B{row}'] = f"{i}. {strategy}"
                        row += 1
                    row += 1
            
            # リスク指標詳細
            if result.strategy_performance:
                row += 1
                ws[f'A{row}'] = 'リスク指標詳細'
                ws[f'A{row}'].font = self.styles['subheader']['font']
                ws[f'A{row}'].fill = self.styles['subheader']['fill']
                
                row += 1
                
                # ヘッダー
                headers = ['戦略名', 'VaR (95%)', 'スキューネス', 'クルトーシス', 'ダウンサイド偏差']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles['header']['font']
                    cell.fill = self.styles['header']['fill']
                
                row += 1
                
                # データ
                for strategy, metrics in result.strategy_performance.items():
                    ws.cell(row=row, column=1, value=strategy)
                    ws.cell(row=row, column=2, value=f"{metrics.get('var_95', 0):.2%}")
                    ws.cell(row=row, column=3, value=f"{metrics.get('skewness', 0):.3f}")
                    ws.cell(row=row, column=4, value=f"{metrics.get('kurtosis', 0):.3f}")
                    ws.cell(row=row, column=5, value=f"{metrics.get('downside_deviation', 0):.2%}")
                    row += 1
            
            # 列幅調整
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 18
            
        except Exception as e:
            logger.error(f"リスク分析シート作成エラー: {e}")

    def _create_regime_analysis_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """レジーム分析シート作成"""
        try:
            ws = wb.create_sheet("市場レジーム分析")
            
            # タイトル
            ws['A1'] = '市場レジーム分析'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            row = 3
            
            if result.regime_analysis:
                # レジーム パフォーマンス
                if 'regime_performance' in result.regime_analysis:
                    ws[f'A{row}'] = 'レジーム別パフォーマンス'
                    ws[f'A{row}'].font = self.styles['subheader']['font']
                    ws[f'A{row}'].fill = self.styles['subheader']['fill']
                    
                    row += 1
                    regime_perf = result.regime_analysis['regime_performance']
                    
                    for regime, metrics in regime_perf.items():
                        ws[f'A{row}'] = regime.replace('_', ' ').title()
                        if isinstance(metrics, dict):
                            ws[f'B{row}'] = f"発生回数: {metrics.get('count', 0)}"
                            ws[f'C{row}'] = f"平均リターン: {metrics.get('avg_return', 0):.2%}"
                        row += 1
                
                # レジーム遷移
                row += 2
                if result.regime_transitions:
                    ws[f'A{row}'] = 'レジーム遷移'
                    ws[f'A{row}'].font = self.styles['subheader']['font']
                    ws[f'A{row}'].fill = self.styles['subheader']['fill']
                    
                    row += 1
                    for i, transition in enumerate(result.regime_transitions[:10], 1):
                        ws[f'A{row}'] = f"遷移 {i}"
                        ws[f'B{row}'] = str(transition)
                        row += 1
            else:
                ws[f'A{row}'] = 'レジーム分析データがありません'
                ws[f'A{row}'].font = Font(italic=True)
            
            # 列幅調整
            for col in range(1, 4):
                ws.column_dimensions[chr(64 + col)].width = 20
            
        except Exception as e:
            logger.error(f"レジーム分析シート作成エラー: {e}")

    def _create_correlation_analysis_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """相関分析シート作成"""
        try:
            ws = wb.create_sheet("相関分析")
            
            # タイトル
            ws['A1'] = '戦略間相関分析'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            row = 3
            
            if result.correlation_analysis:
                # 相関行列
                if 'correlation_matrix' in result.correlation_analysis:
                    ws[f'A{row}'] = '相関行列'
                    ws[f'A{row}'].font = self.styles['subheader']['font']
                    ws[f'A{row}'].fill = self.styles['subheader']['fill']
                    
                    row += 1
                    corr_matrix = result.correlation_analysis['correlation_matrix']
                    
                    if isinstance(corr_matrix, dict):
                        strategies = list(corr_matrix.keys())
                        
                        # ヘッダー
                        ws.cell(row=row, column=1, value='戦略')
                        for col, strategy in enumerate(strategies, 2):
                            ws.cell(row=row, column=col, value=strategy)
                        
                        row += 1
                        
                        # データ
                        for i, strategy1 in enumerate(strategies):
                            ws.cell(row=row + i, column=1, value=strategy1)
                            for j, strategy2 in enumerate(strategies):
                                corr_value = corr_matrix.get(strategy1, {}).get(strategy2, 0)
                                cell = ws.cell(row=row + i, column=j + 2, value=f"{corr_value:.3f}")
                                
                                # 色付け
                                if abs(corr_value) > 0.7:
                                    cell.fill = self.styles['highlight_bad']['fill']
                                elif abs(corr_value) > 0.5:
                                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        
                        row += len(strategies) + 2
                
                # 高相関ペア
                if 'high_correlations' in result.correlation_analysis:
                    ws[f'A{row}'] = '高相関ペア (閾値: 70%以上)'
                    ws[f'A{row}'].font = self.styles['subheader']['font']
                    ws[f'A{row}'].fill = self.styles['subheader']['fill']
                    
                    row += 1
                    high_corrs = result.correlation_analysis['high_correlations']
                    
                    if high_corrs:
                        headers = ['戦略1', '戦略2', '相関係数']
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = self.styles['header']['font']
                            cell.fill = self.styles['header']['fill']
                        
                        row += 1
                        
                        for corr_pair in high_corrs:
                            ws.cell(row=row, column=1, value=corr_pair.get('strategy_1', ''))
                            ws.cell(row=row, column=2, value=corr_pair.get('strategy_2', ''))
                            ws.cell(row=row, column=3, value=f"{corr_pair.get('correlation', 0):.3f}")
                            row += 1
                    else:
                        ws[f'A{row}'] = '高相関ペアは検出されませんでした'
                        row += 1
                
                # 分散化スコア
                row += 2
                ws[f'A{row}'] = '分散化スコア'
                ws[f'A{row}'].font = self.styles['subheader']['font']
                ws[f'A{row}'].fill = self.styles['subheader']['fill']
                
                row += 1
                diversification_score = result.correlation_analysis.get('diversification_score', 0)
                ws[f'A{row}'] = '分散化スコア'
                ws[f'B{row}'] = f"{diversification_score:.1%}"
                
                # スコア評価
                row += 1
                if diversification_score > 0.7:
                    evaluation = "良好（戦略が十分に分散されています）"
                elif diversification_score > 0.5:
                    evaluation = "普通（一部戦略の相関が高い可能性があります）"
                else:
                    evaluation = "要改善（戦略間の相関が高すぎます）"
                
                ws[f'A{row}'] = '評価'
                ws[f'B{row}'] = evaluation
            else:
                ws[f'A{row}'] = '相関分析データがありません'
                ws[f'A{row}'].font = Font(italic=True)
            
            # 列幅調整
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 18
            
        except Exception as e:
            logger.error(f"相関分析シート作成エラー: {e}")

    def _create_recommendations_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """推奨事項シート作成"""
        try:
            ws = wb.create_sheet("推奨事項")
            
            # タイトル
            ws['A1'] = '推奨事項と最適化提案'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            row = 3
            
            # 全体推奨事項
            ws[f'A{row}'] = '全体推奨事項'
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws[f'A{row}'].fill = self.styles['subheader']['fill']
            
            row += 1
            for i, recommendation in enumerate(result.overall_recommendations, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = recommendation
                row += 1
            
            # 最適化提案
            row += 2
            ws[f'A{row}'] = '最適化提案'
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws[f'A{row}'].fill = self.styles['subheader']['fill']
            
            row += 1
            if result.optimization_suggestions:
                headers = ['優先度', 'タイプ', '詳細']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.styles['header']['font']
                    cell.fill = self.styles['header']['fill']
                
                row += 1
                
                for suggestion in result.optimization_suggestions:
                    ws.cell(row=row, column=1, value=suggestion.get('priority', 'medium').upper())
                    ws.cell(row=row, column=2, value=suggestion.get('type', '').replace('_', ' ').title())
                    ws.cell(row=row, column=3, value=suggestion.get('description', ''))
                    
                    # 優先度による色分け
                    priority_cell = ws.cell(row=row, column=1)
                    if suggestion.get('priority') == 'high':
                        priority_cell.fill = self.styles['highlight_bad']['fill']
                    elif suggestion.get('priority') == 'medium':
                        priority_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    
                    row += 1
            else:
                ws[f'A{row}'] = '最適化提案はありません'
                ws[f'A{row}'].font = Font(italic=True)
                row += 1
            
            # アクションアイテム
            row += 2
            ws[f'A{row}'] = 'アクションアイテム'
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws[f'A{row}'].fill = self.styles['subheader']['fill']
            
            row += 1
            action_items = [
                "定期的な戦略パフォーマンスモニタリング",
                "市場レジーム変化への対応準備",
                "リスク管理パラメータの見直し",
                "戦略間相関の継続監視",
                "最適化提案の実装検討"
            ]
            
            for i, item in enumerate(action_items, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = item
                row += 1
            
            # 列幅調整
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            
        except Exception as e:
            logger.error(f"推奨事項シート作成エラー: {e}")

    def _create_detailed_data_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """詳細データシート作成"""
        try:
            ws = wb.create_sheet("詳細データ")
            
            # タイトル
            ws['A1'] = '詳細分析データ'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            row = 3
            
            # 株式選択効果
            if result.stock_selection_effects:
                ws[f'A{row}'] = '株式選択効果分析'
                ws[f'A{row}'].font = self.styles['subheader']['font']
                ws[f'A{row}'].fill = self.styles['subheader']['fill']
                
                row += 1
                for strategy, effects in result.stock_selection_effects.items():
                    if isinstance(effects, dict) and 'strategy_selection_effects' in effects:
                        ws[f'A{row}'] = f'戦略: {strategy}'
                        row += 1
                        
                        selection_data = effects['strategy_selection_effects']
                        for effect_strategy, metrics in selection_data.items():
                            ws[f'B{row}'] = effect_strategy
                            if isinstance(metrics, dict):
                                for metric, value in metrics.items():
                                    ws[f'C{row}'] = f"{metric}: {value}"
                                    row += 1
                            row += 1
                        row += 1
            
            # セクター分析
            if result.sector_analysis:
                row += 1
                ws[f'A{row}'] = 'セクター分析'
                ws[f'A{row}'].font = self.styles['subheader']['font']
                ws[f'A{row}'].fill = self.styles['subheader']['fill']
                
                row += 1
                # セクター分析データを表示（簡略化）
                ws[f'A{row}'] = 'セクター分析データあり'
                row += 1
            
            # 列幅調整
            for col in range(1, 4):
                ws.column_dimensions[chr(64 + col)].width = 25
            
        except Exception as e:
            logger.error(f"詳細データシート作成エラー: {e}")

    def _create_metadata_sheet(self, wb: Workbook, result: 'ComparisonResult'):
        """メタデータシート作成"""
        try:
            ws = wb.create_sheet("メタデータ")
            
            # タイトル
            ws['A1'] = 'レポートメタデータ'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            row = 3
            
            metadata_items = [
                ('分析ID', result.analysis_id),
                ('生成日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                ('分析タイムスタンプ', result.timestamp.strftime('%Y-%m-%d %H:%M:%S')),
                ('分析期間開始', result.analysis_period[0].strftime('%Y-%m-%d')),
                ('分析期間終了', result.analysis_period[1].strftime('%Y-%m-%d')),
                ('分析モード', result.analysis_mode),
                ('信頼度レベル', f"{result.confidence_level:.1%}"),
                ('データ品質スコア', f"{result.data_quality_score:.1%}"),
                ('戦略数', len(result.strategy_performance)),
                ('推奨事項数', len(result.overall_recommendations)),
                ('最適化提案数', len(result.optimization_suggestions)),
                ('生成ツール', 'DSSMS比較分析エンジン v1.0'),
                ('レポート形式', 'Excel拡張レポート')
            ]
            
            for item, value in metadata_items:
                ws[f'A{row}'] = item
                ws[f'B{row}'] = value
                row += 1
            
            # 列幅調整
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            
        except Exception as e:
            logger.error(f"メタデータシート作成エラー: {e}")

    def _add_performance_charts(self, wb: Workbook, result: 'ComparisonResult'):
        """パフォーマンスチャート追加"""
        try:
            if not HAS_OPENPYXL:
                return
            
            # チャート用のシートを作成
            chart_ws = wb.create_sheet("パフォーマンスチャート")
            
            # 簡単なチャートデータ準備
            if result.strategy_performance:
                strategies = list(result.strategy_performance.keys())
                returns = [result.strategy_performance[s].get('annual_return', 0) for s in strategies]
                risks = [result.strategy_performance[s].get('volatility', 0) for s in strategies]
                
                # データをシートに書き込み
                chart_ws['A1'] = '戦略'
                chart_ws['B1'] = '年率リターン'
                chart_ws['C1'] = 'ボラティリティ'
                
                for i, (strategy, ret, risk) in enumerate(zip(strategies, returns, risks), 2):
                    chart_ws[f'A{i}'] = strategy
                    chart_ws[f'B{i}'] = ret
                    chart_ws[f'C{i}'] = risk
                
                # 棒グラフ作成（リターン）
                chart = BarChart()
                chart.title = "戦略別年率リターン"
                chart.y_axis.title = 'リターン (%)'
                chart.x_axis.title = '戦略'
                
                data = Reference(chart_ws, min_col=2, min_row=1, max_row=len(strategies)+1)
                cats = Reference(chart_ws, min_col=1, min_row=2, max_row=len(strategies)+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                chart_ws.add_chart(chart, "E2")
                
        except Exception as e:
            logger.error(f"パフォーマンスチャート追加エラー: {e}")

    def generate_quick_summary_report(
        self, 
        comparison_result: 'ComparisonResult',
        report_name: Optional[str] = None
    ) -> str:
        """
        クイックサマリーレポート生成
        
        Parameters:
            comparison_result: 比較分析結果
            report_name: レポート名
            
        Returns:
            生成されたファイルパス
        """
        try:
            # ファイル名生成
            if report_name is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                report_name = f"dssms_comparison_summary_{timestamp}.xlsx"
            
            if not report_name.endswith('.xlsx'):
                report_name += '.xlsx'
            
            output_path = os.path.join(self.output_dir, report_name)
            
            logger.info(f"クイックサマリーレポート生成開始: {output_path}")
            
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: # pandasのExcelWriterを使用
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # サマリーシート
                summary_data = self._prepare_summary_data(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: summary_data.to_excel(writer, sheet_name='サマリー', index=False)
                
                # 戦略ランキング
                ranking_data = self._prepare_ranking_data(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: ranking_data.to_excel(writer, sheet_name='戦略ランキング', index=False)
                
                # 推奨事項
                rec_data = self._prepare_recommendations_data(comparison_result)
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: rec_data.to_excel(writer, sheet_name='推奨事項', index=False)
            
            # 履歴更新
            self.report_history.append({
                'timestamp': datetime.now(),
                'file_path': output_path,
                'analysis_id': comparison_result.analysis_id,
                'report_type': 'quick_summary'
            })
            
            logger.info(f"クイックサマリーレポート生成完了: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"クイックサマリーレポート生成エラー: {e}")
            return ""

    def _prepare_summary_data(self, result: 'ComparisonResult') -> pd.DataFrame:
        """サマリーデータ準備"""
        try:
            data = []
            
            # 基本情報
            data.append(['分析ID', result.analysis_id])
            data.append(['分析日時', result.timestamp.strftime('%Y-%m-%d %H:%M:%S')])
            data.append(['分析期間', f"{result.analysis_period[0].strftime('%Y-%m-%d')} ～ {result.analysis_period[1].strftime('%Y-%m-%d')}"])
            data.append(['信頼度', f"{result.confidence_level:.1%}"])
            data.append(['データ品質', f"{result.data_quality_score:.1%}"])
            
            # トップ戦略
            if result.strategy_rankings and 'sharpe_ratio' in result.strategy_rankings:
                top_strategy = result.strategy_rankings['sharpe_ratio'][0]
                data.append(['最優秀戦略 (シャープレシオ)', top_strategy])
            
            return pd.DataFrame(data, columns=['項目', '値'])
            
        except Exception as e:
            logger.error(f"サマリーデータ準備エラー: {e}")
            return pd.DataFrame(columns=['項目', '値'])

    def _prepare_performance_data(self, result: 'ComparisonResult') -> pd.DataFrame:
        """パフォーマンスデータ準備"""
        try:
            data = []
            
            for strategy, metrics in result.strategy_performance.items():
                data.append([
                    strategy,
                    f"{metrics.get('total_return', 0):.2%}",
                    f"{metrics.get('annual_return', 0):.2%}",
                    f"{metrics.get('volatility', 0):.2%}",
                    f"{metrics.get('sharpe_ratio', 0):.3f}",
                    f"{metrics.get('max_drawdown', 0):.2%}"
                ])
            
            columns = ['戦略', '総リターン', '年率リターン', 'ボラティリティ', 'シャープレシオ', '最大ドローダウン']
            return pd.DataFrame(data, columns=columns)
            
        except Exception as e:
            logger.error(f"パフォーマンスデータ準備エラー: {e}")
            return pd.DataFrame()

    def _prepare_ranking_data(self, result: 'ComparisonResult') -> pd.DataFrame:
        """ランキングデータ準備"""
        try:
            data = []
            
            for metric, strategies in result.strategy_rankings.items():
                for rank, strategy in enumerate(strategies, 1):
                    data.append([metric.replace('_', ' ').title(), rank, strategy])
            
            return pd.DataFrame(data, columns=['指標', '順位', '戦略'])
            
        except Exception as e:
            logger.error(f"ランキングデータ準備エラー: {e}")
            return pd.DataFrame(columns=['指標', '順位', '戦略'])

    def _prepare_recommendations_data(self, result: 'ComparisonResult') -> pd.DataFrame:
        """推奨事項データ準備"""
        try:
            data = []
            
            for i, recommendation in enumerate(result.overall_recommendations, 1):
                data.append([i, recommendation])
            
            return pd.DataFrame(data, columns=['番号', '推奨事項'])
            
        except Exception as e:
            logger.error(f"推奨事項データ準備エラー: {e}")
            return pd.DataFrame(columns=['番号', '推奨事項'])

    def _prepare_metadata(self, result: 'ComparisonResult') -> pd.DataFrame:
        """メタデータ準備"""
        try:
            data = [
                ['分析ID', result.analysis_id],
                ['生成日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['分析モード', result.analysis_mode],
                ['信頼度レベル', f"{result.confidence_level:.1%}"],
                ['データ品質スコア', f"{result.data_quality_score:.1%}"],
                ['戦略数', len(result.strategy_performance)],
                ['レポート生成ツール', 'DSSMS比較レポート生成器']
            ]
            
            return pd.DataFrame(data, columns=['項目', '値'])
            
        except Exception as e:
            logger.error(f"メタデータ準備エラー: {e}")
            return pd.DataFrame(columns=['項目', '値'])

    def get_report_history(self) -> List[Dict[str, Any]]:
        """レポート履歴取得"""
        return self.report_history.copy()

    def cleanup_old_reports(self, days: int = 30):
        """古いレポートファイル削除"""
        try:
            cutoff_date = datetime.now() - timedelta(days=days)
            
            removed_count = 0
            for report in self.report_history[:]:
                if report['timestamp'] < cutoff_date:
                    try:
                        if os.path.exists(report['file_path']):
                            os.remove(report['file_path'])
                        self.report_history.remove(report)
                        removed_count += 1
                    except Exception as e:
                        logger.warning(f"レポートファイル削除失敗: {e}")
            
            logger.info(f"古いレポート {removed_count} 件を削除しました")
            
        except Exception as e:
            logger.error(f"古いレポート削除エラー: {e}")
