#!/usr/bin/env python3
"""
切り替え分析のログ出力と成功判定ロジックを詳細デバッグ
"""

import pandas as pd
import logging
import openpyxl
from pathlib import Path

# ログレベルを設定
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def analyze_latest_excel_debug():
    """最新のExcelファイルを取得してデバッグ分析"""
    
    # Excelファイルを検索
    excel_files = list(Path('backtest_results/dssms_results').glob('dssms_unified_backtest_*.xlsx'))
    if not excel_files:
        logger.error("Excelファイルが見つかりません")
        return
    
    latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
    logger.info(f"解析対象ファイル: {latest_file}")
    
    try:
        # Excelファイルを読み込み
        workbook = openpyxl.load_workbook(latest_file)
        
        if "切替分析" not in workbook.sheetnames:
            logger.error("切替分析シートが見つかりません")
            return
        
        ws = workbook["切替分析"]
        logger.info(f"切替分析シート読み込み完了: {ws.max_row}行, {ws.max_column}列")
        
        # データ行を解析（ヘッダーは1行目）
        success_count = 0
        fail_count = 0
        positive_performance_count = 0
        
        for row_idx in range(2, min(12, ws.max_row + 1)):  # 最初の10行をデバッグ
            # セルの値を取得
            date_val = ws[f"A{row_idx}"].value
            performance_val = ws[f"G{row_idx}"].value
            success_val = ws[f"H{row_idx}"].value
            
            # パフォーマンス値の詳細解析
            performance_numeric = None
            if performance_val is not None:
                try:
                    if isinstance(performance_val, str):
                        if '%' in performance_val:
                            performance_str = performance_val.replace('%', '').replace(',', '').strip()
                            performance_numeric = float(performance_str)
                        else:
                            performance_numeric = float(performance_val)
                    else:
                        performance_numeric = float(performance_val)
                except (ValueError, TypeError):
                    performance_numeric = 0.0
            
            # 判定ロジックのテスト
            expected_success = "成功" if performance_numeric and performance_numeric > 0 else "失敗"
            
            # カウント
            if success_val == "成功":
                success_count += 1
            else:
                fail_count += 1
            
            if performance_numeric and performance_numeric > 0:
                positive_performance_count += 1
            
            # デバッグ出力
            logger.info(f"Row {row_idx}: Date={date_val}, Performance='{performance_val}' ({type(performance_val)}), "
                       f"Numeric={performance_numeric:.4f}, Success='{success_val}', Expected='{expected_success}', "
                       f"Match={'✓' if success_val == expected_success else '✗'}")
        
        logger.info(f"集計結果: 成功={success_count}, 失敗={fail_count}, 正のパフォーマンス={positive_performance_count}")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"エラーが発生しました: {e}")

def simulate_excel_generation():
    """Excel生成時の処理をシミュレート"""
    logger.info("Excel生成処理のシミュレーション開始")
    
    # テストデータ
    test_switches = [
        {"performance_after": "13.52%", "expected": "成功"},
        {"performance_after": "-8.94%", "expected": "失敗"},
        {"performance_after": "3.61%", "expected": "成功"},
        {"performance_after": 0.1352, "expected": "成功"},
        {"performance_after": -0.0894, "expected": "失敗"},
    ]
    
    for i, switch in enumerate(test_switches, 1):
        performance = switch["performance_after"]
        logger.info(f"=== テストケース {i} ===")
        logger.info(f"入力パフォーマンス: '{performance}' ({type(performance)})")
        
        try:
            # 実際のロジックを再現
            if isinstance(performance, str) and '%' in performance:
                performance_str = performance.replace('%', '').replace(',', '').strip()
                performance_val = float(performance_str)
                logger.info(f"文字列変換: '{performance}' -> '{performance_str}' -> {performance_val}")
            else:
                performance_val = float(performance) if performance is not None else 0.0
                logger.info(f"数値変換: {performance} -> {performance_val}")
            
            is_successful = performance_val > 0
            success_status = "成功" if is_successful else "失敗"
            
            logger.info(f"成功判定: {performance_val} > 0 = {is_successful} -> '{success_status}'")
            logger.info(f"期待値: '{switch['expected']}', 実際: '{success_status}', "
                       f"一致: {'✓' if success_status == switch['expected'] else '✗'}")
            
        except (ValueError, TypeError) as e:
            logger.error(f"変換エラー: {e}")

if __name__ == "__main__":
    logger.info("切り替え分析デバッグツール開始")
    
    print("1. Excel生成ロジックのシミュレーション")
    simulate_excel_generation()
    
    print("\n2. 実際のExcelファイルの解析")
    analyze_latest_excel_debug()
