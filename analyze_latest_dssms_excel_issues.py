#!/usr/bin/env python3
"""
DSSMS最新Excelファイル問題分析ツール
作成日: 2025-09-08

目的:
- 最新のDSSMSバックテストExcelファイルの戦略統計シート問題を詳細分析
- 保有期間24時間固定問題の復活有無を確認
- 戦略名がDSSMSのみになっている問題を検証
"""

import pandas as pd
import json
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def analyze_latest_dssms_excel():
    """最新のDSSMSExcelファイルを詳細分析"""
    print("=" * 80)
    print("DSSMS最新Excelファイル問題分析")
    print("=" * 80)
    
    # 最新ファイルを特定
    results_dir = Path("backtest_results/dssms_results")
    excel_files = list(results_dir.glob("dssms_unified_backtest_*.xlsx"))
    
    if not excel_files:
        print("❌ DSSMSバックテストExcelファイルが見つかりません")
        return None
    
    # 最新ファイルを取得（ファイル名のタイムスタンプでソート）
    latest_file = sorted(excel_files, key=lambda x: x.name)[-1]
    print(f"📄 分析対象ファイル: {latest_file}")
    
    try:
        # Excelファイルを読み込み
        workbook = load_workbook(latest_file)
        print(f"📋 利用可能シート: {workbook.sheetnames}")
        
        analysis_results = {}
        
        # 1. 戦略統計シートの分析
        print("\n🔍 戦略統計シート分析:")
        if "戦略統計" in workbook.sheetnames:
            strategy_sheet = workbook["戦略統計"]
            
            # シートの内容を読み取り
            strategy_data = []
            for row in strategy_sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    strategy_data.append(row)
            
            print(f"   📊 戦略統計シート行数: {len(strategy_data)}")
            if strategy_data:
                print("   📝 先頭5行:")
                for i, row in enumerate(strategy_data[:5]):
                    print(f"      {i+1}: {row}")
            
            # 戦略名のユニーク値をチェック
            if len(strategy_data) > 1:  # ヘッダー行を除く
                strategy_names = [row[0] for row in strategy_data[1:] if row and row[0]]
                unique_strategies = set(strategy_names)
                print(f"   🏷️  戦略名一覧: {list(unique_strategies)}")
                print(f"   📈 ユニーク戦略数: {len(unique_strategies)}")
                
                # 期待される7つの戦略があるかチェック
                expected_strategies = [
                    "VWAPBreakoutStrategy", "MeanReversionStrategy", "TrendFollowingStrategy",
                    "MomentumStrategy", "ContrarianStrategy", "VolatilityBreakoutStrategy", "RSIStrategy"
                ]
                
                missing_strategies = set(expected_strategies) - unique_strategies
                if missing_strategies:
                    print(f"   ⚠️  不足戦略: {list(missing_strategies)}")
                
                # DSSMSのみの問題をチェック
                if unique_strategies == {"DSSMSStrategy"}:
                    print("   ❌ 問題確認: 戦略名が「DSSMSStrategy」のみになっています")
                elif len(unique_strategies) == 7:
                    print("   ✅ 戦略名は正常（7つの個別戦略）")
                
            analysis_results["strategy_sheet"] = {
                "exists": True,
                "row_count": len(strategy_data),
                "strategy_names": list(unique_strategies) if len(strategy_data) > 1 else [],
                "data_sample": strategy_data[:3] if strategy_data else []
            }
        else:
            print("   ❌ 戦略統計シートが見つかりません")
            analysis_results["strategy_sheet"] = {"exists": False}
        
        # 2. 取引履歴シートの分析（保有期間問題チェック）
        print("\n🔍 取引履歴シート分析:")
        if "取引履歴" in workbook.sheetnames:
            trade_sheet = workbook["取引履歴"]
            
            # シートの内容を読み取り
            trade_data = []
            for row in trade_sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    trade_data.append(row)
            
            print(f"   📊 取引履歴シート行数: {len(trade_data)}")
            
            if len(trade_data) > 1:  # ヘッダー行を除く
                # ヘッダー行を確認
                if trade_data:
                    headers = trade_data[0]
                    print(f"   📝 ヘッダー: {headers}")
                    
                    # 保有期間列のインデックスを探す
                    holding_period_index = None
                    for i, header in enumerate(headers):
                        if header and "保有期間" in str(header):
                            holding_period_index = i
                            break
                    
                    if holding_period_index is not None:
                        # 保有期間データを分析
                        holding_periods = []
                        for row in trade_data[1:6]:  # 最初の5行をサンプル分析
                            if row and len(row) > holding_period_index:
                                holding_periods.append(row[holding_period_index])
                        
                        print(f"   ⏱️  保有期間サンプル: {holding_periods}")
                        
                        # 24時間固定問題をチェック
                        unique_periods = set(str(p) for p in holding_periods if p is not None)
                        if len(unique_periods) == 1 and "24.0" in unique_periods:
                            print("   ❌ 問題確認: 保有期間が24時間固定になっています")
                        elif len(unique_periods) > 1:
                            print("   ✅ 保有期間は正常（多様な値）")
                        else:
                            print(f"   ⚠️  保有期間の状況: {unique_periods}")
            
            analysis_results["trade_sheet"] = {
                "exists": True,
                "row_count": len(trade_data),
                "sample_data": trade_data[:3] if trade_data else []
            }
        else:
            print("   ❌ 取引履歴シートが見つかりません")
            analysis_results["trade_sheet"] = {"exists": False}
        
        # 3. 対応するJSONファイルも分析
        print("\n🔍 対応JSONファイル分析:")
        json_file = latest_file.with_suffix('').name.replace('dssms_unified_backtest_', 'dssms_unified_data_') + '.json'
        json_path = results_dir / json_file
        
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # 戦略統計データを確認
            if "strategy_statistics" in json_data:
                strategy_stats = json_data["strategy_statistics"]
                print(f"   📊 JSON戦略統計キー数: {len(strategy_stats)}")
                print(f"   🏷️  JSON戦略名: {list(strategy_stats.keys())}")
                
                # JSONの戦略データが正しいかチェック
                if len(strategy_stats) == 7:
                    print("   ✅ JSON戦略データは正常（7戦略）")
                elif len(strategy_stats) == 1 and "DSSMSStrategy" in strategy_stats:
                    print("   ❌ JSON戦略データも単一戦略になっています")
                
                analysis_results["json_strategy_data"] = {
                    "strategy_count": len(strategy_stats),
                    "strategy_names": list(strategy_stats.keys()),
                    "sample_stats": {k: v for k, v in list(strategy_stats.items())[:2]}
                }
            
            # 取引履歴も確認
            if "trades" in json_data:
                trades = json_data["trades"]
                print(f"   📈 JSON取引数: {len(trades)}")
                if trades:
                    sample_trade = trades[0]
                    print(f"   📝 サンプル取引: {sample_trade}")
                
                analysis_results["json_trades"] = {
                    "trade_count": len(trades),
                    "sample_trade": trades[0] if trades else None
                }
        else:
            print(f"   ❌ 対応JSONファイルが見つかりません: {json_path}")
            analysis_results["json_data"] = {"exists": False}
        
        # 4. 結果サマリー
        print("\n📋 分析結果サマリー:")
        issues_found = []
        
        if analysis_results.get("strategy_sheet", {}).get("exists"):
            strategy_names = analysis_results["strategy_sheet"].get("strategy_names", [])
            if strategy_names == ["DSSMSStrategy"]:
                issues_found.append("戦略統計シートに7つの戦略ではなくDSSMSStrategyのみ表示")
            elif len(strategy_names) != 7:
                issues_found.append(f"戦略数が不正: {len(strategy_names)}個（期待値: 7個）")
        
        if issues_found:
            print("   ❌ 発見された問題:")
            for issue in issues_found:
                print(f"      - {issue}")
        else:
            print("   ✅ 特定の問題は検出されませんでした")
        
        return analysis_results, latest_file
        
    except Exception as e:
        print(f"❌ 分析中にエラーが発生: {e}")
        return None, latest_file

def main():
    """メイン実行関数"""
    analysis_results, latest_file = analyze_latest_dssms_excel()
    
    if analysis_results:
        print(f"\n📄 詳細分析完了: {latest_file}")
        print("💡 次のステップ:")
        print("   1. 統一出力エンジンの戦略統計生成ロジックを確認")
        print("   2. DSSMSバックテスター内の戦略データ処理を検証")
        print("   3. 必要に応じて修正を実行")
    
    return analysis_results

if __name__ == "__main__":
    main()
