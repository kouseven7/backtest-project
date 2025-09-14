#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Task 6.3: Excel出力内容の正確性検証
DSSMS統一出力エンジンが生成するExcelファイルの数学的正確性・完全性を検証

目的:
1. 最新Excel（dssms_unified_backtest_20250910_213413.xlsx）の詳細分析
2. 各シート（取引履歴・損益推移・戦略統計・切替分析）の数値検証
3. Problem 1-14で特定された問題がExcel内容に反映されているか確認
4. 手計算による数値検証（勝率・保有期間・切替数等）
5. 85.0点エンジンによる出力データの信頼性評価

Author: GitHub Copilot Agent
Created: 2025-09-14
Task: 6.3 Excel出力内容の正確性検証
"""

import os
import sys
import json
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
import warnings

# 警告を抑制
warnings.filterwarnings('ignore')

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class Task63ExcelAccuracyVerifier:
    """Task 6.3: Excel出力内容の正確性検証システム"""
    
    def __init__(self):
        self.project_root = Path(__file__).parent
        self.results = {
            "task_6_3_results": {
                "timestamp": datetime.now().strftime("%Y%m%d_%H%M%S"),
                "verification_summary": {},
                "sheet_analysis": {},
                "numerical_verification": {},
                "problem_correlation": {},
                "engine_reliability": {},
                "issues_found": [],
                "recommendations": []
            }
        }
        
    def find_latest_excel(self) -> Optional[Path]:
        """最新のDSSMS Excelファイルを特定"""
        excel_patterns = [
            "dssms_unified_backtest_*.xlsx",
            "DSSMS_*.xlsx", 
            "*dssms*.xlsx"
        ]
        
        latest_file = None
        latest_time = None
        
        # プロジェクトルートとサブディレクトリを検索
        search_paths = [
            self.project_root,
            self.project_root / "backtest_results",
            self.project_root / "backtest_results" / "dssms_results",
            self.project_root / "output",
            self.project_root / "output" / "backtest_results"
        ]
        
        for search_path in search_paths:
            if search_path.exists():
                for pattern in excel_patterns:
                    for excel_file in search_path.glob(pattern):
                        file_time = excel_file.stat().st_mtime
                        if latest_time is None or file_time > latest_time:
                            latest_time = file_time
                            latest_file = excel_file
                    
        return latest_file
    
    def analyze_sheet_structure(self, excel_path: Path) -> Dict[str, Any]:
        """Excelファイルのシート構造分析"""
        logger.info(f"シート構造分析開始: {excel_path.name}")
        
        structure_analysis = {}
        
        try:
            workbook = load_workbook(excel_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # シートの基本情報
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # ヘッダー行の取得
                headers = []
                if max_row > 0:
                    for col in range(1, min(max_col + 1, 20)):  # 最大20列まで
                        cell_value = sheet.cell(1, col).value
                        if cell_value:
                            headers.append(str(cell_value))
                
                # データ行数の計算
                data_rows = max_row - 1 if max_row > 1 else 0
                
                structure_analysis[sheet_name] = {
                    "max_row": max_row,
                    "max_col": max_col,
                    "data_rows": data_rows,
                    "headers": headers,
                    "has_data": data_rows > 0
                }
                
                logger.info(f"  {sheet_name}: {data_rows}行, {max_col}列, ヘッダー: {len(headers)}個")
                
        except Exception as e:
            logger.error(f"シート構造分析エラー: {e}")
            structure_analysis["error"] = str(e)
            
        return structure_analysis
    
    def verify_trade_history_sheet(self, excel_path: Path) -> Dict[str, Any]:
        """取引履歴シートの数値検証"""
        logger.info("取引履歴シートの数値検証開始")
        
        verification = {
            "sheet_found": False,
            "total_trades": 0,
            "strategy_distribution": {},
            "holding_period_analysis": {},
            "price_analysis": {},
            "pnl_analysis": {},
            "issues": []
        }
        
        try:
            # 複数の可能なシート名をチェック
            possible_names = ["取引履歴", "Trade History", "TradeHistory", "Trades"]
            workbook = load_workbook(excel_path, data_only=True)
            
            trade_sheet = None
            for name in possible_names:
                if name in workbook.sheetnames:
                    trade_sheet = workbook[name]
                    verification["sheet_found"] = True
                    break
            
            if not trade_sheet:
                verification["issues"].append("取引履歴シートが見つからない")
                return verification
            
            # データの読み取り
            trades_data = []
            headers = []
            
            # ヘッダー行の取得
            for col in range(1, trade_sheet.max_column + 1):
                header = trade_sheet.cell(1, col).value
                if header:
                    headers.append(str(header))
                else:
                    break
            
            # データ行の読み取り
            for row in range(2, trade_sheet.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = trade_sheet.cell(row, col).value
                    row_data[header] = cell_value
                trades_data.append(row_data)
            
            verification["total_trades"] = len(trades_data)
            logger.info(f"総取引数: {len(trades_data)}")
            
            # 戦略分布の分析
            strategy_counts = {}
            holding_periods = []
            entry_prices = []
            exit_prices = []
            pnl_values = []
            
            for trade in trades_data:
                # 戦略名の分析
                strategy = trade.get("戦略名", trade.get("Strategy", "Unknown"))
                if strategy:
                    strategy_counts[strategy] = strategy_counts.get(strategy, 0) + 1
                
                # 保有期間の分析
                holding_period = trade.get("保有期間", trade.get("Holding Period"))
                if holding_period and isinstance(holding_period, (int, float)):
                    holding_periods.append(holding_period)
                
                # 価格データの分析
                entry_price = trade.get("エントリー価格", trade.get("Entry Price"))
                exit_price = trade.get("エグジット価格", trade.get("Exit Price"))
                
                if entry_price and isinstance(entry_price, (int, float)):
                    entry_prices.append(entry_price)
                if exit_price and isinstance(exit_price, (int, float)):
                    exit_prices.append(exit_price)
                
                # 損益の分析
                pnl = trade.get("損益", trade.get("PnL", trade.get("Profit/Loss")))
                if pnl and isinstance(pnl, (int, float)):
                    pnl_values.append(pnl)
            
            verification["strategy_distribution"] = strategy_counts
            
            # 保有期間分析
            if holding_periods:
                verification["holding_period_analysis"] = {
                    "count": len(holding_periods),
                    "average": np.mean(holding_periods),
                    "min": min(holding_periods),
                    "max": max(holding_periods),
                    "unique_values": len(set(holding_periods)),
                    "all_24_hours": all(h == 24 for h in holding_periods if h)
                }
                
                # 24時間固定問題のチェック
                if verification["holding_period_analysis"]["all_24_hours"]:
                    verification["issues"].append("保有期間が24時間で固定されている（Problem 3関連）")
            
            # 価格分析
            if entry_prices:
                unique_entry_prices = len(set(entry_prices))
                verification["price_analysis"]["entry_prices"] = {
                    "count": len(entry_prices),
                    "unique_count": unique_entry_prices,
                    "all_same": unique_entry_prices == 1,
                    "average": np.mean(entry_prices)
                }
                
                if unique_entry_prices == 1:
                    verification["issues"].append(f"エントリー価格が固定値({entry_prices[0]})（Problem 2関連）")
            
            if exit_prices:
                unique_exit_prices = len(set(exit_prices))
                verification["price_analysis"]["exit_prices"] = {
                    "count": len(exit_prices),
                    "unique_count": unique_exit_prices,
                    "all_same": unique_exit_prices == 1,
                    "average": np.mean(exit_prices)
                }
                
                if unique_exit_prices == 1:
                    verification["issues"].append(f"エグジット価格が固定値({exit_prices[0]})（Problem 2関連）")
            
            # 損益分析
            if pnl_values:
                verification["pnl_analysis"] = {
                    "total_pnl": sum(pnl_values),
                    "average_pnl": np.mean(pnl_values),
                    "positive_trades": len([p for p in pnl_values if p > 0]),
                    "negative_trades": len([p for p in pnl_values if p < 0]),
                    "win_rate": len([p for p in pnl_values if p > 0]) / len(pnl_values) if pnl_values else 0
                }
            
            # 戦略名問題のチェック
            if len(strategy_counts) == 1 and "DSSMS" in list(strategy_counts.keys())[0]:
                verification["issues"].append("戦略名がDSSMS系のみで個別戦略が区別されていない（Problem 1関連）")
            
        except Exception as e:
            logger.error(f"取引履歴シート検証エラー: {e}")
            verification["issues"].append(f"検証エラー: {e}")
            
        return verification
    
    def verify_strategy_statistics_sheet(self, excel_path: Path) -> Dict[str, Any]:
        """戦略統計シートの数値検証"""
        logger.info("戦略統計シートの数値検証開始")
        
        verification = {
            "sheet_found": False,
            "strategy_count": 0,
            "strategies_listed": [],
            "metrics_analysis": {},
            "issues": []
        }
        
        try:
            possible_names = ["戦略統計", "Strategy Statistics", "StrategyStats", "Strategy"]
            workbook = load_workbook(excel_path, data_only=True)
            
            stats_sheet = None
            for name in possible_names:
                if name in workbook.sheetnames:
                    stats_sheet = workbook[name]
                    verification["sheet_found"] = True
                    break
            
            if not stats_sheet:
                verification["issues"].append("戦略統計シートが見つからない")
                return verification
            
            # 戦略名の収集
            strategies = []
            for row in range(2, stats_sheet.max_row + 1):
                strategy_name = stats_sheet.cell(row, 1).value
                if strategy_name:
                    strategies.append(str(strategy_name))
            
            verification["strategy_count"] = len(strategies)
            verification["strategies_listed"] = strategies
            
            # 期待される7戦略との比較
            expected_strategies = [
                "VWAPBreakoutStrategy", "GoldenCrossStrategy", "BollingerBandStrategy",
                "MACDStrategy", "RSIStrategy", "MeanReversionStrategy", "MomentumStrategy"
            ]
            
            if len(strategies) < 7:
                verification["issues"].append(f"戦略数が不足（期待: 7, 実際: {len(strategies)}）（Problem 4関連）")
            
            # DSSMSのみの問題チェック
            if len(strategies) == 1 and "DSSMS" in strategies[0]:
                verification["issues"].append("戦略統計がDSSMSのみで個別戦略統計がない（Problem 4関連）")
            
        except Exception as e:
            logger.error(f"戦略統計シート検証エラー: {e}")
            verification["issues"].append(f"検証エラー: {e}")
            
        return verification
    
    def verify_switch_analysis_sheet(self, excel_path: Path) -> Dict[str, Any]:
        """切替分析シートの数値検証"""
        logger.info("切替分析シートの数値検証開始")
        
        verification = {
            "sheet_found": False,
            "switch_count": 0,
            "success_rate_analysis": {},
            "frequency_analysis": {},
            "issues": []
        }
        
        try:
            possible_names = ["切替分析", "Switch Analysis", "SwitchAnalysis", "切り替え分析"]
            workbook = load_workbook(excel_path, data_only=True)
            
            switch_sheet = None
            for name in possible_names:
                if name in workbook.sheetnames:
                    switch_sheet = workbook[name]
                    verification["sheet_found"] = True
                    break
            
            if not switch_sheet:
                verification["issues"].append("切替分析シートが見つからない")
                return verification
            
            # 切替データの読み取り
            switch_data = []
            for row in range(2, switch_sheet.max_row + 1):
                date = switch_sheet.cell(row, 1).value
                from_symbol = switch_sheet.cell(row, 2).value
                to_symbol = switch_sheet.cell(row, 3).value
                
                if date and from_symbol and to_symbol:
                    switch_data.append({
                        "date": date,
                        "from": from_symbol,
                        "to": to_symbol
                    })
            
            verification["switch_count"] = len(switch_data)
            
            # 切替頻度の分析
            if len(switch_data) < 10:  # 期待される切替数と比較
                verification["issues"].append(f"切替数が極端に少ない（{len(switch_data)}回）（Problem 5関連）")
            
        except Exception as e:
            logger.error(f"切替分析シート検証エラー: {e}")
            verification["issues"].append(f"検証エラー: {e}")
            
        return verification
    
    def correlate_with_known_problems(self, all_issues: List[str]) -> Dict[str, Any]:
        """発見された問題とProblem 1-14との相関分析"""
        
        problem_mapping = {
            "Problem 1": "戦略名がDSSMS系のみ",
            "Problem 2": "価格が固定値",
            "Problem 3": "保有期間が24時間固定",
            "Problem 4": "戦略統計が個別戦略で生成されていない",
            "Problem 5": "切替数が激減",
            "Problem 6": "勝率計算の不正確性",
            "Problem 7": "統計計算未実施",
            "Problem 8": "Excel出力サイズ問題"
        }
        
        correlation = {}
        
        for issue in all_issues:
            for problem_id, description in problem_mapping.items():
                if any(keyword in issue for keyword in description.split()):
                    if problem_id not in correlation:
                        correlation[problem_id] = []
                    correlation[problem_id].append(issue)
        
        return correlation
    
    def evaluate_engine_reliability(self, verification_results: Dict[str, Any]) -> Dict[str, Any]:
        """85.0点エンジンの出力データ信頼性評価"""
        
        total_issues = len(verification_results.get("issues_found", []))
        
        reliability_score = max(0, 100 - (total_issues * 10))  # 問題1つにつき-10点
        
        reliability_assessment = {
            "total_issues_found": total_issues,
            "reliability_score": reliability_score,
            "engine_quality_score": 85.0,  # Task 4.2結果
            "quality_vs_output_mismatch": reliability_score < 70,  # 70点未満で不整合
            "assessment": ""
        }
        
        if reliability_score >= 80:
            reliability_assessment["assessment"] = "高信頼性: 85.0点エンジンの品質が出力に反映されている"
        elif reliability_score >= 60:
            reliability_assessment["assessment"] = "中信頼性: 一部問題があるが許容範囲"
        else:
            reliability_assessment["assessment"] = "低信頼性: 85.0点エンジンの品質が出力に反映されていない（重大な不整合）"
        
        return reliability_assessment
    
    def run_verification(self) -> Dict[str, Any]:
        """Task 6.3の完全検証実行"""
        logger.info("Task 6.3: Excel出力内容の正確性検証開始")
        
        # 1. 最新Excelファイルの特定
        latest_excel = self.find_latest_excel()
        
        if not latest_excel:
            logger.error("検証対象のExcelファイルが見つからない")
            self.results["task_6_3_results"]["verification_summary"]["status"] = "FAILED"
            self.results["task_6_3_results"]["verification_summary"]["error"] = "Excelファイルが見つからない"
            return self.results
        
        logger.info(f"検証対象ファイル: {latest_excel.name}")
        self.results["task_6_3_results"]["verification_summary"]["target_file"] = latest_excel.name
        self.results["task_6_3_results"]["verification_summary"]["file_size"] = latest_excel.stat().st_size
        
        # 2. シート構造分析
        structure_analysis = self.analyze_sheet_structure(latest_excel)
        self.results["task_6_3_results"]["sheet_analysis"] = structure_analysis
        
        # 3. 各シートの数値検証
        trade_verification = self.verify_trade_history_sheet(latest_excel)
        strategy_verification = self.verify_strategy_statistics_sheet(latest_excel)
        switch_verification = self.verify_switch_analysis_sheet(latest_excel)
        
        self.results["task_6_3_results"]["numerical_verification"] = {
            "trade_history": trade_verification,
            "strategy_statistics": strategy_verification,
            "switch_analysis": switch_verification
        }
        
        # 4. 全問題の集約
        all_issues = []
        all_issues.extend(trade_verification.get("issues", []))
        all_issues.extend(strategy_verification.get("issues", []))
        all_issues.extend(switch_verification.get("issues", []))
        
        self.results["task_6_3_results"]["issues_found"] = all_issues
        
        # 5. Problem 1-14との相関分析
        problem_correlation = self.correlate_with_known_problems(all_issues)
        self.results["task_6_3_results"]["problem_correlation"] = problem_correlation
        
        # 6. エンジン信頼性評価
        engine_reliability = self.evaluate_engine_reliability(self.results["task_6_3_results"])
        self.results["task_6_3_results"]["engine_reliability"] = engine_reliability
        
        # 7. 推奨事項の生成
        recommendations = []
        
        if len(all_issues) > 5:
            recommendations.append("Critical: 85.0点エンジンにも関わらず重大な出力問題が多数存在")
            recommendations.append("エンジン品質スコアと実際の出力品質に大きな乖離がある")
        
        if any("Problem 1" in str(problem_correlation.get(p, [])) for p in problem_correlation):
            recommendations.append("戦略名表示ロジックの修正が必要")
        
        if any("Problem 3" in str(problem_correlation.get(p, [])) for p in problem_correlation):
            recommendations.append("保有期間計算ロジックの修正が必要")
        
        self.results["task_6_3_results"]["recommendations"] = recommendations
        
        # 8. 検証サマリー
        self.results["task_6_3_results"]["verification_summary"]["status"] = "COMPLETED"
        self.results["task_6_3_results"]["verification_summary"]["total_issues"] = len(all_issues)
        self.results["task_6_3_results"]["verification_summary"]["reliability_score"] = engine_reliability["reliability_score"]
        
        return self.results
    
    def save_results(self):
        """検証結果の保存"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.project_root / f"task_6_3_results_{timestamp}.json"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, ensure_ascii=False, indent=2, default=str)
        
        logger.info(f"Task 6.3検証結果保存: {output_file}")
        return output_file

def main():
    """Task 6.3メイン実行"""
    print("="*80)
    print("Task 6.3: Excel出力内容の正確性検証")
    print("="*80)
    
    verifier = Task63ExcelAccuracyVerifier()
    
    # 検証実行
    results = verifier.run_verification()
    
    # 結果保存
    output_file = verifier.save_results()
    
    # サマリー表示
    summary = results["task_6_3_results"]["verification_summary"]
    reliability = results["task_6_3_results"]["engine_reliability"]
    issues = results["task_6_3_results"]["issues_found"]
    
    print(f"\n【検証結果サマリー】")
    print(f"対象ファイル: {summary.get('target_file', 'N/A')}")
    print(f"ファイルサイズ: {summary.get('file_size', 0):,} bytes")
    print(f"検証ステータス: {summary.get('status', 'UNKNOWN')}")
    print(f"発見された問題数: {summary.get('total_issues', 0)}")
    print(f"エンジン信頼性スコア: {reliability.get('reliability_score', 0)}/100")
    print(f"品質評価: {reliability.get('assessment', 'N/A')}")
    
    if issues:
        print(f"\n【発見された問題】")
        for i, issue in enumerate(issues, 1):
            print(f"{i:2d}. {issue}")
    
    recommendations = results["task_6_3_results"]["recommendations"]
    if recommendations:
        print(f"\n【推奨事項】")
        for i, rec in enumerate(recommendations, 1):
            print(f"{i:2d}. {rec}")
    
    print(f"\n詳細結果: {output_file}")
    print("="*80)
    
    return output_file

if __name__ == "__main__":
    main()