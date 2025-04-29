"""
Module: Excel Result Exporter
File: excel_result_exporter.py
Description:
    バックテスト結果をExcelファイルに出力するための関数群を提供します。
    - Workbookの存在確認および新規作成
    - 複数シートへのDataFrame出力（タイムゾーン付き日時の処理含む）
    - 損益推移グラフの追加
    - 取引履歴からピボットテーブルの作成

Author: imega
Created: 2023-04-01
Modified: 2025-04-14

Dependencies:
  - os
  - logging
  - pandas
  - openpyxl
  - datetime
  - config.logger_config
"""

import os
import logging
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict
from datetime import datetime

import sys

# プロジェクトのルートディレクトリを `sys.path` に追加
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from config.logger_config import setup_logger

logger = setup_logger(__name__)

def ensure_workbook_exists(workbook_path: str) -> None:
    """
    指定したパスにExcelファイルが存在しない場合、必要なシートとヘッダーを含む新規Workbookを作成して保存します。
    """
    output_dir = os.path.dirname(workbook_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.info(f"出力ディレクトリを作成しました: {output_dir}")
    
    if not os.path.exists(workbook_path):
        logger.info(f"Workbookが存在しません。新規に作成します: {workbook_path}")
        wb = Workbook()
        ws_pnl = wb.active
        ws_pnl.title = "損益推移"
        ws_pnl.append(["日付", "日次損益", "累積損益"])
        
        ws_trade = wb.create_sheet(title="取引履歴")
        ws_trade.append(["日付", "銘柄", "戦略", "エントリー", "イグジット", "取引結果", "取引量", "手数料"])
        
        ws_metrics = wb.create_sheet(title="パフォーマンス指標")
        ws_metrics.append(["指標", "値"])
        ws_metrics.append(["総取引数", "0"])
        ws_metrics.append(["勝率", "0%"])
        ws_metrics.append(["損益合計", "0円"])
        ws_metrics.append(["平均損益", "0円"])
        ws_metrics.append(["最大利益", "0円"])
        ws_metrics.append(["最大損失", "0円"])
        ws_metrics.append(["最大ドローダウン(金額)", "0円"])
        
        try:
            wb.save(workbook_path)
            logger.info(f"新規Workbookを作成し、保存しました: {workbook_path}")
        except Exception as e:
            logger.error(f"Workbookの保存中にエラーが発生しました: {e}")
            raise
    else:
        logger.info(f"Workbookは既に存在します: {workbook_path}")
        # 既存のWorkbookに必要なシートが含まれているか確認
        wb = openpyxl.load_workbook(workbook_path)
        required_sheets = ["損益推移", "取引履歴", "パフォーマンス指標"]
        
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                logger.info(f"シート '{sheet_name}' が存在しないため追加します")
                ws = wb.create_sheet(title=sheet_name)
                
                if sheet_name == "損益推移":
                    ws.append(["日付", "日次損益", "累積損益"])
                elif sheet_name == "取引履歴":
                    ws.append(["日付", "銘柄", "戦略", "エントリー", "イグジット", "取引結果", "取引量", "手数料"])
                elif sheet_name == "パフォーマンス指標":
                    ws.append(["指標", "値"])
                    ws.append(["総取引数", "0"])
                    ws.append(["勝率", "0%"])
                    ws.append(["損益合計", "0円"])
                    ws.append(["平均損益", "0円"])
                    ws.append(["最大利益", "0円"])
                    ws.append(["最大損失", "0円"])
                    ws.append(["最大ドローダウン(金額)", "0円"])
        
        try:
            wb.save(workbook_path)
        except Exception as e:
            logger.error(f"Workbookの更新中にエラーが発生しました: {e}")
            raise

def _remove_timezone_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrame内の各セルに対して、もしタイムゾーン付きdatetimeであればタイムゾーン情報を除去します.
    
    Returns:
        pd.DataFrame: タイムゾーン情報を除去したDataFrame
    """
    def remove_tz(x):
        try:
            # datetime型でtzinfoが存在する場合、tz_localize(None)を適用
            if hasattr(x, 'tzinfo') and x.tzinfo is not None:
                return x.tz_localize(None)
        except Exception:
            pass
        return x
    return df.applymap(remove_tz)

def save_backtest_results(trade_results: Dict[str, pd.DataFrame], output_file: str) -> None:
    """
    バックテスト結果をExcelファイルの各シートに保存します
    Parameters:
        trade_results (Dict[str, pd.DataFrame]): 各種結果データフレームの辞書
        output_file (str): 結果を保存するExcelファイルのパス
    """
    # Excelファイルが存在するか確認し、存在しない場合は作成
    ensure_workbook_exists(output_file)
    
    try:
        # ExcelWriterを作成し、各シートにデータを書き込む
        # mode="a" は追加モード、if_sheet_exists="replace" は既存のシートを置き換え
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            for sheet_name, df in trade_results.items():
                if isinstance(df, pd.DataFrame):
                    # データフレームのインデックスからタイムゾーン情報を削除
                    processed_df = df.copy()
                    
                    # インデックスがDatetimeIndexの場合、タイムゾーン情報を削除
                    if isinstance(processed_df.index, pd.DatetimeIndex) and processed_df.index.tz is not None:
                        processed_df.index = processed_df.index.tz_localize(None)
                        logger.info(f"シート '{sheet_name}' のインデックスからタイムゾーン情報を削除しました")
                    
                    # データフレーム内の日時データからもタイムゾーン情報を削除
                    for col in processed_df.columns:
                        if isinstance(processed_df[col].dtype, pd.DatetimeTZDtype):
                            processed_df[col] = processed_df[col].dt.tz_localize(None)
                            logger.info(f"シート '{sheet_name}' の列 '{col}' からタイムゾーン情報を削除しました")
                    
                    # 処理されたデータフレームをExcelに書き込む
                    processed_df.to_excel(writer, sheet_name=sheet_name)
                    logger.info(f"シート '{sheet_name}' にデータを書き込みました")
    except PermissionError:
        logger.error(f"ファイル {output_file} が開かれているため、書き込みができません。")
        
        # CSV形式でバックアップ
        base_dir = os.path.dirname(output_file)
        base_name = os.path.basename(output_file).replace(".xlsx", "")
        backup_dir = os.path.join(base_dir, "backup")
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        for sheet_name, df in trade_results.items():
            if isinstance(df, pd.DataFrame):
                backup_file = os.path.join(backup_dir, f"{base_name}_{sheet_name}_{timestamp}.csv")
                df.to_csv(backup_file)
                logger.info(f"バックアップとして {backup_file} にデータを保存しました")
        
        raise PermissionError(f"ファイル {output_file} が開かれているため、書き込みができません。バックアップとしてCSVファイルを保存しました。")
    except Exception as e:
        logger.error(f"データの保存中にエラーが発生しました: {e}")
        raise

def add_pnl_chart(workbook_path: str, sheet_name: str = "損益推移", chart_title: str = "累積損益推移") -> None:
    """
    指定したExcelファイルの指定シートに、累積損益の折れ線グラフを追加します.
    シート内は、A列に「日付」、B列に「累積損益」が存在する前提です.
    
    Parameters:
        workbook_path (str): Excelファイルのパス
        sheet_name (str): グラフを追加するシート名（デフォルトは「損益推移」）
        chart_title (str): グラフのタイトル
    """
    try:
        wb = openpyxl.load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            logger.warning(f"シート '{sheet_name}' が存在しません。グラフは作成されません。")
            return
            
        ws = wb[sheet_name]
        max_row = ws.max_row
        
        if max_row <= 1:  # ヘッダーだけの場合
            logger.warning(f"シート '{sheet_name}' にデータがありません。グラフは作成されません。")
            return
            
        chart = LineChart()
        chart.title = chart_title
        chart.style = 10
        chart.y_axis.title = '累積損益'
        chart.x_axis.title = '日付'
        
        data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D2")
        wb.save(workbook_path)
        logger.info(f"チャート '{chart_title}' をシート '{sheet_name}' に追加し、保存しました。")
    except Exception as e:
        logger.exception(f"チャートの追加中にエラーが発生しました: {e}")
        raise

def create_pivot_from_trade_history(
    workbook_path: str, 
    trade_sheet: str = "取引履歴", 
    pivot_sheet: str = "Pivot_取引履歴"
) -> None:
    """
    指定したExcelファイルの取引履歴シートからデータを読み込み、
    銘柄ごとの取引数と取引結果合計を算出したピボットテーブルを新規シートに出力します.
    
    Parameters:
        workbook_path (str): Excelファイルのパス
        trade_sheet (str): 取引履歴が格納されているシート名（デフォルトは「取引履歴」）
        pivot_sheet (str): ピボットテーブルを出力するシート名（デフォルトは「Pivot_取引履歴」）
    """
    try:
        # まず対象ファイルとシートの存在を確認
        if not os.path.exists(workbook_path):
            logger.warning(f"ファイル '{workbook_path}' が存在しません。ピボットテーブルは作成されません。")
            return
            
        wb = openpyxl.load_workbook(workbook_path)
        if trade_sheet not in wb.sheetnames:
            logger.warning(f"シート '{trade_sheet}' が存在しません。ピボットテーブルは作成されません。")
            return
            
        df = pd.read_excel(workbook_path, sheet_name=trade_sheet)
        
        # 取引履歴にデータがあるか確認
        if len(df) <= 0:
            logger.warning(f"シート '{trade_sheet}' にデータがありません。ピボットテーブルは作成されません。")
            return
            
        # 必要なカラムが存在することを確認
        required_columns = ["銘柄", "取引結果"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.warning(f"シート '{trade_sheet}' に必要なカラム {missing_columns} がありません。ピボットテーブルは作成されません。")
            return
            
        pivot = df.pivot_table(index="銘柄", aggfunc={"取引結果": ["count", "sum"]})
        pivot.columns = ["取引数", "取引結果合計"]
        pivot.reset_index(inplace=True)
        
        if pivot_sheet in wb.sheetnames:
            ws = wb[pivot_sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(title=pivot_sheet)
        
        for r_idx, row in enumerate(dataframe_to_rows(pivot, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(workbook_path)
        logger.info(f"ピボットテーブルをシート '{pivot_sheet}' に作成し、保存しました。")
    except Exception as e:
        logger.exception(f"ピボットテーブルの作成中にエラーが発生しました: {e}")
        raise

def save_performance_metrics_to_excel(performance_metrics: pd.DataFrame, output_file: str):
    """
    パフォーマンス指標をExcelファイルに保存します。

    Parameters:
        performance_metrics (pd.DataFrame): パフォーマンス指標を含むデータフレーム。
        output_file (str): 結果を保存するExcelファイルのパス。
    """
    from openpyxl import load_workbook

    try:
        # 既存のExcelファイルを開く
        wb = load_workbook(output_file)

        # パフォーマンス指標シートを作成または上書き
        if "パフォーマンス指標" in wb.sheetnames:
            ws = wb["パフォーマンス指標"]
            # 既存の内容をクリア
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(title="パフォーマンス指標")

        # パフォーマンス指標を書き込む
        for r_idx, row in enumerate(performance_metrics.itertuples(index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 保存
        wb.save(output_file)
        wb.close()
        logger.info(f"パフォーマンス指標をExcelに保存しました: {output_file}")

    except Exception as e:
        logger.error(f"パフォーマンス指標の保存中にエラーが発生しました: {e}")
        raise

def simulate_and_save(result_data: pd.DataFrame, ticker: str):
    """
    バックテストシミュレーションを実行し、結果をExcelに出力します。
    """
    import trade_simulation as trade_simulation
    trade_results = trade_simulation.simulate_trades(result_data, ticker)

    # パフォーマンス指標を追加
    trade_results = add_performance_metrics(trade_results)

    logger.info("バックテスト（トレードシミュレーション）完了")

    output_dir = r"C:\\Users\\imega\\Documents\\my_backtest_project\\backtest_results"
    # ディレクトリが存在しない場合は作成
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.info(f"出力ディレクトリを作成しました: {output_dir}")

    # 実行日時をファイル名に含める
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"backtest_results_{now}.xlsx")

    try:
        from output.excel_result_exporter import ensure_workbook_exists, add_pnl_chart, create_pivot_from_trade_history, save_backtest_results

        # 新しいExcelファイルを作成
        ensure_workbook_exists(output_file)

        # 結果を保存
        save_backtest_results(trade_results, output_file)
        logger.info(f"バックテスト結果をExcelに出力完了: {output_file}")

        # パフォーマンス指標を保存
        save_performance_metrics_to_excel(trade_results["パフォーマンス指標"], output_file)

        # チャートとピボットテーブルを追加
        add_pnl_chart(output_file, sheet_name="損益推移", chart_title="累積損益推移")
        create_pivot_from_trade_history(output_file, trade_sheet="取引履歴", pivot_sheet="Pivot_取引履歴")
        logger.info("Excelのチャート、ピボットテーブル追加完了")

    except PermissionError as e:
        logger.warning(f"Excelファイルにアクセスできません: {e}")
        # アクセスできない場合はCSVに出力
        csv_dir = os.path.join(output_dir, "csv_backup")
        os.makedirs(csv_dir, exist_ok=True)

        for sheet_name, df in trade_results.items():
            if isinstance(df, pd.DataFrame):
                csv_file = os.path.join(csv_dir, f"backtest_{sheet_name}_{now}.csv")
                df.to_csv(csv_file)
                logger.info(f"代替として結果をCSVに出力しました: {csv_file}")

        # 元のExcelファイルを別名で保存してみる
        alt_output_file = os.path.join(output_dir, f"backtest_results_alt_{now}.xlsx")
        try:
            ensure_workbook_exists(alt_output_file)
            save_backtest_results(trade_results, alt_output_file)
            logger.info(f"代替Excelファイルに結果を出力しました: {alt_output_file}")
            output_file = alt_output_file  # 成功した場合は新しいファイル名を使用
        except Exception as ex:
            logger.error(f"代替Excelファイルの作成にも失敗しました: {ex}")
            # CSVファイルだけで十分とする

    return trade_results

def run_trade_simulation():
    """
    トレードシミュレーションを実行するメイン関数。
    """
    try:
        # 1. 戦略パラメータの取得（例: GC戦略）
        param_manager = StrategyParameterManager(config_file)
        gc_params = param_manager.get_params("GC戦略")
        logger.info("GC戦略のパラメータを取得しました。")

        # 2. 市場データの取得（例: 銘柄設定シートから取得）
        stock_params = pd.read_excel(config_file, sheet_name="銘柄設定")
        ticker = stock_params["銘柄"].iloc[0]
        start_date = stock_params["開始日"].iloc[0]
        end_date = stock_params["終了日"].iloc[0]
        logger.info(f"市場データの取得対象: {ticker} {start_date}〜{end_date}")
        
        # 市場データの取得
        stock_data = fetch_yahoo_data(ticker, start_date, end_date)
        
        # 3. 戦略クラスのインスタンス化とシグナル生成
        strategy = GCStrategy(stock_data, gc_params, price_column="Adj Close")
        # generate_signals() メソッドがなく、代わりに backtest() メソッドを使用
        result_data = strategy.backtest()
        logger.info("シグナルの生成が完了しました。")
        
        # 4. トレードシミュレーションの実行
        trade_results = simulate_trades(result_data, ticker)
        
        # 5. 結果の保存
        output_file = r"C:\Users\imega\Documents\my_backtest_project\backtest_results\backtest_results.xlsx"
        save_backtest_results(trade_results, output_file)
        logger.info("トレードシミュレーションの結果を保存しました。")
    
    except Exception as e:
        logger.exception("トレードシミュレーションの実行中にエラーが発生しました。")
        raise
