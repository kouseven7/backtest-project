#!/usr/bin/env python3
"""
Excelファイルの詳細なセル検証スクリプト
セルの値とフォーマットを詳しく確認します
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime
import logging

# ログ設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def main():
    """メイン実行関数"""
    
    # 最新のファイルを見つける
    dssms_dir = 'backtest_results/dssms_results'
    excel_files = [f for f in os.listdir(dssms_dir) if f.startswith('dssms_unified_backtest_') and f.endswith('.xlsx')]
    
    if not excel_files:
        logger.error("Excelファイルが見つかりません")
        return
    
    latest_file = max(excel_files, key=lambda x: os.path.getctime(os.path.join(dssms_dir, x)))
    file_path = os.path.join(dssms_dir, latest_file)
    
    logger.info(f"分析対象ファイル: {file_path}")
    
    # openpyxlで詳細分析
    analyze_excel_cells(file_path)

def analyze_excel_cells(file_path):
    """Excelセルの詳細分析"""
    
    try:
        # ワークブックを開く
        workbook = openpyxl.load_workbook(file_path)
        
        # 切替分析シートを取得
        if '切替分析' not in workbook.sheetnames:
            logger.error("切替分析シートが見つかりません")
            logger.info(f"利用可能シート: {workbook.sheetnames}")
            return
            
        worksheet = workbook['切替分析']
        
        logger.info("=== セル詳細分析 ===")
        logger.info(f"シート名: {worksheet.title}")
        logger.info(f"最大行数: {worksheet.max_row}")
        logger.info(f"最大列数: {worksheet.max_column}")
        
        # ヘッダー行の確認
        headers = []
        for col in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col)
            headers.append(header_cell.value)
        
        logger.info(f"ヘッダー: {headers}")
        
        # パフォーマンス列と成功判定列のインデックスを探す
        perf_col = None
        success_col = None
        
        for i, header in enumerate(headers):
            if 'パフォーマンス' in str(header):
                perf_col = i + 1
            elif '成功' in str(header) or '判定' in str(header):
                success_col = i + 1
        
        logger.info(f"パフォーマンス列: {perf_col}, 成功判定列: {success_col}")
        
        if perf_col is None or success_col is None:
            logger.error("必要な列が見つかりません")
            return
        
        # 最初の10行を詳しく分析
        logger.info("\\n=== セル詳細（最初の10データ行） ===")
        
        for row in range(2, min(12, worksheet.max_row + 1)):  # 2行目から（ヘッダー除く）
            perf_cell = worksheet.cell(row=row, column=perf_col)
            success_cell = worksheet.cell(row=row, column=success_col)
            
            logger.info(f"Row {row}:")
            logger.info(f"  パフォーマンス - 値: {perf_cell.value}, 型: {type(perf_cell.value)}, フォーマット: {perf_cell.number_format}")
            logger.info(f"  成功判定 - 値: {success_cell.value}, 型: {type(success_cell.value)}")
            
            # 数値変換テスト
            try:
                if isinstance(perf_cell.value, str):
                    # %記号を除去して変換
                    clean_value = perf_cell.value.replace('%', '').replace(',', '')
                    numeric_value = float(clean_value)
                    expected_success = "成功" if numeric_value > 0 else "失敗"
                    logger.info(f"  計算結果: {clean_value} -> {numeric_value} -> 期待値: {expected_success}")
                elif isinstance(perf_cell.value, (int, float)):
                    expected_success = "成功" if perf_cell.value > 0 else "失敗"
                    logger.info(f"  計算結果: {perf_cell.value} -> 期待値: {expected_success}")
                else:
                    logger.info(f"  計算結果: 変換不可能な型")
                    
            except Exception as e:
                logger.info(f"  計算結果: 変換エラー - {e}")
            
            logger.info("")
        
        # 成功・失敗の分布を確認
        success_count = 0
        failure_count = 0
        total_rows = worksheet.max_row - 1  # ヘッダー除く
        
        logger.info("=== 全体分布確認 ===")
        
        for row in range(2, worksheet.max_row + 1):
            success_cell = worksheet.cell(row=row, column=success_col)
            if success_cell.value == "成功":
                success_count += 1
            elif success_cell.value == "失敗":
                failure_count += 1
        
        logger.info(f"総行数（データ）: {total_rows}")
        logger.info(f"成功: {success_count}件 ({success_count/total_rows*100:.1f}%)")
        logger.info(f"失敗: {failure_count}件 ({failure_count/total_rows*100:.1f}%)")
        logger.info(f"その他: {total_rows - success_count - failure_count}件")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Excelファイル分析エラー: {e}")
        import traceback
        logger.error(traceback.format_exc())

if __name__ == "__main__":
    main()
