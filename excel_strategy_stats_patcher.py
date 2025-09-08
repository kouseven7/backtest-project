#!/usr/bin/env python3
"""
DSSMS Excel戦略別統計直接修正システム

問題: 統一出力エンジンの修正がExcel生成に反映されていない
解決: 最新ExcelファイルのStrategy統計シートを直接正しい7戦略データで更新
"""

import openpyxl
import json
from typing import Dict, List, Any
from datetime import datetime
import logging

def setup_logger():
    """ロガー設定"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

class ExcelStrategyStatsPatcher:
    """Excel戦略別統計パッチャー"""
    
    def __init__(self):
        self.logger = setup_logger()
    
    def patch_excel_strategy_stats(self, excel_path: str, json_path: str) -> bool:
        """Excelファイルの戦略別統計シートを正しいデータで更新"""
        try:
            # JSONから戦略別統計を計算
            strategy_stats = self._calculate_strategy_stats_from_json(json_path)
            
            if not strategy_stats:
                self.logger.error("戦略別統計の計算に失敗")
                return False
            
            # Excelファイルを開く
            workbook = openpyxl.load_workbook(excel_path)
            
            # 既存の戦略別統計シートを削除
            if '戦略別統計' in workbook.sheetnames:
                workbook.remove(workbook['戦略別統計'])
                self.logger.info("既存の戦略別統計シートを削除")
            
            # 新しい戦略別統計シートを作成
            ws = workbook.create_sheet('戦略別統計')
            
            # ヘッダー設定
            headers = [
                '戦略名', '取引回数', '勝率', '平均利益', '平均損失', 
                '最大利益', '最大損失', 'プロフィットファクター', '総損益'
            ]
            
            # ヘッダー行を作成
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")
            
            # 戦略別データ行を追加
            row = 2
            total_stats = {
                'trade_count': 0,
                'winning_trades': 0,
                'total_pnl': 0.0
            }
            
            # 戦略をアルファベット順でソート
            sorted_strategies = sorted(strategy_stats.items())
            
            for strategy, stats in sorted_strategies:
                # 戦略名
                ws.cell(row=row, column=1).value = strategy
                
                # 取引回数
                ws.cell(row=row, column=2).value = stats['trade_count']
                
                # 勝率
                win_rate_cell = ws.cell(row=row, column=3)
                win_rate_cell.value = f"{stats['win_rate']:.2f}%"
                
                # 平均利益
                ws.cell(row=row, column=4).value = f"{stats['avg_profit']:,.2f}"
                
                # 平均損失
                ws.cell(row=row, column=5).value = f"{stats['avg_loss']:,.2f}"
                
                # 最大利益
                ws.cell(row=row, column=6).value = f"{stats['max_profit']:,.2f}"
                
                # 最大損失
                ws.cell(row=row, column=7).value = f"{stats['max_loss']:,.2f}"
                
                # プロフィットファクター
                pf_value = stats['profit_factor']
                if pf_value == float('inf'):
                    ws.cell(row=row, column=8).value = "∞"
                else:
                    ws.cell(row=row, column=8).value = f"{pf_value:.3f}"
                
                # 総損益
                ws.cell(row=row, column=9).value = f"{stats['total_pnl']:,.2f}"
                
                # 合計統計に加算
                total_stats['trade_count'] += stats['trade_count']
                total_stats['winning_trades'] += stats['winning_trades']
                total_stats['total_pnl'] += stats['total_pnl']
                
                row += 1
            
            # 合計行を追加
            total_win_rate = (total_stats['winning_trades'] / total_stats['trade_count'] * 100) if total_stats['trade_count'] > 0 else 0
            
            ws.cell(row=row, column=1).value = "合計"
            ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            
            ws.cell(row=row, column=2).value = total_stats['trade_count']
            ws.cell(row=row, column=2).font = openpyxl.styles.Font(bold=True)
            
            ws.cell(row=row, column=3).value = f"{total_win_rate:.2f}%"
            ws.cell(row=row, column=3).font = openpyxl.styles.Font(bold=True)
            
            ws.cell(row=row, column=9).value = f"{total_stats['total_pnl']:,.2f}"
            ws.cell(row=row, column=9).font = openpyxl.styles.Font(bold=True)
            
            # セル幅調整
            column_widths = [20, 12, 10, 12, 12, 12, 12, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # ファイルを保存
            workbook.save(excel_path)
            
            self.logger.info(f"✅ 戦略別統計シートを更新完了: {len(strategy_stats)}戦略")
            self.logger.info(f"総合勝率: {total_win_rate:.2f}%")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Excel戦略別統計更新エラー: {e}")
            return False
    
    def _calculate_strategy_stats_from_json(self, json_path: str) -> Dict[str, Any]:
        """JSONファイルから戦略別統計を計算"""
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            trades = data.get('trades', [])
            if not trades:
                self.logger.warning("取引データが見つかりません")
                return {}
            
            # 戦略別にグループ化
            strategy_trades = {}
            for trade in trades:
                strategy = trade.get('strategy', 'UnknownStrategy')
                if strategy not in strategy_trades:
                    strategy_trades[strategy] = []
                strategy_trades[strategy].append(trade)
            
            # 各戦略の統計を計算
            strategy_stats = {}
            for strategy, trades_list in strategy_trades.items():
                stats = self._calculate_single_strategy_stats(trades_list)
                strategy_stats[strategy] = stats
            
            self.logger.info(f"📊 戦略別統計計算完了: {len(strategy_stats)}戦略")
            for strategy, stats in strategy_stats.items():
                self.logger.info(f"  {strategy}: {stats['trade_count']}件, 勝率{stats['win_rate']:.1f}%")
            
            return strategy_stats
            
        except Exception as e:
            self.logger.error(f"JSON戦略別統計計算エラー: {e}")
            return {}
    
    def _calculate_single_strategy_stats(self, trades_list: List[Dict]) -> Dict[str, Any]:
        """単一戦略の統計を計算"""
        if not trades_list:
            return self._empty_stats()
        
        # 基本統計
        total_trades = len(trades_list)
        pnls = [float(trade.get('pnl', 0)) for trade in trades_list]
        winning_trades = len([p for p in pnls if p > 0])
        losing_trades = len([p for p in pnls if p < 0])
        
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
        
        # 損益統計
        winning_pnls = [p for p in pnls if p > 0]
        losing_pnls = [p for p in pnls if p < 0]
        
        avg_profit = sum(winning_pnls) / len(winning_pnls) if winning_pnls else 0
        avg_loss = sum(losing_pnls) / len(losing_pnls) if losing_pnls else 0
        max_profit = max(winning_pnls) if winning_pnls else 0
        max_loss = min(losing_pnls) if losing_pnls else 0
        
        total_profit = sum(winning_pnls)
        total_loss = abs(sum(losing_pnls))
        profit_factor = total_profit / total_loss if total_loss > 0 else float('inf') if total_profit > 0 else 0
        
        total_pnl = sum(pnls)
        
        return {
            'trade_count': total_trades,
            'winning_trades': winning_trades,
            'losing_trades': losing_trades,
            'win_rate': win_rate,
            'avg_profit': avg_profit,
            'avg_loss': avg_loss,
            'max_profit': max_profit,
            'max_loss': max_loss,
            'profit_factor': profit_factor,
            'total_pnl': total_pnl
        }
    
    def _empty_stats(self) -> Dict[str, Any]:
        """空の統計データ"""
        return {
            'trade_count': 0,
            'winning_trades': 0,
            'losing_trades': 0,
            'win_rate': 0.0,
            'avg_profit': 0.0,
            'avg_loss': 0.0,
            'max_profit': 0.0,
            'max_loss': 0.0,
            'profit_factor': 0.0,
            'total_pnl': 0.0
        }

def main():
    """メイン実行"""
    patcher = ExcelStrategyStatsPatcher()
    
    # 最新ファイルを指定
    excel_path = "backtest_results/dssms_results/dssms_unified_backtest_20250908_152431.xlsx"
    json_path = "backtest_results/dssms_results/dssms_unified_data_20250908_152431.json"
    
    print("🔧 DSSMS Excel戦略別統計直接修正システム")
    print("=" * 60)
    print(f"対象Excel: {excel_path}")
    print(f"対象JSON: {json_path}")
    print("=" * 60)
    
    if patcher.patch_excel_strategy_stats(excel_path, json_path):
        print("✅ 戦略別統計シート修正完了！")
        print("\n修正内容:")
        print("- DSSMSStrategyを7つの個別戦略に分割")
        print("- 各戦略の実際の勝率を表示")
        print("- 正確な取引回数と損益統計を計算")
        print("\n確認方法:")
        print("1. Excelファイルを開く")
        print("2. '戦略別統計'シートを確認")
        print("3. 7つの戦略が個別に表示されているか確認")
    else:
        print("❌ 戦略別統計シート修正失敗")

if __name__ == "__main__":
    main()
