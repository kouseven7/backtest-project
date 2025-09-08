#!/usr/bin/env python3
"""
最新DSSMS戦略別統計アナライザー

最新生成されたDSSMSファイルを対象に戦略別統計問題をチェック
"""

import pandas as pd
import json
import logging
import openpyxl
from datetime import datetime
from pathlib import Path

def setup_logger():
    """ロガー設定"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

def analyze_latest_strategy_stats():
    """最新ファイルの戦略別統計を分析"""
    logger = setup_logger()
    
    # 最新ファイル指定
    latest_excel = "backtest_results/dssms_results/dssms_unified_backtest_20250908_152431.xlsx"
    latest_json = "backtest_results/dssms_results/dssms_unified_data_20250908_152431.json"
    
    logger.info("🔍 最新DSSMS戦略別統計分析開始")
    logger.info(f"対象Excel: {latest_excel}")
    logger.info(f"対象JSON: {latest_json}")
    
    # JSONデータから戦略別統計を確認
    try:
        with open(latest_json, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        trades = json_data.get('trades', [])
        logger.info(f"📦 JSON取引データ: {len(trades)}件")
        
        # 戦略別に集計
        strategy_trades = {}
        for trade in trades:
            strategy = trade.get('strategy', 'Unknown')
            if strategy not in strategy_trades:
                strategy_trades[strategy] = []
            strategy_trades[strategy].append(trade)
        
        logger.info(f"📊 JSON戦略種類: {len(strategy_trades)}種類")
        for strategy, trade_list in strategy_trades.items():
            pnls = [float(t.get('pnl', 0)) for t in trade_list]
            wins = len([p for p in pnls if p > 0])
            win_rate = (wins / len(pnls) * 100) if pnls else 0
            logger.info(f"  {strategy}: {len(trade_list)}件, 勝率{win_rate:.1f}%")
    
    except Exception as e:
        logger.error(f"JSON分析エラー: {e}")
    
    # Excelファイルの戦略別統計を確認
    try:
        workbook = openpyxl.load_workbook(latest_excel)
        logger.info(f"📊 Excel利用可能シート: {workbook.sheetnames}")
        
        if '戦略別統計' in workbook.sheetnames:
            ws = workbook['戦略別統計']
            logger.info("✅ 戦略別統計シート発見")
            
            # データを読み取り
            data = []
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    data.append(row)
            
            logger.info(f"📋 戦略別統計データ: {len(data)}行")
            
            if data:
                headers = data[0]
                logger.info(f"ヘッダー: {headers}")
                
                for i, row in enumerate(data[1:], 1):
                    if row and len(row) > 0:
                        strategy_name = row[0] if row[0] else "不明"
                        win_rate = row[2] if len(row) > 2 else "不明"
                        logger.info(f"  行{i}: {strategy_name}, 勝率={win_rate}")
            
            # 問題チェック
            strategy_names = [row[0] for row in data[1:] if row and row[0]]
            unique_strategies = list(set(strategy_names))
            
            if len(unique_strategies) == 1 and 'DSSMS' in str(unique_strategies[0]):
                logger.warning("❌ 問題: DSSMSのみ表示、7戦略が不足")
            elif len(unique_strategies) >= 7:
                logger.info("✅ 解決: 7つ以上の戦略が表示されています")
            else:
                logger.warning(f"⚠️ 部分解決: {len(unique_strategies)}戦略表示（7戦略未満）")
            
        else:
            logger.warning("❌ 戦略別統計シートが見つかりません")
    
    except Exception as e:
        logger.error(f"Excel分析エラー: {e}")
    
    logger.info("🔍 分析完了")

if __name__ == "__main__":
    analyze_latest_strategy_stats()
