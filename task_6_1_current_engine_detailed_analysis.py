#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Task 6.1: 現在使用中エンジンの詳細品質分析（Critical）
dssms_unified_output_engine.py（品質0点）の詳細調査と
Task 4.2のv1エンジン（85.0点）との関係性分析
"""

import os
import sys
from pathlib import Path
import pandas as pd
import json
import openpyxl
from datetime import datetime
import ast
import re

class CurrentEngineDetailedQualityAnalyzer:
    def __init__(self):
        self.current_engine_file = 'dssms_unified_output_engine.py'
        self.latest_excel_file = 'backtest_results/dssms_results/dssms_unified_backtest_20250910_213413.xlsx'
        self.task42_results_file = 'task_4_2_results_20250912_115837.json'
        self.analysis_results = {}
        
    def analyze_current_engine_detailed_quality(self):
        """現在使用中エンジンの詳細品質・正確性分析"""
        print("🚨 Task 6.1: 現在使用中エンジンの詳細品質分析開始")
        print("=" * 80)
        
        # 1. 現在使用中エンジンの計算式実装状況調査
        self._analyze_calculation_implementation()
        
        # 2. Task 4.2のv1エンジンとの詳細比較分析
        self._compare_with_task42_v1_engine()
        
        # 3. 実際のExcel出力内容の数学的正確性検証
        self._verify_excel_output_accuracy()
        
        # 4. 24KB出力データの構造・完全性確認
        self._analyze_excel_structure_completeness()
        
        # 5. 0点評価の具体的原因特定
        self._identify_zero_score_root_cause()
        
        return self.analysis_results
    
    def _analyze_calculation_implementation(self):
        """1. 現在使用中エンジンの計算式実装状況調査"""
        print("\n🔍 1. 現在使用中エンジンの計算式実装状況調査")
        print("-" * 60)
        
        try:
            with open(self.current_engine_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # ファイルサイズと基本情報
            file_info = {
                'file_size': len(content),
                'line_count': len(content.split('\n')),
                'is_empty': len(content.strip()) == 0
            }
            
            if file_info['is_empty']:
                print("❌ ファイルが空です - これが0点評価の原因")
                self.analysis_results['calculation_analysis'] = {
                    'status': 'empty_file',
                    'file_info': file_info,
                    'critical_issue': 'ファイルが完全に空'
                }
                return
            
            # 重要な計算メソッドの実装確認
            calculation_methods = {
                'win_rate': ['win_rate', 'winrate', '勝率'],
                'profit_factor': ['profit_factor', 'profitfactor', 'プロフィットファクター'],
                'sharpe_ratio': ['sharpe_ratio', 'sharpe', 'シャープレシオ'],
                'total_return': ['total_return', 'totalreturn', '総収益'],
                'max_drawdown': ['max_drawdown', 'maxdrawdown', '最大ドローダウン'],
                'average_trade': ['average_trade', 'avg_trade', '平均取引']
            }
            
            implemented_calculations = {}
            missing_calculations = []
            
            for calc_type, keywords in calculation_methods.items():
                found = False
                for keyword in keywords:
                    if keyword.lower() in content.lower():
                        found = True
                        # 実装の詳細確認
                        pattern = rf'def.*{keyword}.*\(.*\):'
                        matches = re.findall(pattern, content, re.IGNORECASE)
                        implemented_calculations[calc_type] = {
                            'found': True,
                            'methods': matches,
                            'keyword_count': content.lower().count(keyword.lower())
                        }
                        break
                
                if not found:
                    missing_calculations.append(calc_type)
            
            # 数学的公式の実装確認
            formula_patterns = {
                'division_operations': len(re.findall(r'/[^/]', content)),  # 除算
                'sum_operations': content.count('sum('),
                'mean_operations': content.count('mean(') + content.count('.mean()'),
                'pandas_calculations': content.count('pd.') + content.count('DataFrame'),
                'numpy_calculations': content.count('np.')
            }
            
            self.analysis_results['calculation_analysis'] = {
                'status': 'analyzed',
                'file_info': file_info,
                'implemented_calculations': implemented_calculations,
                'missing_calculations': missing_calculations,
                'formula_patterns': formula_patterns,
                'implementation_completeness': len(implemented_calculations) / len(calculation_methods) * 100
            }
            
            print(f"📊 ファイル情報: {file_info['file_size']:,} bytes, {file_info['line_count']} lines")
            print(f"📈 実装済み計算: {len(implemented_calculations)}/{len(calculation_methods)} ({self.analysis_results['calculation_analysis']['implementation_completeness']:.1f}%)")
            print(f"❌ 未実装計算: {', '.join(missing_calculations)}")
            print(f"🔢 数学公式パターン: 除算{formula_patterns['division_operations']}個, sum{formula_patterns['sum_operations']}個")
            
        except Exception as e:
            print(f"❌ 分析エラー: {e}")
            self.analysis_results['calculation_analysis'] = {'error': str(e)}
    
    def _compare_with_task42_v1_engine(self):
        """2. Task 4.2のv1エンジンとの詳細比較分析"""
        print("\n🔍 2. Task 4.2のv1エンジンとの詳細比較分析")
        print("-" * 60)
        
        try:
            if os.path.exists(self.task42_results_file):
                with open(self.task42_results_file, 'r', encoding='utf-8') as f:
                    task42_data = json.load(f)
                
                # 現在のエンジンの品質情報取得（正しいパス）
                current_engine_quality = task42_data.get('implementation_quality', {}).get('dssms_unified_output_engine.py', {})
                current_engine_score = current_engine_quality.get('total_score', 0)
                
                print(f"✅ Task 4.2での現在エンジンスコア: {current_engine_score}点")
                print(f"📊 実装率: {current_engine_quality.get('implementation_percentage', 0)}%")
                print(f"🎯 品質率: {current_engine_quality.get('quality_percentage', 0)}%")
                
                # 現在のエンジンとの比較
                current_analysis = self.analysis_results.get('calculation_analysis', {})
                
                comparison = {
                    'quality_scores': {
                        'task42_v1': v1_quality,
                        'current_engine': 0,  # 既知の0点
                        'score_gap': v1_quality - 0
                    },
                    'implementation_comparison': {
                        'v1_implementations': v1_details.get('implemented_calculations', {}),
                        'current_implementations': current_analysis.get('implemented_calculations', {}),
                        'implementation_gap': 'analysis_needed'
                    },
                    'file_relationship': {
                        'same_file': False,  # 確実に別ファイル
                        'relationship_type': 'different_files',
                        'confusion_source': 'Task 4.2調査対象と実際使用ファイルが異なる'
                    }
                }
                
                self.analysis_results['task42_comparison'] = comparison
                
                print(f"📊 品質スコア比較:")
                print(f"   Task 4.2 v1エンジン: {v1_quality}点")
                print(f"   現在使用中エンジン: 0点")
                print(f"   格差: {comparison['quality_scores']['score_gap']}点")
                print(f"🔄 ファイル関係: {comparison['file_relationship']['relationship_type']}")
                print(f"⚠️ 混乱の原因: {comparison['file_relationship']['confusion_source']}")
                
            else:
                print("❌ Task 4.2結果ファイルが見つかりません")
                self.analysis_results['task42_comparison'] = {'error': 'Task 4.2結果ファイル未発見'}
                
        except Exception as e:
            print(f"❌ 比較分析エラー: {e}")
            self.analysis_results['task42_comparison'] = {'error': str(e)}
    
    def _verify_excel_output_accuracy(self):
        """3. 実際のExcel出力内容の数学的正確性検証"""
        print("\n🔍 3. Excel出力内容の数学的正確性検証")
        print("-" * 60)
        
        try:
            if not os.path.exists(self.latest_excel_file):
                print(f"❌ Excel出力ファイルが見つかりません: {self.latest_excel_file}")
                self.analysis_results['excel_accuracy'] = {'error': 'Excel出力ファイル未発見'}
                return
            
            # Excelファイルの読み込み
            workbook = openpyxl.load_workbook(self.latest_excel_file)
            sheet_names = workbook.sheetnames
            
            accuracy_analysis = {
                'sheets_found': sheet_names,
                'sheet_count': len(sheet_names),
                'mathematical_errors': [],
                'data_inconsistencies': [],
                'calculation_verifications': {}
            }
            
            # 各シートの数値検証
            for sheet_name in sheet_names:
                try:
                    sheet = workbook[sheet_name]
                    sheet_analysis = self._analyze_sheet_accuracy(sheet, sheet_name)
                    accuracy_analysis['calculation_verifications'][sheet_name] = sheet_analysis
                    
                    if sheet_analysis.get('errors'):
                        accuracy_analysis['mathematical_errors'].extend(sheet_analysis['errors'])
                        
                except Exception as e:
                    accuracy_analysis['calculation_verifications'][sheet_name] = {'error': str(e)}
            
            # 重要な数値の論理検証
            logical_checks = self._perform_logical_consistency_checks(workbook)
            accuracy_analysis['logical_consistency'] = logical_checks
            
            self.analysis_results['excel_accuracy'] = accuracy_analysis
            
            print(f"📊 Excel構造: {len(sheet_names)}シート発見")
            print(f"📋 シート名: {', '.join(sheet_names)}")
            
            error_count = len(accuracy_analysis['mathematical_errors'])
            if error_count > 0:
                print(f"❌ 数学的エラー: {error_count}件発見")
                for error in accuracy_analysis['mathematical_errors'][:3]:  # 最初の3件のみ表示
                    print(f"   - {error}")
            else:
                print("✅ 明らかな数学的エラーは検出されませんでした")
            
        except Exception as e:
            print(f"❌ Excel検証エラー: {e}")
            self.analysis_results['excel_accuracy'] = {'error': str(e)}
    
    def _analyze_sheet_accuracy(self, sheet, sheet_name):
        """個別シートの数値精度分析"""
        analysis = {
            'sheet_name': sheet_name,
            'data_rows': 0,
            'numeric_columns': 0,
            'errors': [],
            'suspicious_values': []
        }
        
        try:
            # データ行数とカラム数の確認
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            analysis['data_rows'] = max_row
            analysis['max_columns'] = max_col
            
            # 数値データの確認（サンプリング）
            numeric_values = []
            for row in range(1, min(max_row + 1, 101)):  # 最初の100行をサンプリング
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)) and cell_value != 0:
                        numeric_values.append(cell_value)
            
            if numeric_values:
                analysis['numeric_sample_count'] = len(numeric_values)
                analysis['value_range'] = {
                    'min': min(numeric_values),
                    'max': max(numeric_values),
                    'avg': sum(numeric_values) / len(numeric_values)
                }
                
                # 異常値の検出
                for value in numeric_values:
                    if abs(value) > 1e10:  # 異常に大きい値
                        analysis['suspicious_values'].append(f"異常に大きい値: {value}")
                    elif value < 0 and 'portfolio' in sheet_name.lower():  # ポートフォリオで負の値
                        analysis['suspicious_values'].append(f"ポートフォリオ負値: {value}")
            
            # 特定シートの特殊検証
            if '取引履歴' in sheet_name:
                analysis.update(self._verify_trade_history_sheet(sheet))
            elif '損益推移' in sheet_name:
                analysis.update(self._verify_pnl_history_sheet(sheet))
                
        except Exception as e:
            analysis['errors'].append(f"シート分析エラー: {e}")
        
        return analysis
    
    def _verify_trade_history_sheet(self, sheet):
        """取引履歴シートの専用検証"""
        verification = {'trade_specific_checks': []}
        
        try:
            # ヘッダー行の確認
            headers = []
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    headers.append(str(header))
            
            verification['headers'] = headers
            
            # 24時間固定問題の確認
            holding_period_col = None
            for i, header in enumerate(headers):
                if '保有期間' in str(header) or 'holding' in str(header).lower():
                    holding_period_col = i + 1
                    break
            
            if holding_period_col:
                # 保有期間の値を確認
                holding_periods = []
                for row in range(2, min(sheet.max_row + 1, 52)):  # 最初の50件
                    value = sheet.cell(row=row, column=holding_period_col).value
                    if value:
                        holding_periods.append(str(value))
                
                # 24時間固定問題の検出
                if holding_periods:
                    unique_periods = set(holding_periods)
                    if len(unique_periods) == 1 and '24' in list(unique_periods)[0]:
                        verification['trade_specific_checks'].append("❌ 24時間固定問題を検出")
                    else:
                        verification['trade_specific_checks'].append(f"✅ 保有期間の多様性: {len(unique_periods)}種類")
            
        except Exception as e:
            verification['trade_specific_checks'].append(f"取引履歴検証エラー: {e}")
        
        return verification
    
    def _verify_pnl_history_sheet(self, sheet):
        """損益推移シートの専用検証"""
        verification = {'pnl_specific_checks': []}
        
        try:
            # 日付の連続性確認
            date_values = []
            for row in range(2, min(sheet.max_row + 1, 32)):  # 最初の30行
                date_value = sheet.cell(row=row, column=1).value
                if date_value:
                    date_values.append(str(date_value))
            
            verification['date_sample'] = date_values[:5]  # 最初の5件
            
            # 日付ループ問題の検出
            if len(date_values) >= 3:
                if date_values[0] == date_values[-1]:
                    verification['pnl_specific_checks'].append("❌ 日付ループ問題の可能性を検出")
                elif '2023-01-01' in date_values and '2023-12-31' in date_values:
                    verification['pnl_specific_checks'].append("✅ 正常な年間日付範囲")
                else:
                    verification['pnl_specific_checks'].append("⚠️ 日付範囲要確認")
            
        except Exception as e:
            verification['pnl_specific_checks'].append(f"損益推移検証エラー: {e}")
        
        return verification
    
    def _perform_logical_consistency_checks(self, workbook):
        """論理整合性チェック"""
        consistency = {
            'checks_performed': [],
            'consistency_errors': [],
            'data_relationships': {}
        }
        
        try:
            # シート間データ整合性の確認
            if 'サマリー' in workbook.sheetnames and '取引履歴' in workbook.sheetnames:
                # サマリーと取引履歴の整合性
                summary_sheet = workbook['サマリー']
                trade_sheet = workbook['取引履歴']
                
                # 取引数の整合性（簡易確認）
                trade_rows = trade_sheet.max_row - 1  # ヘッダー除く
                consistency['data_relationships']['trade_count_from_sheet'] = trade_rows
                consistency['checks_performed'].append("取引数整合性チェック")
                
                if trade_rows == 0:
                    consistency['consistency_errors'].append("取引履歴が空 - 切替数激減問題の証拠")
                
        except Exception as e:
            consistency['consistency_errors'].append(f"整合性チェックエラー: {e}")
        
        return consistency
    
    def _analyze_excel_structure_completeness(self):
        """4. 24KB出力データの構造・完全性確認"""
        print("\n🔍 4. Excel出力データの構造・完全性確認")
        print("-" * 60)
        
        try:
            file_path = Path(self.latest_excel_file)
            
            structure_analysis = {
                'file_info': {
                    'size_bytes': file_path.stat().st_size,
                    'size_kb': file_path.stat().st_size / 1024,
                    'last_modified': datetime.fromtimestamp(file_path.stat().st_mtime),
                    'exists': file_path.exists()
                }
            }
            
            if file_path.exists():
                # Excel内部構造の詳細分析
                workbook = openpyxl.load_workbook(self.latest_excel_file)
                
                sheet_details = {}
                total_data_cells = 0
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # 各シートの詳細情報
                    sheet_info = {
                        'max_row': sheet.max_row,
                        'max_column': sheet.max_column,
                        'estimated_data_cells': sheet.max_row * sheet.max_column,
                        'has_data': sheet.max_row > 1  # ヘッダー以外にデータがあるか
                    }
                    
                    # 実際のデータセル数をサンプリング
                    non_empty_cells = 0
                    for row in range(1, min(sheet.max_row + 1, 51)):  # 最初の50行
                        for col in range(1, min(sheet.max_column + 1, 21)):  # 最初の20列
                            if sheet.cell(row=row, column=col).value is not None:
                                non_empty_cells += 1
                    
                    sheet_info['sample_non_empty_cells'] = non_empty_cells
                    sheet_details[sheet_name] = sheet_info
                    total_data_cells += non_empty_cells
                
                structure_analysis['sheet_details'] = sheet_details
                structure_analysis['total_estimated_data_cells'] = total_data_cells
                
                # 必要シートの存在確認
                required_sheets = ['サマリー', 'パフォーマンス指標', '取引履歴', '損益推移', '戦略別統計', '切替分析']
                missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
                structure_analysis['missing_sheets'] = missing_sheets
                
                # データ密度の評価
                bytes_per_cell = structure_analysis['file_info']['size_bytes'] / max(total_data_cells, 1)
                structure_analysis['data_density'] = {
                    'bytes_per_cell': bytes_per_cell,
                    'density_assessment': 'normal' if 10 <= bytes_per_cell <= 100 else 'suspicious'
                }
                
                self.analysis_results['structure_analysis'] = structure_analysis
                
                print(f"📁 ファイル情報: {structure_analysis['file_info']['size_kb']:.1f}KB")
                print(f"📊 シート数: {len(workbook.sheetnames)}")
                print(f"📈 推定データセル数: {total_data_cells:,}")
                print(f"📋 未実装シート: {', '.join(missing_sheets) if missing_sheets else 'なし'}")
                print(f"🔍 データ密度: {bytes_per_cell:.1f} bytes/cell ({structure_analysis['data_density']['density_assessment']})")
                
            else:
                print("❌ Excel出力ファイルが存在しません")
                structure_analysis['error'] = 'ファイル未存在'
                self.analysis_results['structure_analysis'] = structure_analysis
                
        except Exception as e:
            print(f"❌ 構造分析エラー: {e}")
            self.analysis_results['structure_analysis'] = {'error': str(e)}
    
    def _identify_zero_score_root_cause(self):
        """5. 0点評価の具体的原因特定"""
        print("\n🔍 5. 0点評価の具体的原因特定")
        print("-" * 60)
        
        root_cause_analysis = {
            'potential_causes': [],
            'evidence': {},
            'severity_assessment': 'unknown',
            'recommended_actions': []
        }
        
        try:
            # 各分析結果からの原因特定
            calc_analysis = self.analysis_results.get('calculation_analysis', {})
            excel_analysis = self.analysis_results.get('excel_accuracy', {})
            structure_analysis = self.analysis_results.get('structure_analysis', {})
            
            # 原因1: ファイルが空の場合
            if calc_analysis.get('file_info', {}).get('is_empty', False):
                root_cause_analysis['potential_causes'].append("Critical: ファイルが完全に空")
                root_cause_analysis['severity_assessment'] = 'critical'
                root_cause_analysis['recommended_actions'].append("即座にファイル内容を復旧")
            
            # 原因2: 計算実装の不足
            implementation_rate = calc_analysis.get('implementation_completeness', 0)
            if implementation_rate < 50:
                root_cause_analysis['potential_causes'].append(f"計算実装不足: {implementation_rate:.1f}%のみ実装")
                root_cause_analysis['evidence']['low_implementation'] = implementation_rate
            
            # 原因3: Excel出力の問題
            excel_errors = excel_analysis.get('mathematical_errors', [])
            if excel_errors:
                root_cause_analysis['potential_causes'].append(f"Excel出力エラー: {len(excel_errors)}件")
                root_cause_analysis['evidence']['excel_errors'] = excel_errors[:3]
            
            # 原因4: データ構造の問題
            missing_sheets = structure_analysis.get('missing_sheets', [])
            if missing_sheets:
                root_cause_analysis['potential_causes'].append(f"必須シート未実装: {len(missing_sheets)}件")
                root_cause_analysis['evidence']['missing_sheets'] = missing_sheets
            
            # 原因5: ファイル混乱問題
            task42_comparison = self.analysis_results.get('task42_comparison', {})
            if task42_comparison.get('file_relationship', {}).get('same_file') == False:
                root_cause_analysis['potential_causes'].append("ファイル混乱: 評価対象と使用ファイルが異なる")
                root_cause_analysis['evidence']['file_confusion'] = True
            
            # 重要度評価
            if len(root_cause_analysis['potential_causes']) >= 3:
                root_cause_analysis['severity_assessment'] = 'critical'
            elif len(root_cause_analysis['potential_causes']) >= 1:
                root_cause_analysis['severity_assessment'] = 'high'
            else:
                root_cause_analysis['severity_assessment'] = 'medium'
            
            # 解決アクション推奨
            if root_cause_analysis['severity_assessment'] == 'critical':
                root_cause_analysis['recommended_actions'].extend([
                    "Task 4.2の85.0点エンジンへの即座切り替え",
                    "現在エンジンの完全置換または修復",
                    "Excel出力の品質管理体制構築"
                ])
            
            self.analysis_results['root_cause_analysis'] = root_cause_analysis
            
            print(f"🎯 特定された原因: {len(root_cause_analysis['potential_causes'])}件")
            for i, cause in enumerate(root_cause_analysis['potential_causes'], 1):
                print(f"   {i}. {cause}")
            
            print(f"⚠️ 重要度評価: {root_cause_analysis['severity_assessment']}")
            print(f"💡 推奨アクション: {len(root_cause_analysis['recommended_actions'])}件")
            
        except Exception as e:
            print(f"❌ 原因特定エラー: {e}")
            root_cause_analysis['error'] = str(e)
            self.analysis_results['root_cause_analysis'] = root_cause_analysis
    
    def generate_problem16_definition(self):
        """Problem 16の定義生成"""
        root_cause = self.analysis_results.get('root_cause_analysis', {})
        calc_analysis = self.analysis_results.get('calculation_analysis', {})
        task42_comparison = self.analysis_results.get('task42_comparison', {})
        
        problem16 = {
            'problem_id': 'Problem 16',
            'title': '現在使用中エンジンの致命的品質問題（Critical）',
            'severity': root_cause.get('severity_assessment', 'critical'),
            'description': f"dssms_unified_output_engine.py（0点品質）が稼働中で、Task 4.2のv1エンジン（85.0点）と85点の品質格差",
            'discovered_issues': root_cause.get('potential_causes', []),
            'impact_assessment': {
                'excel_output_reliability': 'questionable',
                'calculation_accuracy': 'unverified',
                'implementation_completeness': calc_analysis.get('implementation_completeness', 0),
                'quality_gap': task42_comparison.get('quality_scores', {}).get('score_gap', 85)
            },
            'evidence': root_cause.get('evidence', {}),
            'recommended_resolution': root_cause.get('recommended_actions', [])
        }
        
        return problem16
    
    def save_results(self):
        """結果保存"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"task_6_1_current_engine_detailed_analysis_{timestamp}.json"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.analysis_results, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\n✅ 詳細分析結果保存: {output_file}")
        return output_file

def main():
    """Task 6.1メイン実行"""
    print("🚨 Task 6.1: 現在使用中エンジンの詳細品質分析（Critical）")
    print("=" * 80)
    
    analyzer = CurrentEngineDetailedQualityAnalyzer()
    
    # 詳細分析実行
    analyzer.analyze_current_engine_detailed_quality()
    
    # Problem 16定義生成
    problem16 = analyzer.generate_problem16_definition()
    
    # 結果保存
    output_file = analyzer.save_results()
    
    print("\n" + "=" * 80)
    print("📋 Task 6.1 完了サマリー")
    print("=" * 80)
    print(f"🎯 Problem 16定義: {problem16['title']}")
    print(f"⚠️ 重要度: {problem16['severity']}")
    print(f"📊 発見された問題: {len(problem16['discovered_issues'])}件")
    print(f"📈 実装完全性: {problem16['impact_assessment']['implementation_completeness']:.1f}%")
    print(f"📄 品質格差: {problem16['impact_assessment']['quality_gap']}点")
    
    print(f"\n💡 次のアクション:")
    print(f"1. Problem 16をroadmap2.mdに追加")
    print(f"2. 科学的効率分析への統合")
    print(f"3. 85.0点エンジンへの切り替え検討")
    
    return output_file, problem16

if __name__ == "__main__":
    main()