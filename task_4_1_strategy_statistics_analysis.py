#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Task 4.1: 戦略別統計未計算問題の特定

戦略別統計シートの計算ロジック問題を特定し、
勝率、平均利益/損失、プロフィットファクター等の
統計項目未計算の原因を調査する
"""

import pandas as pd
import numpy as np
import json
from datetime import datetime, timedelta
from pathlib import Path
import sys
import traceback
import re

def analyze_strategy_statistics_data_source():
    """戦略統計計算のデータソース調査"""
    
    results = {
        "analysis_timestamp": datetime.now().strftime("%Y%m%d_%H%M%S"),
        "findings": {},
        "data_sources": {},
        "calculation_implementations": {},
        "problems_identified": [],
        "recommendations": []
    }
    
    print("🔍 Task 4.1: 戦略別統計未計算問題の特定を開始")
    print("=" * 60)
    
    # 1. DSSMSBacktester.get_strategy_statistics()の実装状況
    print("\n📊 Step 1: DSSMSBacktester統計メソッド分析")
    try:
        # DSSMSBacktesterクラスを調査
        backtester_path = Path("src/dssms/dssms_backtester.py")
        if backtester_path.exists():
            with open(backtester_path, 'r', encoding='utf-8') as f:
                backtester_content = f.read()
            
            # 統計関連メソッドを検索
            statistics_methods = []
            method_keywords = [
                "get_strategy_statistics", "calculate_statistics", 
                "performance_metrics", "win_rate", "profit_factor",
                "statistics", "get_performance_metrics"
            ]
            
            for keyword in method_keywords:
                for i, line in enumerate(backtester_content.split('\n'), 1):
                    if keyword.lower() in line.lower() and ('def ' in line or 'class ' in line):
                        statistics_methods.append(f"Line {i}: {line.strip()}")
            
            # 勝率計算の実装を検索
            win_rate_patterns = []
            profit_patterns = []
            for i, line in enumerate(backtester_content.split('\n'), 1):
                line_lower = line.lower()
                if 'win' in line_lower and ('rate' in line_lower or 'ratio' in line_lower):
                    win_rate_patterns.append(f"Line {i}: {line.strip()}")
                if 'profit' in line_lower and ('factor' in line_lower or 'loss' in line_lower):
                    profit_patterns.append(f"Line {i}: {line.strip()}")
            
            results["data_sources"]["backtester_analysis"] = {
                "file_size": len(backtester_content),
                "statistics_methods_found": len(statistics_methods),
                "method_details": statistics_methods[:10],
                "win_rate_implementations": win_rate_patterns[:5],
                "profit_implementations": profit_patterns[:5]
            }
            
            print(f"✅ DSSMSBacktester統計メソッド: {len(statistics_methods)}件発見")
            print(f"📈 勝率関連実装: {len(win_rate_patterns)}件")
            print(f"💰 利益関連実装: {len(profit_patterns)}件")
            
            if statistics_methods:
                print("🔍 発見された統計メソッド:")
                for method in statistics_methods[:3]:
                    print(f"   {method}")
        else:
            results["problems_identified"].append("DSSMSBacktesterファイルが見つからない")
            print("❌ DSSMSBacktesterファイルが見つかりません")
            
    except Exception as e:
        print(f"❌ DSSMSBacktester分析エラー: {e}")
        results["problems_identified"].append(f"DSSMSBacktester分析エラー: {e}")
    
    # 2. trade_historyデータの構造と内容分析
    print("\n💼 Step 2: trade_historyデータ構造分析")
    try:
        # 各統一エンジンでのtrade_history処理を比較
        engine_files = [
            "dssms_unified_output_engine.py",
            "dssms_unified_output_engine_fixed.py", 
            "dssms_unified_output_engine_fixed_v3.py",
            "dssms_unified_output_engine_fixed_v4.py"
        ]
        
        trade_history_analysis = {}
        for engine_file in engine_files:
            engine_path = Path(engine_file)
            if engine_path.exists():
                with open(engine_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # trade_historyの使用箇所を検索
                trade_usage = []
                for i, line in enumerate(content.split('\n'), 1):
                    if 'trade_history' in line.lower():
                        trade_usage.append(f"Line {i}: {line.strip()}")
                
                # 統計計算の実装を検索
                stats_calculations = []
                stats_keywords = ["win_rate", "profit_factor", "average", "total_profit", "total_loss"]
                for keyword in stats_keywords:
                    for i, line in enumerate(content.split('\n'), 1):
                        if keyword in line.lower():
                            stats_calculations.append(f"{keyword} Line {i}: {line.strip()}")
                
                trade_history_analysis[engine_file] = {
                    "trade_history_usage": len(trade_usage),
                    "usage_details": trade_usage[:5],
                    "statistics_calculations": len(stats_calculations),
                    "calculation_details": stats_calculations[:5]
                }
                
                print(f"📁 {engine_file}:")
                print(f"   trade_history使用: {len(trade_usage)}箇所")
                print(f"   統計計算実装: {len(stats_calculations)}箇所")
                
        results["data_sources"]["trade_history_analysis"] = trade_history_analysis
        
    except Exception as e:
        print(f"❌ trade_history分析エラー: {e}")
        results["problems_identified"].append(f"trade_history分析エラー: {e}")
    
    # 3. switch_historyからの統計計算可能性
    print("\n🔄 Step 3: switch_historyからの統計計算可能性分析")
    try:
        switch_statistics_potential = {}
        
        for engine_file in engine_files:
            engine_path = Path(engine_file)
            if engine_path.exists():
                with open(engine_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # switch_historyの統計利用を検索
                switch_stats_usage = []
                for i, line in enumerate(content.split('\n'), 1):
                    if 'switch_history' in line.lower() and any(keyword in line.lower() 
                        for keyword in ['profit', 'loss', 'gain', 'return', 'performance']):
                        switch_stats_usage.append(f"Line {i}: {line.strip()}")
                
                # 切替ベースの統計計算を検索
                switch_calculations = []
                switch_keywords = ["switch_count", "holding_period", "switch_profit", "switch_performance"]
                for keyword in switch_keywords:
                    for i, line in enumerate(content.split('\n'), 1):
                        if keyword in line.lower():
                            switch_calculations.append(f"{keyword} Line {i}: {line.strip()}")
                
                switch_statistics_potential[engine_file] = {
                    "switch_stats_usage": len(switch_stats_usage),
                    "usage_details": switch_stats_usage[:3],
                    "switch_calculations": len(switch_calculations),
                    "calculation_details": switch_calculations[:3]
                }
                
                print(f"📁 {engine_file}:")
                print(f"   switch統計利用: {len(switch_stats_usage)}箇所")
                print(f"   switch計算実装: {len(switch_calculations)}箇所")
                
        results["data_sources"]["switch_statistics_potential"] = switch_statistics_potential
        
    except Exception as e:
        print(f"❌ switch_history統計分析エラー: {e}")
        results["problems_identified"].append(f"switch_history統計分析エラー: {e}")
    
    # 4. 各統一エンジンでの統計計算実装比較
    print("\n🧮 Step 4: 統計計算実装比較")
    try:
        calculation_implementations = {}
        
        # 重要な統計計算の実装状況をチェック
        target_calculations = {
            "win_rate": "勝率計算",
            "profit_factor": "プロフィットファクター",
            "average_profit": "平均利益",
            "average_loss": "平均損失", 
            "total_trades": "総取引数",
            "profitable_trades": "利益取引数"
        }
        
        for engine_file in engine_files:
            engine_path = Path(engine_file)
            if engine_path.exists():
                with open(engine_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                implementation_status = {}
                for calc_name, calc_desc in target_calculations.items():
                    # 各計算の実装を検索
                    implementations = []
                    for i, line in enumerate(content.split('\n'), 1):
                        if calc_name.replace('_', '').lower() in line.replace('_', '').lower():
                            implementations.append(f"Line {i}: {line.strip()}")
                    
                    implementation_status[calc_name] = {
                        "implemented": len(implementations) > 0,
                        "implementation_count": len(implementations),
                        "details": implementations[:2]
                    }
                
                calculation_implementations[engine_file] = implementation_status
                
                # 実装状況の集計
                implemented_count = sum(1 for calc in implementation_status.values() if calc["implemented"])
                print(f"📁 {engine_file}: {implemented_count}/{len(target_calculations)}項目実装")
                
        results["calculation_implementations"] = calculation_implementations
        
    except Exception as e:
        print(f"❌ 統計計算実装比較エラー: {e}")
        results["problems_identified"].append(f"統計計算実装比較エラー: {e}")
    
    # 5. Excel出力での戦略別統計シート分析
    print("\n📊 Step 5: Excel戦略別統計シート分析")
    try:
        # 最新のExcelファイルを検索
        excel_files = list(Path("backtest_results/dssms_results").glob("*.xlsx"))
        excel_files = [f for f in excel_files if not f.name.startswith("~$")]  # ロックファイル除外
        
        if excel_files:
            latest_excel = max(excel_files, key=lambda x: x.stat().st_mtime)
            print(f"📄 最新Excelファイル: {latest_excel.name}")
            
            # Excelファイルの戦略別統計シートを分析
            try:
                import openpyxl
                wb = openpyxl.load_workbook(latest_excel)
                
                excel_statistics_analysis = {
                    "file_name": latest_excel.name,
                    "sheet_names": wb.sheetnames,
                    "strategy_stats_analysis": {}
                }
                
                # 戦略別統計シートの分析
                strategy_sheet_names = [name for name in wb.sheetnames if "戦略" in name or "統計" in name]
                
                for sheet_name in strategy_sheet_names:
                    ws = wb[sheet_name]
                    
                    # シートの内容を分析
                    sheet_data = []
                    for row in range(1, min(11, ws.max_row + 1)):  # 最初10行
                        row_data = {}
                        for col in range(1, min(6, ws.max_column + 1)):  # 最初5列
                            cell_value = ws.cell(row=row, column=col).value
                            row_data[f"col_{col}"] = str(cell_value) if cell_value is not None else ""
                        sheet_data.append(row_data)
                    
                    # 統計項目の存在確認
                    stats_found = {
                        "勝率": False,
                        "利益": False,
                        "損失": False,
                        "プロフィットファクター": False,
                        "取引数": False
                    }
                    
                    for row_data in sheet_data:
                        for cell_value in row_data.values():
                            if isinstance(cell_value, str):
                                for stat_name in stats_found:
                                    if stat_name in cell_value:
                                        stats_found[stat_name] = True
                    
                    excel_statistics_analysis["strategy_stats_analysis"][sheet_name] = {
                        "max_row": ws.max_row,
                        "max_column": ws.max_column,
                        "sample_data": sheet_data,
                        "statistics_items_found": stats_found
                    }
                    
                    print(f"✅ {sheet_name}シート: {ws.max_row}行 x {ws.max_column}列")
                    stats_count = sum(stats_found.values())
                    print(f"   統計項目発見: {stats_count}/5項目")
                
                results["findings"]["excel_statistics_analysis"] = excel_statistics_analysis
                
            except Exception as excel_error:
                print(f"❌ Excel分析エラー: {excel_error}")
                results["problems_identified"].append(f"Excel分析エラー: {excel_error}")
                
        else:
            print("❌ Excelファイルが見つかりません")
            results["problems_identified"].append("Excelファイルが見つからない")
            
    except Exception as e:
        print(f"❌ Excel統計分析エラー: {e}")
        results["problems_identified"].append(f"Excel統計分析エラー: {e}")
    
    # 6. 問題の特定と推奨事項
    print("\n🎯 Step 6: 統計計算問題の特定")
    
    # 実装格差の分析
    implementation_gaps = []
    
    # エンジン間の実装格差をチェック
    if "calculation_implementations" in results:
        for engine, implementations in results["calculation_implementations"].items():
            if isinstance(implementations, dict):
                implemented_count = sum(1 for calc in implementations.values() 
                                      if isinstance(calc, dict) and calc.get("implemented", False))
                if implemented_count < 3:  # 6項目中3項目未満
                    implementation_gaps.append(f"{engine}: {implemented_count}/6項目のみ実装")
    
    # v3エンジンの問題
    v3_issues = []
    if "dssms_unified_output_engine_fixed_v3.py" in results.get("calculation_implementations", {}):
        v3_data = results["calculation_implementations"]["dssms_unified_output_engine_fixed_v3.py"]
        if isinstance(v3_data, dict):
            v3_implemented = sum(1 for calc in v3_data.values() 
                               if isinstance(calc, dict) and calc.get("implemented", False))
            if v3_implemented == 0:
                v3_issues.append("v3エンジンで統計計算が完全未実装")
    
    # 統計データソース問題
    data_source_issues = []
    if "data_sources" in results:
        backtester_data = results["data_sources"].get("backtester_analysis", {})
        if backtester_data.get("statistics_methods_found", 0) < 2:
            data_source_issues.append("DSSMSBacktesterに統計メソッドが不足")
    
    results["problems_identified"].extend(implementation_gaps)
    results["problems_identified"].extend(v3_issues)
    results["problems_identified"].extend(data_source_issues)
    
    # 推奨事項の生成
    print("\n💡 推奨事項:")
    recommendations = [
        "v3エンジンの統計計算実装（現在完全未実装）",
        "DSSMSBacktester.get_strategy_statistics()メソッド実装",
        "trade_historyからの統計計算ロジック統一",
        "Excel戦略別統計シートの自動計算機能追加",
        "switch_historyベースの統計計算代替実装"
    ]
    
    for i, rec in enumerate(recommendations, 1):
        print(f"{i}. {rec}")
        results["recommendations"].append(rec)
    
    print("\n❌ 特定された問題:")
    for problem in results["problems_identified"]:
        print(f"   {problem}")
    
    # 結果をJSONファイルに保存
    output_file = f"task_4_1_results_{results['analysis_timestamp']}.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 分析結果を保存: {output_file}")
    print("=" * 60)
    print("🔍 Task 4.1: 戦略別統計未計算問題の特定完了")
    
    return results

if __name__ == "__main__":
    try:
        results = analyze_strategy_statistics_data_source()
        
        # 重要な発見事項を表示
        print("\n🎯 重要な発見事項:")
        for problem in results["problems_identified"]:
            print(f"❌ {problem}")
            
    except Exception as e:
        print(f"❌ Task 4.1実行エラー: {e}")
        traceback.print_exc()