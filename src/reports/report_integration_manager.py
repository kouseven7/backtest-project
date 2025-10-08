"""
レポート統合管理システム
DSSMS Phase 3 Task 3.1: レポート生成システム改良

既存のsimple_excel_exporter.pyとの統合、空レポート問題修正、
エラー診断機能の統合を行います。
"""

import logging
import pandas as pd
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import sys
import traceback
from dataclasses import dataclass, asdict

# プロジェクトルートをパスに追加
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from config.logger_config import setup_logger
from src.reports.error_diagnostic_reporter import ErrorDiagnosticReporter, ErrorDiagnostic, SystemHealthReport

# 既存のExcel出力システムをインポート
try:
    from output.simple_excel_exporter import SimpleExcelExporter, save_backtest_results_simple
    EXCEL_EXPORTER_AVAILABLE = True
except ImportError as e:
    logger = setup_logger(__name__)
    logger.warning(f"simple_excel_exporterのインポートに失敗: {e}")
    EXCEL_EXPORTER_AVAILABLE = False

logger = setup_logger(__name__)

@dataclass
class ReportGenerationRequest:
    """レポート生成リクエストのデータクラス"""
    report_type: str  # EXCEL, HTML, COMBINED
    include_error_diagnostics: bool
    include_performance_metrics: bool
    include_risk_analysis: bool
    output_directory: str
    custom_filename: Optional[str] = None
    analysis_period_hours: int = 24
    emergency_mode: bool = False

@dataclass
class ReportGenerationResult:
    """レポート生成結果のデータクラス"""
    success: bool
    generated_files: List[str]
    error_message: Optional[str]
    warnings: List[str]
    execution_time_seconds: float
    report_summary: Dict[str, Any]

class ReportIntegrationManager:
    """レポート統合管理クラス"""
    
    def __init__(self, config_path: Optional[str] = None):
        """
        初期化
        
        Parameters:
            config_path (str): 設定ファイルパス（オプション）
        """
        self.config_path = config_path
        self.config = self._load_config()
        self.error_reporter = ErrorDiagnosticReporter()
        self.excel_exporter = SimpleExcelExporter() if EXCEL_EXPORTER_AVAILABLE else None
        
        # 出力ディレクトリの確保
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
        
    def _load_config(self) -> Dict[str, Any]:
        """設定ファイルを読み込み"""
        default_config = {
            "default_output_directory": "output",
            "enable_error_diagnostics": True,
            "enable_performance_metrics": True,
            "enable_risk_analysis": True,
            "max_report_age_hours": 168,  # 1週間
            "emergency_mode_triggers": [
                "ZERO_SWITCHING_SUCCESS",
                "PORTFOLIO_VALUE_COLLAPSE", 
                "TOTAL_RETURN_FAILURE"
            ],
            "report_templates": {
                "excel_template": "templates/excel_report_template.xlsx",
                "html_template": "templates/html_report_template.html"
            }
        }
        
        if self.config_path and Path(self.config_path).exists():
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    user_config = json.load(f)
                default_config.update(user_config)
                logger.info(f"設定ファイルを読み込み: {self.config_path}")
            except Exception as e:
                logger.warning(f"設定ファイル読み込みエラー: {e}")
                
        return default_config
        
    def generate_comprehensive_report(self, request: ReportGenerationRequest) -> ReportGenerationResult:
        """
        包括的レポートを生成
        
        Parameters:
            request (ReportGenerationRequest): レポート生成リクエスト
            
        Returns:
            ReportGenerationResult: 生成結果
        """
        start_time = datetime.now()
        generated_files = []
        warnings = []
        
        try:
            logger.info(f"包括的レポート生成開始: {request.report_type}")
            
            # 緊急モード判定
            if self._should_activate_emergency_mode(request):
                request.emergency_mode = True
                warnings.append("緊急モードが有効化されました")
                
            # エラー診断の実行
            diagnostics = []
            health_report = None
            if request.include_error_diagnostics:
                diagnostics, health_report = self._perform_error_diagnostics(request)
                
            # パフォーマンス指標の収集
            performance_metrics = {}
            if request.include_performance_metrics:
                performance_metrics = self._collect_performance_metrics(request)
                
            # リスク分析の実行
            risk_analysis = {}
            if request.include_risk_analysis:
                risk_analysis = self._perform_risk_analysis(request)
                
            # レポート生成
            if request.report_type in ['EXCEL', 'COMBINED']:
                excel_file = self._generate_excel_report(
                    request, diagnostics, health_report, performance_metrics, risk_analysis
                )
                if excel_file:
                    generated_files.append(excel_file)
                    
            if request.report_type in ['HTML', 'COMBINED']:
                html_file = self._generate_html_report(
                    request, diagnostics, health_report, performance_metrics, risk_analysis
                )
                if html_file:
                    generated_files.append(html_file)
                    
            # サマリー情報の構築
            execution_time = (datetime.now() - start_time).total_seconds()
            report_summary = self._build_report_summary(
                diagnostics, health_report, performance_metrics, risk_analysis
            )
            
            logger.info(f"レポート生成完了: {len(generated_files)}ファイル, {execution_time:.2f}秒")
            
            return ReportGenerationResult(
                success=True,
                generated_files=generated_files,
                error_message=None,
                warnings=warnings,
                execution_time_seconds=execution_time,
                report_summary=report_summary
            )
            
        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds()
            error_msg = f"レポート生成エラー: {e}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            
            return ReportGenerationResult(
                success=False,
                generated_files=generated_files,
                error_message=error_msg,
                warnings=warnings,
                execution_time_seconds=execution_time,
                report_summary={}
            )
            
    def _should_activate_emergency_mode(self, request: ReportGenerationRequest) -> bool:
        """緊急モード有効化判定"""
        if request.emergency_mode:
            return True
            
        try:
            # 最近のログから緊急トリガーを検索
            log_entries = self._collect_recent_logs(hours_back=1)  # 直近1時間
            emergency_triggers = self.config.get("emergency_mode_triggers", [])
            
            for entry in log_entries:
                for trigger in emergency_triggers:
                    if trigger.lower() in entry.get('message', '').lower():
                        logger.warning(f"緊急モードトリガー検出: {trigger}")
                        return True
                        
        except Exception as e:
            logger.warning(f"緊急モード判定エラー: {e}")
            
        return False
        
    def _collect_recent_logs(self, hours_back: int = 1) -> List[Dict[str, Any]]:
        """最近のログを収集"""
        try:
            return self.error_reporter._collect_log_entries(hours_back)
        except Exception as e:
            logger.warning(f"ログ収集エラー: {e}")
            return []
            
    def _perform_error_diagnostics(self, request: ReportGenerationRequest) -> Tuple[List[ErrorDiagnostic], SystemHealthReport]:
        """エラー診断を実行"""
        try:
            logger.info("エラー診断を実行中...")
            return self.error_reporter.analyze_logs(hours_back=request.analysis_period_hours)
        except Exception as e:
            logger.error(f"エラー診断実行エラー: {e}")
            return [], self.error_reporter._create_emergency_health_report()
            
    def _collect_performance_metrics(self, request: ReportGenerationRequest) -> Dict[str, Any]:
        """パフォーマンス指標を収集"""
        try:
            logger.info("パフォーマンス指標を収集中...")
            
            # 基本的なシステム指標
            metrics = {
                "collection_timestamp": datetime.now().isoformat(),
                "analysis_period_hours": request.analysis_period_hours,
                "system_uptime": self._calculate_system_uptime(),
                "log_file_count": len(list(Path("logs").glob("*.log"))),
                "recent_error_rate": self._calculate_recent_error_rate(),
                "emergency_mode": request.emergency_mode
            }
            
            # DSSMS固有の指標
            dssms_metrics = self._collect_dssms_metrics()
            metrics.update(dssms_metrics)
            
            return metrics
            
        except Exception as e:
            logger.error(f"パフォーマンス指標収集エラー: {e}")
            return {
                "collection_timestamp": datetime.now().isoformat(),
                "error_message": str(e),
                "emergency_mode": request.emergency_mode
            }
            
    def _collect_dssms_metrics(self) -> Dict[str, Any]:
        """DSSMS固有のメトリクスを収集"""
        try:
            # ログから関連情報を抽出
            log_entries = self._collect_recent_logs(hours_back=24)
            
            switching_success_entries = [
                e for e in log_entries 
                if 'switching success rate' in e.get('message', '').lower()
            ]
            
            portfolio_value_entries = [
                e for e in log_entries 
                if 'portfolio value' in e.get('message', '').lower()
            ]
            
            return {
                "switching_events_detected": len(switching_success_entries),
                "portfolio_calculations_detected": len(portfolio_value_entries),
                "dssms_log_entries": len([
                    e for e in log_entries 
                    if 'dssms' in e.get('message', '').lower()
                ])
            }
            
        except Exception as e:
            logger.warning(f"DSSMS指標収集エラー: {e}")
            return {"dssms_metrics_error": str(e)}
            
    def _calculate_system_uptime(self) -> str:
        """システム稼働時間を計算"""
        try:
            # 最も古いログファイルの作成時刻から推定
            log_files = list(Path("logs").glob("*.log"))
            if log_files:
                oldest_log = min(log_files, key=lambda x: x.stat().st_ctime)
                uptime = datetime.now() - datetime.fromtimestamp(oldest_log.stat().st_ctime)
                return str(uptime)
            return "不明"
        except Exception:
            return "計算エラー"
            
    def _calculate_recent_error_rate(self) -> float:
        """最近のエラー率を計算"""
        try:
            recent_logs = self._collect_recent_logs(hours_back=1)
            if not recent_logs:
                return 0.0
                
            error_logs = [
                e for e in recent_logs 
                if e.get('level', '').upper() in ['ERROR', 'CRITICAL', 'FATAL']
            ]
            
            return len(error_logs) / len(recent_logs) * 100.0
            
        except Exception:
            return -1.0  # エラー率計算不可
            
    def _perform_risk_analysis(self, request: ReportGenerationRequest) -> Dict[str, Any]:
        """リスク分析を実行"""
        try:
            logger.info("リスク分析を実行中...")
            
            # 基本的なリスク指標
            risk_metrics = {
                "analysis_timestamp": datetime.now().isoformat(),
                "critical_risk_level": "UNKNOWN",
                "identified_risks": [],
                "mitigation_recommendations": []
            }
            
            # ログからリスク指標を抽出
            log_entries = self._collect_recent_logs(hours_back=request.analysis_period_hours)
            
            # 重大なリスク要因の検出
            critical_patterns = [
                ("portfolio value.*0\\.01", "PORTFOLIO_COLLAPSE"),
                ("total return.*-100%", "TOTAL_LOSS"),
                ("switching success rate.*0%", "SWITCHING_FAILURE"),
                ("max drawdown.*100%", "MAXIMUM_DRAWDOWN")
            ]
            
            detected_risks = []
            for pattern, risk_type in critical_patterns:
                import re
                matching_entries = [
                    e for e in log_entries 
                    if re.search(pattern, e.get('message', ''), re.IGNORECASE)
                ]
                if matching_entries:
                    detected_risks.append({
                        "risk_type": risk_type,
                        "frequency": len(matching_entries),
                        "last_occurrence": max(e['timestamp'] for e in matching_entries).isoformat()
                    })
                    
            risk_metrics["identified_risks"] = detected_risks
            
            # リスクレベル判定
            if any(risk["risk_type"] in ["PORTFOLIO_COLLAPSE", "TOTAL_LOSS"] for risk in detected_risks):
                risk_metrics["critical_risk_level"] = "CRITICAL"
            elif any(risk["risk_type"] == "SWITCHING_FAILURE" for risk in detected_risks):
                risk_metrics["critical_risk_level"] = "HIGH"
            elif detected_risks:
                risk_metrics["critical_risk_level"] = "MEDIUM"
            else:
                risk_metrics["critical_risk_level"] = "LOW"
                
            # 軽減策の推奨
            if detected_risks:
                risk_metrics["mitigation_recommendations"] = [
                    "緊急診断システムの実行",
                    "ポートフォリオ計算エンジンの検証",
                    "戦略切り替えロジックの修復",
                    "リスク管理パラメータの再調整"
                ]
            else:
                risk_metrics["mitigation_recommendations"] = [
                    "定期的な監視の継続",
                    "予防的メンテナンスの実施"
                ]
                
            return risk_metrics
            
        except Exception as e:
            logger.error(f"リスク分析エラー: {e}")
            return {
                "analysis_timestamp": datetime.now().isoformat(),
                "critical_risk_level": "ANALYSIS_ERROR",
                "error_message": str(e)
            }
            
    def _generate_excel_report(self, request: ReportGenerationRequest, 
                             diagnostics: List[ErrorDiagnostic],
                             health_report: Optional[SystemHealthReport],
                             performance_metrics: Dict[str, Any],
                             risk_analysis: Dict[str, Any]) -> Optional[str]:
        """Excel形式のレポートを生成"""
        if not EXCEL_EXPORTER_AVAILABLE:
            logger.warning("Excel出力機能が利用できません")
            return None
            
        try:
            logger.info("Excelレポートを生成中...")
            
            # ファイル名の生成
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = request.custom_filename or f"dssms_comprehensive_report_{timestamp}.xlsx"
            output_path = Path(request.output_directory) / filename
            
            # Excel workbookを作成
            try:
                # openpyxl遅延インポート (TODO-PERF-001: Stage 3)
import src.utils.openpyxl_lazy_wrapper as openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: wb = openpyxl.Workbook()
                
                # サマリーシート
                ws_summary = wb.active
                ws_summary.title = "サマリー"
                self._create_summary_sheet(ws_summary, health_report, performance_metrics, risk_analysis)
                
                # エラー診断シート
                if diagnostics:
                    ws_errors = wb.create_sheet("エラー診断")
                    self._create_error_diagnostics_sheet(ws_errors, diagnostics)
                    
                # パフォーマンス指標シート
                if performance_metrics:
                    ws_performance = wb.create_sheet("パフォーマンス")
                    self._create_performance_sheet(ws_performance, performance_metrics)
                    
                # リスク分析シート
                if risk_analysis:
                    ws_risk = wb.create_sheet("リスク分析")
                    self._create_risk_analysis_sheet(ws_risk, risk_analysis)
                    
                # ファイルを保存
                wb.save(output_path)
                logger.info(f"Excelレポートを保存: {output_path}")
                return str(output_path)
                
            except ImportError:
                logger.warning("openpyxlが利用できません。簡易Excel出力を試行します。")
                return self._generate_simple_excel_report(request, diagnostics, health_report, performance_metrics, risk_analysis)
                
        except Exception as e:
            logger.error(f"Excelレポート生成エラー: {e}")
            return None
            
    def _generate_simple_excel_report(self, request: ReportGenerationRequest,
                                    diagnostics: List[ErrorDiagnostic],
                                    health_report: Optional[SystemHealthReport],
                                    performance_metrics: Dict[str, Any],
                                    risk_analysis: Dict[str, Any]) -> Optional[str]:
        """簡易Excel形式のレポートを生成（既存システム活用）"""
        try:
            if not self.excel_exporter:
                return None
                
            # ダミーデータでExcel出力をテスト
            dummy_data = pd.DataFrame({
                'Date': [datetime.now()],
                'Close': [100.0],
                'Entry_Signal': [0],
                'Exit_Signal': [0]
            })
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"dssms_simple_report_{timestamp}.xlsx"
            output_path = str(Path(request.output_directory) / filename)
            
            # 簡易レポートデータを追加
            report_data = {
                'diagnostics_count': len(diagnostics),
                'health_score': health_report.overall_health_score if health_report else 0.0,
                'performance_metrics': performance_metrics,
                'risk_level': risk_analysis.get('critical_risk_level', 'UNKNOWN')
            }
            
            result_path = self.excel_exporter.export_backtest_results(
                dummy_data, "DSSMS_REPORT", request.output_directory, filename
            )
            
            logger.info(f"簡易Excelレポートを生成: {result_path}")
            return result_path
            
        except Exception as e:
            logger.error(f"簡易Excelレポート生成エラー: {e}")
            return None
            
    def _generate_html_report(self, request: ReportGenerationRequest,
                            diagnostics: List[ErrorDiagnostic],
                            health_report: Optional[SystemHealthReport],
                            performance_metrics: Dict[str, Any],
                            risk_analysis: Dict[str, Any]) -> Optional[str]:
        """HTML形式のレポートを生成"""
        try:
            logger.info("HTMLレポートを生成中...")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = request.custom_filename or f"dssms_comprehensive_report_{timestamp}.html"
            output_path = Path(request.output_directory) / filename
            
            # HTMLコンテンツを生成
            html_content = self._build_comprehensive_html(
                diagnostics, health_report, performance_metrics, risk_analysis, request
            )
            
            # ファイルに出力
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
            logger.info(f"HTMLレポートを保存: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"HTMLレポート生成エラー: {e}")
            return None
            
    def _build_comprehensive_html(self, diagnostics: List[ErrorDiagnostic],
                                health_report: Optional[SystemHealthReport],
                                performance_metrics: Dict[str, Any],
                                risk_analysis: Dict[str, Any],
                                request: ReportGenerationRequest) -> str:
        """包括的HTMLコンテンツを構築"""
        
        # 基本HTMLテンプレート
        html_template = """
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DSSMS 包括的システムレポート</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; margin: -30px -30px 30px -30px; border-radius: 10px 10px 0 0; }}
        .emergency {{ background: linear-gradient(135deg, #ff416c 0%, #ff4b2b 100%); }}
        .section {{ margin: 30px 0; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; }}
        .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin: 20px 0; }}
        .metric-card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; }}
        .health-score {{ font-size: 48px; font-weight: bold; margin: 10px 0; }}
        .critical {{ color: #dc3545; }}
        .high {{ color: #fd7e14; }}
        .medium {{ color: #ffc107; }}
        .low {{ color: #28a745; }}
        .error-list {{ max-height: 400px; overflow-y: auto; }}
        .error-item {{ border-left: 4px solid #dc3545; padding: 15px; margin: 10px 0; background: #f8f9fa; }}
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #f8f9fa; font-weight: bold; }}
        .timestamp {{ color: #6c757d; font-size: 0.9em; }}
        .recommendations {{ background: #e7f3ff; border-left: 4px solid #007bff; padding: 20px; margin: 20px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header {emergency_class}">
            <h1>DSSMS 包括的システムレポート</h1>
            <p class="timestamp">生成日時: {timestamp}</p>
            {emergency_notice}
        </div>

        <!-- システム概要セクション -->
        <div class="section">
            <h2>システム概要</h2>
            <div class="metric-grid">
                <div class="metric-card">
                    <div class="health-score {health_class}">{health_score:.1f}</div>
                    <div>システム健康度</div>
                </div>
                <div class="metric-card">
                    <div class="health-score {availability_class}">{availability:.1f}%</div>
                    <div>システム可用性</div>
                </div>
                <div class="metric-card">
                    <div class="health-score {risk_class}">{risk_level}</div>
                    <div>リスクレベル</div>
                </div>
                <div class="metric-card">
                    <div class="health-score">{error_count}</div>
                    <div>検出エラー数</div>
                </div>
            </div>
        </div>

        <!-- エラー診断セクション -->
        {error_section}

        <!-- パフォーマンス指標セクション -->
        {performance_section}

        <!-- リスク分析セクション -->
        {risk_section}

        <!-- 推奨事項セクション -->
        {recommendations_section}
    </div>
</body>
</html>
        """
        
        # データの準備
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        emergency_class = "emergency" if request.emergency_mode else ""
        emergency_notice = "<h3>[WARNING] 緊急モード有効 [WARNING]</h3>" if request.emergency_mode else ""
        
        health_score = health_report.overall_health_score if health_report else 0.0
        availability = health_report.system_availability if health_report else 0.0
        health_class = self._get_severity_class(health_score, [30, 60, 80])
        availability_class = self._get_severity_class(availability, [50, 75, 90])
        
        risk_level = risk_analysis.get('critical_risk_level', 'UNKNOWN')
        risk_class = self._get_risk_class(risk_level)
        
        error_count = len(diagnostics)
        
        # セクションの構築
        error_section = self._build_error_section(diagnostics)
        performance_section = self._build_performance_section(performance_metrics)
        risk_section = self._build_risk_section(risk_analysis)
        recommendations_section = self._build_recommendations_section(health_report, diagnostics, risk_analysis)
        
        return html_template.format(
            timestamp=timestamp,
            emergency_class=emergency_class,
            emergency_notice=emergency_notice,
            health_score=health_score,
            health_class=health_class,
            availability=availability,
            availability_class=availability_class,
            risk_level=risk_level,
            risk_class=risk_class,
            error_count=error_count,
            error_section=error_section,
            performance_section=performance_section,
            risk_section=risk_section,
            recommendations_section=recommendations_section
        )
        
    def _get_severity_class(self, value: float, thresholds: List[float]) -> str:
        """数値から重要度クラスを取得"""
        if value >= thresholds[2]:
            return "low"
        elif value >= thresholds[1]:
            return "medium" 
        elif value >= thresholds[0]:
            return "high"
        else:
            return "critical"
            
    def _get_risk_class(self, risk_level: str) -> str:
        """リスクレベルからクラスを取得"""
        mapping = {
            "CRITICAL": "critical",
            "HIGH": "high", 
            "MEDIUM": "medium",
            "LOW": "low"
        }
        return mapping.get(risk_level, "medium")
        
    def _build_error_section(self, diagnostics: List[ErrorDiagnostic]) -> str:
        """エラーセクションを構築"""
        if not diagnostics:
            return """
            <div class="section">
                <h2>エラー診断</h2>
                <p style="color: #28a745; font-weight: bold;">[OK] エラーは検出されませんでした</p>
            </div>
            """
            
        error_items = ""
        for diag in diagnostics[:10]:  # 最大10件表示
            severity_class = diag.severity.lower()
            error_items += f"""
            <div class="error-item">
                <h4 class="{severity_class}">{diag.error_type} ({diag.severity})</h4>
                <p><strong>発生頻度:</strong> {diag.frequency}回</p>
                <p><strong>根本原因:</strong> {diag.root_cause}</p>
                <p><strong>修復提案:</strong> {'; '.join(diag.suggested_fixes[:2])}</p>
            </div>
            """
            
        return f"""
        <div class="section">
            <h2>エラー診断 ({len(diagnostics)}件)</h2>
            <div class="error-list">
                {error_items}
            </div>
        </div>
        """
        
    def _build_performance_section(self, performance_metrics: Dict[str, Any]) -> str:
        """パフォーマンスセクションを構築"""
        if not performance_metrics:
            return ""
            
        rows = ""
        for key, value in performance_metrics.items():
            if key not in ['collection_timestamp']:
                rows += f"<tr><td>{key}</td><td>{value}</td></tr>"
                
        return f"""
        <div class="section">
            <h2>パフォーマンス指標</h2>
            <table>
                <thead>
                    <tr><th>指標</th><th>値</th></tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
        </div>
        """
        
    def _build_risk_section(self, risk_analysis: Dict[str, Any]) -> str:
        """リスクセクションを構築"""
        if not risk_analysis:
            return ""
            
        risks = risk_analysis.get('identified_risks', [])
        risk_items = ""
        for risk in risks[:5]:  # 最大5件表示
            risk_items += f"""
            <tr>
                <td>{risk.get('risk_type', 'Unknown')}</td>
                <td>{risk.get('frequency', 0)}</td>
                <td>{risk.get('last_occurrence', 'Unknown')}</td>
            </tr>
            """
            
        return f"""
        <div class="section">
            <h2>リスク分析</h2>
            <p><strong>総合リスクレベル:</strong> <span class="{self._get_risk_class(risk_analysis.get('critical_risk_level', 'UNKNOWN'))}">{risk_analysis.get('critical_risk_level', 'UNKNOWN')}</span></p>
            {f'<table><thead><tr><th>リスク種別</th><th>頻度</th><th>最終発生</th></tr></thead><tbody>{risk_items}</tbody></table>' if risk_items else '<p>特定されたリスクはありません。</p>'}
        </div>
        """
        
    def _build_recommendations_section(self, health_report: Optional[SystemHealthReport],
                                     diagnostics: List[ErrorDiagnostic],
                                     risk_analysis: Dict[str, Any]) -> str:
        """推奨事項セクションを構築"""
        recommendations = []
        
        # 健康度レポートからの推奨事項
        if health_report and health_report.recommendations:
            recommendations.extend(health_report.recommendations)
            
        # リスク分析からの推奨事項
        risk_recommendations = risk_analysis.get('mitigation_recommendations', [])
        recommendations.extend(risk_recommendations)
        
        # 診断結果からの推奨事項
        for diag in diagnostics[:3]:  # 重要度上位3件
            recommendations.extend(diag.suggested_fixes[:1])  # 各診断から1つ
            
        # 重複除去
        unique_recommendations = list(dict.fromkeys(recommendations))
        
        if not unique_recommendations:
            unique_recommendations = ["定期的な監視を継続してください"]
            
        rec_items = "".join(f"<li>{rec}</li>" for rec in unique_recommendations[:10])
        
        return f"""
        <div class="recommendations">
            <h2>推奨事項</h2>
            <ul>
                {rec_items}
            </ul>
        </div>
        """
        
    def _create_summary_sheet(self, worksheet, health_report, performance_metrics, risk_analysis):
        """Excelサマリーシートを作成"""
        try:
            worksheet["A1"] = "DSSMS システムサマリー"
            worksheet["A2"] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            if health_report:
                worksheet["A4"] = "システム健康度"
                worksheet["B4"] = f"{health_report.overall_health_score:.1f}/100.0"
                worksheet["A5"] = "致命的問題数"
                worksheet["B5"] = health_report.critical_issues_count
                worksheet["A6"] = "高優先度問題数"
                worksheet["B6"] = health_report.high_priority_issues_count
                
        except Exception as e:
            logger.warning(f"Excelサマリーシート作成エラー: {e}")
            
    def _create_error_diagnostics_sheet(self, worksheet, diagnostics):
        """Excelエラー診断シートを作成"""
        try:
            headers = ["エラータイプ", "重要度", "頻度", "根本原因", "修復提案"]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
                
            for row, diag in enumerate(diagnostics, 2):
                worksheet.cell(row=row, column=1, value=diag.error_type)
                worksheet.cell(row=row, column=2, value=diag.severity)
                worksheet.cell(row=row, column=3, value=diag.frequency)
                worksheet.cell(row=row, column=4, value=diag.root_cause)
                worksheet.cell(row=row, column=5, value="; ".join(diag.suggested_fixes))
                
        except Exception as e:
            logger.warning(f"Excelエラー診断シート作成エラー: {e}")
            
    def _create_performance_sheet(self, worksheet, performance_metrics):
        """Excelパフォーマンスシートを作成"""
        try:
            worksheet.cell(row=1, column=1, value="指標")
            worksheet.cell(row=1, column=2, value="値")
            
            for row, (key, value) in enumerate(performance_metrics.items(), 2):
                worksheet.cell(row=row, column=1, value=key)
                worksheet.cell(row=row, column=2, value=str(value))
                
        except Exception as e:
            logger.warning(f"Excelパフォーマンスシート作成エラー: {e}")
            
    def _create_risk_analysis_sheet(self, worksheet, risk_analysis):
        """Excelリスク分析シートを作成"""
        try:
            worksheet.cell(row=1, column=1, value="リスク分析結果")
            worksheet.cell(row=2, column=1, value="総合リスクレベル")
            worksheet.cell(row=2, column=2, value=risk_analysis.get('critical_risk_level', 'UNKNOWN'))
            
            risks = risk_analysis.get('identified_risks', [])
            if risks:
                worksheet.cell(row=4, column=1, value="特定されたリスク")
                headers = ["リスク種別", "頻度", "最終発生"]
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=5, column=col, value=header)
                    
                for row, risk in enumerate(risks, 6):
                    worksheet.cell(row=row, column=1, value=risk.get('risk_type', ''))
                    worksheet.cell(row=row, column=2, value=risk.get('frequency', 0))
                    worksheet.cell(row=row, column=3, value=risk.get('last_occurrence', ''))
                    
        except Exception as e:
            logger.warning(f"Excelリスク分析シート作成エラー: {e}")
            
    def _build_report_summary(self, diagnostics: List[ErrorDiagnostic],
                            health_report: Optional[SystemHealthReport],
                            performance_metrics: Dict[str, Any],
                            risk_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """レポートサマリーを構築"""
        return {
            "generation_timestamp": datetime.now().isoformat(),
            "total_errors_detected": len(diagnostics),
            "critical_errors": len([d for d in diagnostics if d.severity == 'CRITICAL']),
            "system_health_score": health_report.overall_health_score if health_report else 0.0,
            "risk_level": risk_analysis.get('critical_risk_level', 'UNKNOWN'),
            "emergency_mode": any('emergency' in str(v).lower() for v in performance_metrics.values()),
            "key_issues": [d.error_type for d in diagnostics[:5]]
        }

if __name__ == "__main__":
    # テスト実行
    manager = ReportIntegrationManager()
    
    request = ReportGenerationRequest(
        report_type="COMBINED",
        include_error_diagnostics=True,
        include_performance_metrics=True,
        include_risk_analysis=True,
        output_directory="output",
        analysis_period_hours=24
    )
    
    result = manager.generate_comprehensive_report(request)
    
    print(f"レポート生成結果: {'成功' if result.success else '失敗'}")
    print(f"生成ファイル数: {len(result.generated_files)}")
    print(f"実行時間: {result.execution_time_seconds:.2f}秒")
    
    if result.generated_files:
        print("生成されたファイル:")
        for file in result.generated_files:
            print(f"  - {file}")
            
    if result.warnings:
        print("警告:")
        for warning in result.warnings:
            print(f"  - {warning}")
