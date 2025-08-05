"""
戦略比較レポート生成システム
フェーズ2：多形式レポート出力（Excel/HTML/JSON）

戦略パフォーマンス比較、市場環境別分析、
インタラクティブダッシュボード、統合レポートを生成します。
"""

import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple, Union
import sys
import os
import json
import warnings
from dataclasses import dataclass, asdict
from collections import defaultdict
import base64
from io import BytesIO

# プロジェクトルートをパスに追加
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from src.config.logger_config import setup_logger

# Excel操作用
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxlが利用できません。Excel出力機能が制限されます。")

# 可視化用
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    import seaborn as sns
    plt.style.use('seaborn-v0_8')
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    try:
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        import seaborn as sns
        plt.style.use('seaborn')
        MATPLOTLIB_AVAILABLE = True
    except ImportError:
        MATPLOTLIB_AVAILABLE = False
        logging.warning("Matplotlibが利用できません。一部の可視化機能が制限されます。")

# インタラクティブ可視化用
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    import plotly.offline as pyo
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    logging.warning("Plotlyが利用できません。インタラクティブ可視化機能が制限されます。")

@dataclass
class ReportConfig:
    """レポート設定"""
    output_formats: List[str] = None  # ['excel', 'html', 'json', 'pdf']
    include_charts: bool = True
    include_interactive_dashboard: bool = True
    include_statistical_analysis: bool = True
    include_market_breakdown: bool = True
    chart_style: str = "professional"  # professional, minimal, colorful
    language: str = "ja"  # ja, en
    output_directory: str = "output/reports"
    
    def __post_init__(self):
        if self.output_formats is None:
            self.output_formats = ['excel', 'html', 'json']

@dataclass
class ReportSection:
    """レポートセクション"""
    title: str
    content: Any
    section_type: str  # table, chart, text, dashboard
    metadata: Dict[str, Any] = None

class StrategyComparisonReporter:
    """戦略比較レポート生成システム"""
    
    def __init__(self, config: ReportConfig = None):
        """
        初期化
        
        Args:
            config: レポート設定
        """
        self.config = config or ReportConfig()
        self.logger = setup_logger(__name__)
        
        # 出力ディレクトリの作成
        self.output_dir = Path(self.config.output_directory)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # レポートセクション
        self.report_sections = []
        
        # スタイル設定
        self._setup_styles()
        
        self.logger.info("戦略比較レポート生成システム初期化完了")

    def _setup_styles(self):
        """スタイル設定"""
        # Excel用スタイル
        self.excel_styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'highlight_good': {
                'fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            },
            'highlight_bad': {
                'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            }
        }
        
        # プロット用カラーパレット
        if self.config.chart_style == "professional":
            self.color_palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f']
        elif self.config.chart_style == "colorful":
            self.color_palette = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', '#FF9FF3', '#54A0FF', '#5F27CD']
        else:  # minimal
            self.color_palette = ['#2C3E50', '#34495E', '#7F8C8D', '#95A5A6', '#BDC3C7', '#ECF0F1', '#E74C3C', '#3498DB']

    def generate_comprehensive_report(self, aggregated_results: Dict[str, Any]) -> Dict[str, str]:
        """
        統合レポートの生成
        
        Args:
            aggregated_results: パフォーマンス集約結果
            
        Returns:
            生成されたレポートファイルのパス辞書
        """
        self.logger.info("統合レポート生成開始")
        
        try:
            # レポートセクションの生成
            self._generate_report_sections(aggregated_results)
            
            generated_files = {}
            
            # Excel形式でのレポート生成
            if 'excel' in self.config.output_formats and OPENPYXL_AVAILABLE:
                excel_path = self._generate_excel_report(aggregated_results)
                generated_files['excel'] = str(excel_path)
            
            # HTML形式でのレポート生成
            if 'html' in self.config.output_formats:
                html_path = self._generate_html_report(aggregated_results)
                generated_files['html'] = str(html_path)
            
            # JSON形式でのレポート生成
            if 'json' in self.config.output_formats:
                json_path = self._generate_json_report(aggregated_results)
                generated_files['json'] = str(json_path)
            
            # インタラクティブダッシュボードの生成
            if self.config.include_interactive_dashboard and PLOTLY_AVAILABLE:
                dashboard_path = self._generate_interactive_dashboard(aggregated_results)
                generated_files['dashboard'] = str(dashboard_path)
            
            self.logger.info(f"統合レポート生成完了: {len(generated_files)}形式")
            return generated_files
            
        except Exception as e:
            self.logger.error(f"統合レポート生成失敗: {e}")
            raise

    def _generate_report_sections(self, aggregated_results: Dict[str, Any]):
        """レポートセクションの生成"""
        try:
            # 1. エグゼクティブサマリー
            self._add_executive_summary(aggregated_results)
            
            # 2. 戦略パフォーマンス概要
            self._add_strategy_performance_overview(aggregated_results)
            
            # 3. 市場環境別分析
            if self.config.include_market_breakdown:
                self._add_market_environment_analysis(aggregated_results)
            
            # 4. 統計的分析
            if self.config.include_statistical_analysis:
                self._add_statistical_analysis(aggregated_results)
            
            # 5. 相関分析
            self._add_correlation_analysis(aggregated_results)
            
            # 6. パフォーマンスランキング
            self._add_performance_rankings(aggregated_results)
            
            # 7. 時系列分析
            self._add_time_series_analysis(aggregated_results)
            
            # 8. リスク分析
            self._add_risk_analysis(aggregated_results)
            
            # 9. 推奨事項
            self._add_recommendations(aggregated_results)
            
            self.logger.info(f"レポートセクション生成完了: {len(self.report_sections)}セクション")
            
        except Exception as e:
            self.logger.error(f"レポートセクション生成失敗: {e}")

    def _add_executive_summary(self, results: Dict[str, Any]):
        """エグゼクティブサマリーの追加"""
        try:
            summary = results.get('summary', {})
            rankings = results.get('performance_rankings', {})
            
            content = {
                'analysis_overview': {
                    'total_results': summary.get('total_results', 0),
                    'strategies_analyzed': summary.get('strategies_analyzed', 0),
                    'symbols_analyzed': summary.get('symbols_analyzed', 0),
                    'analysis_date': summary.get('aggregation_timestamp', datetime.now().isoformat())
                },
                'key_findings': [],
                'top_strategies': list(rankings.get('overall', {}).keys())[:5] if rankings.get('overall') else []
            }
            
            # 主要な発見事項の生成
            if rankings.get('overall'):
                best_strategy = list(rankings['overall'].keys())[0]
                content['key_findings'].append(f"最高パフォーマンス戦略: {best_strategy}")
            
            if 'strategy_market_performance' in results:
                market_data = results['strategy_market_performance']
                if market_data:
                    content['key_findings'].append(f"分析対象戦略数: {len(market_data)}")
            
            self.report_sections.append(ReportSection(
                title="エグゼクティブサマリー",
                content=content,
                section_type="text"
            ))
            
        except Exception as e:
            self.logger.warning(f"エグゼクティブサマリー生成失敗: {e}")

    def _add_strategy_performance_overview(self, results: Dict[str, Any]):
        """戦略パフォーマンス概要の追加"""
        try:
            strategy_market_performance = results.get('strategy_market_performance', {})
            
            if not strategy_market_performance:
                return
            
            # 戦略別サマリーテーブルの作成
            summary_data = []
            
            for strategy, market_data in strategy_market_performance.items():
                strategy_summary = {'Strategy': strategy}
                
                # 全市場環境での平均パフォーマンス
                all_metrics = defaultdict(list)
                for market_state, metrics in market_data.items():
                    for metric_name, metric_data in metrics.items():
                        if isinstance(metric_data, dict) and 'mean' in metric_data:
                            all_metrics[metric_name].append(metric_data['mean'])
                
                # 平均値の計算
                for metric_name, values in all_metrics.items():
                    if values:
                        strategy_summary[f'Avg_{metric_name}'] = np.mean(values)
                
                summary_data.append(strategy_summary)
            
            summary_df = pd.DataFrame(summary_data)
            
            self.report_sections.append(ReportSection(
                title="戦略パフォーマンス概要",
                content=summary_df,
                section_type="table"
            ))
            
        except Exception as e:
            self.logger.warning(f"戦略パフォーマンス概要生成失敗: {e}")

    def _add_market_environment_analysis(self, results: Dict[str, Any]):
        """市場環境別分析の追加"""
        try:
            strategy_market_performance = results.get('strategy_market_performance', {})
            
            if not strategy_market_performance:
                return
            
            # 市場環境別のパフォーマンス比較テーブル
            market_analysis = {}
            
            # 全ての市場環境を取得
            all_markets = set()
            for strategy_data in strategy_market_performance.values():
                all_markets.update(strategy_data.keys())
            
            for market_state in all_markets:
                market_performance = []
                
                for strategy, market_data in strategy_market_performance.items():
                    if market_state in market_data:
                        metrics = market_data[market_state]
                        strategy_metrics = {'Strategy': strategy}
                        
                        for metric_name, metric_data in metrics.items():
                            if isinstance(metric_data, dict) and 'mean' in metric_data:
                                strategy_metrics[metric_name] = metric_data['mean']
                        
                        market_performance.append(strategy_metrics)
                
                if market_performance:
                    market_analysis[market_state] = pd.DataFrame(market_performance)
            
            self.report_sections.append(ReportSection(
                title="市場環境別分析",
                content=market_analysis,
                section_type="table"
            ))
            
        except Exception as e:
            self.logger.warning(f"市場環境別分析生成失敗: {e}")

    def _add_statistical_analysis(self, results: Dict[str, Any]):
        """統計的分析の追加"""
        try:
            statistical_analysis = results.get('statistical_analysis', {})
            
            if not statistical_analysis:
                return
            
            # 統計テスト結果のサマリー
            stats_summary = {}
            
            # 戦略間比較テスト
            if 'strategy_tests' in statistical_analysis:
                strategy_tests = statistical_analysis['strategy_tests']
                
                if 'pairwise_t_tests' in strategy_tests:
                    significant_pairs = []
                    for test in strategy_tests['pairwise_t_tests']:
                        if test.get('significant', False):
                            significant_pairs.append({
                                'Strategy1': test['strategy1'],
                                'Strategy2': test['strategy2'],
                                'Better_Strategy': test['better_strategy'],
                                'P_Value': test['p_value']
                            })
                    
                    if significant_pairs:
                        stats_summary['significant_strategy_differences'] = pd.DataFrame(significant_pairs)
                
                # 正規性テスト
                if 'normality_tests' in strategy_tests:
                    normality_results = []
                    for strategy, test_result in strategy_tests['normality_tests'].items():
                        normality_results.append({
                            'Strategy': strategy,
                            'Shapiro_P_Value': test_result['shapiro_p_value'],
                            'Is_Normal': test_result['is_normal']
                        })
                    
                    if normality_results:
                        stats_summary['normality_tests'] = pd.DataFrame(normality_results)
            
            # 市場環境テスト
            if 'market_tests' in statistical_analysis:
                market_tests = statistical_analysis['market_tests']
                if 'anova_test' in market_tests:
                    anova_result = market_tests['anova_test']
                    stats_summary['market_environment_anova'] = {
                        'F_Statistic': anova_result['f_statistic'],
                        'P_Value': anova_result['p_value'],
                        'Significant': anova_result['significant']
                    }
            
            if stats_summary:
                self.report_sections.append(ReportSection(
                    title="統計的分析",
                    content=stats_summary,
                    section_type="table"
                ))
            
        except Exception as e:
            self.logger.warning(f"統計的分析生成失敗: {e}")

    def _add_correlation_analysis(self, results: Dict[str, Any]):
        """相関分析の追加"""
        try:
            correlation_analysis = results.get('correlation_analysis', {})
            
            if not correlation_analysis:
                return
            
            correlation_content = {}
            
            # 戦略間相関
            if 'strategy_correlation_matrix' in correlation_analysis:
                correlation_data = correlation_analysis['strategy_correlation_matrix']
                if 'matrix' in correlation_data and 'strategies' in correlation_data:
                    correlation_df = pd.DataFrame(
                        correlation_data['matrix'],
                        index=correlation_data['strategies'],
                        columns=correlation_data['strategies']
                    )
                    correlation_content['strategy_correlation_matrix'] = correlation_df
            
            # 高相関ペア
            if 'high_correlation_pairs' in correlation_analysis:
                high_pairs = correlation_analysis['high_correlation_pairs']
                if high_pairs:
                    pairs_df = pd.DataFrame(high_pairs)
                    correlation_content['high_correlation_pairs'] = pairs_df
            
            if correlation_content:
                self.report_sections.append(ReportSection(
                    title="相関分析",
                    content=correlation_content,
                    section_type="table"
                ))
            
        except Exception as e:
            self.logger.warning(f"相関分析生成失敗: {e}")

    def _add_performance_rankings(self, results: Dict[str, Any]):
        """パフォーマンスランキングの追加"""
        try:
            rankings = results.get('performance_rankings', {})
            
            if not rankings:
                return
            
            ranking_content = {}
            
            # 総合ランキング
            if 'overall' in rankings:
                overall_ranking = []
                for rank, (strategy, score) in enumerate(rankings['overall'].items(), 1):
                    overall_ranking.append({
                        'Rank': rank,
                        'Strategy': strategy,
                        'Score': round(score, 4)
                    })
                ranking_content['overall_ranking'] = pd.DataFrame(overall_ranking)
            
            # 市場環境別ランキング
            if 'by_market_environment' in rankings:
                market_rankings = {}
                for market, market_ranking in rankings['by_market_environment'].items():
                    market_ranking_list = []
                    for rank, (strategy, score) in enumerate(market_ranking.items(), 1):
                        market_ranking_list.append({
                            'Rank': rank,
                            'Strategy': strategy,
                            'Score': round(score, 4)
                        })
                    market_rankings[market] = pd.DataFrame(market_ranking_list)
                
                ranking_content['market_environment_rankings'] = market_rankings
            
            # メトリクス別ランキング
            if 'by_metric' in rankings:
                metric_rankings = {}
                for metric, metric_ranking in rankings['by_metric'].items():
                    metric_ranking_list = []
                    for rank, (strategy, score) in enumerate(metric_ranking.items(), 1):
                        metric_ranking_list.append({
                            'Rank': rank,
                            'Strategy': strategy,
                            'Score': round(score, 4)
                        })
                    metric_rankings[metric] = pd.DataFrame(metric_ranking_list)
                
                ranking_content['metric_rankings'] = metric_rankings
            
            if ranking_content:
                self.report_sections.append(ReportSection(
                    title="パフォーマンスランキング",
                    content=ranking_content,
                    section_type="table"
                ))
            
        except Exception as e:
            self.logger.warning(f"パフォーマンスランキング生成失敗: {e}")

    def _add_time_series_analysis(self, results: Dict[str, Any]):
        """時系列分析の追加"""
        try:
            time_series_analysis = results.get('time_series_analysis', {})
            
            if not time_series_analysis:
                return
            
            time_series_content = {}
            
            # 時系列データ
            if 'time_series_data' in time_series_analysis:
                ts_data = time_series_analysis['time_series_data']
                
                # 時系列パフォーマンステーブルの作成
                time_series_table = []
                for time_period, strategy_data in ts_data.items():
                    for strategy, metrics in strategy_data.items():
                        for metric_name, metric_data in metrics.items():
                            if isinstance(metric_data, dict) and 'mean' in metric_data:
                                time_series_table.append({
                                    'Time_Period': time_period,
                                    'Strategy': strategy,
                                    'Metric': metric_name,
                                    'Value': metric_data['mean']
                                })
                
                if time_series_table:
                    time_series_content['time_series_performance'] = pd.DataFrame(time_series_table)
            
            # トレンド分析
            if 'trend_analysis' in time_series_analysis:
                trend_data = time_series_analysis['trend_analysis']
                
                trend_summary = []
                for strategy, strategy_trends in trend_data.items():
                    for metric, trend_info in strategy_trends.items():
                        if isinstance(trend_info, dict):
                            trend_summary.append({
                                'Strategy': strategy,
                                'Metric': metric,
                                'Trend_Direction': trend_info.get('trend_direction', 'unknown'),
                                'Slope': trend_info.get('slope', 0),
                                'R_Squared': trend_info.get('r_squared', 0),
                                'Significance': trend_info.get('significance', 'unknown')
                            })
                
                if trend_summary:
                    time_series_content['trend_analysis'] = pd.DataFrame(trend_summary)
            
            if time_series_content:
                self.report_sections.append(ReportSection(
                    title="時系列分析",
                    content=time_series_content,
                    section_type="table"
                ))
            
        except Exception as e:
            self.logger.warning(f"時系列分析生成失敗: {e}")

    def _add_risk_analysis(self, results: Dict[str, Any]):
        """リスク分析の追加"""
        try:
            strategy_market_performance = results.get('strategy_market_performance', {})
            
            if not strategy_market_performance:
                return
            
            # リスクメトリクスの抽出
            risk_metrics = []
            
            for strategy, market_data in strategy_market_performance.items():
                strategy_risks = {'Strategy': strategy}
                
                # リスク指標の集約
                volatilities = []
                max_drawdowns = []
                sharpe_ratios = []
                
                for market_state, metrics in market_data.items():
                    if 'volatility' in metrics and 'mean' in metrics['volatility']:
                        volatilities.append(metrics['volatility']['mean'])
                    
                    if 'max_drawdown' in metrics and 'mean' in metrics['max_drawdown']:
                        max_drawdowns.append(metrics['max_drawdown']['mean'])
                    
                    if 'sharpe_ratio' in metrics and 'mean' in metrics['sharpe_ratio']:
                        sharpe_ratios.append(metrics['sharpe_ratio']['mean'])
                
                # 平均リスク指標
                if volatilities:
                    strategy_risks['Avg_Volatility'] = np.mean(volatilities)
                if max_drawdowns:
                    strategy_risks['Avg_Max_Drawdown'] = np.mean(max_drawdowns)
                if sharpe_ratios:
                    strategy_risks['Avg_Sharpe_Ratio'] = np.mean(sharpe_ratios)
                
                # リスク分類
                if volatilities:
                    avg_vol = np.mean(volatilities)
                    if avg_vol < 0.1:
                        risk_level = "Low"
                    elif avg_vol < 0.2:
                        risk_level = "Medium"
                    else:
                        risk_level = "High"
                    strategy_risks['Risk_Level'] = risk_level
                
                risk_metrics.append(strategy_risks)
            
            if risk_metrics:
                risk_df = pd.DataFrame(risk_metrics)
                
                self.report_sections.append(ReportSection(
                    title="リスク分析",
                    content=risk_df,
                    section_type="table"
                ))
            
        except Exception as e:
            self.logger.warning(f"リスク分析生成失敗: {e}")

    def _add_recommendations(self, results: Dict[str, Any]):
        """推奨事項の追加"""
        try:
            rankings = results.get('performance_rankings', {})
            risk_analysis = None
            
            # リスク分析セクションからリスクデータを取得
            for section in self.report_sections:
                if section.title == "リスク分析" and isinstance(section.content, pd.DataFrame):
                    risk_analysis = section.content
                    break
            
            recommendations = []
            
            # パフォーマンスベースの推奨事項
            if 'overall' in rankings:
                top_strategies = list(rankings['overall'].keys())[:3]
                recommendations.append({
                    'Category': 'パフォーマンス',
                    'Recommendation': f"トップ3戦略の採用を推奨: {', '.join(top_strategies)}",
                    'Priority': 'High'
                })
            
            # リスクベースの推奨事項
            if risk_analysis is not None and 'Risk_Level' in risk_analysis.columns:
                low_risk_strategies = risk_analysis[risk_analysis['Risk_Level'] == 'Low']['Strategy'].tolist()
                if low_risk_strategies:
                    recommendations.append({
                        'Category': 'リスク管理',
                        'Recommendation': f"低リスク戦略の活用: {', '.join(low_risk_strategies)}",
                        'Priority': 'Medium'
                    })
            
            # 市場環境別の推奨事項
            if 'by_market_environment' in rankings:
                for market, market_ranking in rankings['by_market_environment'].items():
                    if market_ranking:
                        best_strategy = list(market_ranking.keys())[0]
                        recommendations.append({
                            'Category': f'市場環境: {market}',
                            'Recommendation': f"{market}環境では{best_strategy}戦略が最適",
                            'Priority': 'Medium'
                        })
            
            # 一般的な推奨事項
            recommendations.extend([
                {
                    'Category': '分散化',
                    'Recommendation': '複数戦略の組み合わせによるリスク分散を検討',
                    'Priority': 'Medium'
                },
                {
                    'Category': '継続監視',
                    'Recommendation': '定期的なパフォーマンス評価と戦略見直し',
                    'Priority': 'High'
                }
            ])
            
            recommendations_df = pd.DataFrame(recommendations)
            
            self.report_sections.append(ReportSection(
                title="推奨事項",
                content=recommendations_df,
                section_type="table"
            ))
            
        except Exception as e:
            self.logger.warning(f"推奨事項生成失敗: {e}")

    def _generate_excel_report(self, results: Dict[str, Any]) -> Path:
        """Excelレポートの生成"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_path = self.output_dir / f"strategy_comparison_report_{timestamp}.xlsx"
            
            with pd.ExcelWriter(str(excel_path), engine='openpyxl') as writer:
                # 各セクションをシートとして追加
                for section in self.report_sections:
                    sheet_name = section.title[:31]  # Excel sheet name limit
                    
                    if section.section_type == "table":
                        if isinstance(section.content, pd.DataFrame):
                            section.content.to_excel(writer, sheet_name=sheet_name, index=False)
                        elif isinstance(section.content, dict):
                            # 辞書の場合は複数のテーブルを含む可能性がある
                            if len(section.content) == 1:
                                # 単一テーブル
                                key, df = next(iter(section.content.items()))
                                if isinstance(df, pd.DataFrame):
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                            else:
                                # 複数テーブルを縦に連結
                                start_row = 0
                                for key, df in section.content.items():
                                    if isinstance(df, pd.DataFrame):
                                        # サブタイトルを追加
                                        subtitle_df = pd.DataFrame([[key]], columns=[''])
                                        subtitle_df.to_excel(writer, sheet_name=sheet_name, 
                                                           startrow=start_row, index=False, header=False)
                                        start_row += 2
                                        
                                        # データを追加
                                        df.to_excel(writer, sheet_name=sheet_name, 
                                                  startrow=start_row, index=False)
                                        start_row += len(df) + 3
                    
                    elif section.section_type == "text":
                        # テキストコンテンツをDataFrameに変換
                        if isinstance(section.content, dict):
                            text_data = []
                            for key, value in section.content.items():
                                if isinstance(value, (list, dict)):
                                    text_data.append([key, str(value)])
                                else:
                                    text_data.append([key, value])
                            
                            text_df = pd.DataFrame(text_data, columns=['項目', '内容'])
                            text_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # スタイルの適用
                self._apply_excel_styles(writer)
            
            self.logger.info(f"Excelレポート生成完了: {excel_path}")
            return excel_path
            
        except Exception as e:
            self.logger.error(f"Excelレポート生成失敗: {e}")
            raise

    def _apply_excel_styles(self, writer):
        """Excelスタイルの適用"""
        try:
            if not hasattr(writer, 'sheets'):
                return
            
            for sheet_name, worksheet in writer.sheets.items():
                # ヘッダーのスタイル適用
                for cell in worksheet[1]:  # 最初の行
                    for key, value in self.excel_styles['header'].items():
                        setattr(cell, key, value)
                
                # データ部分のスタイル適用
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        for key, value in self.excel_styles['data'].items():
                            if key != 'fill':  # fillは条件付きで適用
                                setattr(cell, key, value)
                
                # 列幅の自動調整
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            self.logger.warning(f"Excelスタイル適用失敗: {e}")

    def _generate_html_report(self, results: Dict[str, Any]) -> Path:
        """HTMLレポートの生成"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            html_path = self.output_dir / f"strategy_comparison_report_{timestamp}.html"
            
            # HTMLテンプレートの生成
            html_content = self._create_html_template()
            
            # セクションコンテンツの追加
            sections_html = ""
            for i, section in enumerate(self.report_sections):
                sections_html += f'<div class="section" id="section-{i}">\n'
                sections_html += f'<h2>{section.title}</h2>\n'
                
                if section.section_type == "table":
                    if isinstance(section.content, pd.DataFrame):
                        sections_html += section.content.to_html(classes='table table-striped', escape=False)
                    elif isinstance(section.content, dict):
                        for key, df in section.content.items():
                            if isinstance(df, pd.DataFrame):
                                sections_html += f'<h3>{key}</h3>\n'
                                sections_html += df.to_html(classes='table table-striped', escape=False)
                
                elif section.section_type == "text":
                    if isinstance(section.content, dict):
                        sections_html += '<div class="text-content">\n'
                        for key, value in section.content.items():
                            sections_html += f'<p><strong>{key}:</strong> {value}</p>\n'
                        sections_html += '</div>\n'
                
                sections_html += '</div>\n\n'
            
            # チャートの追加
            if self.config.include_charts and PLOTLY_AVAILABLE:
                charts_html = self._generate_html_charts(results)
                sections_html += charts_html
            
            # HTMLコンテンツの完成
            html_content = html_content.replace('{{CONTENT}}', sections_html)
            html_content = html_content.replace('{{TITLE}}', '戦略比較レポート')
            html_content = html_content.replace('{{TIMESTAMP}}', datetime.now().strftime('%Y年%m月%d日 %H:%M'))
            
            # ファイルの保存
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"HTMLレポート生成完了: {html_path}")
            return html_path
            
        except Exception as e:
            self.logger.error(f"HTMLレポート生成失敗: {e}")
            raise

    def _create_html_template(self) -> str:
        """HTMLテンプレートの作成"""
        return """
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{TITLE}}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        .section { margin-bottom: 40px; padding: 20px; background: #f8f9fa; border-radius: 8px; }
        .table { font-size: 0.9em; }
        .chart-container { margin: 20px 0; }
        .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; margin-bottom: 30px; }
        .nav-sidebar { position: fixed; top: 0; left: 0; width: 250px; height: 100vh; background: #343a40; padding: 20px; overflow-y: auto; }
        .nav-sidebar a { color: #adb5bd; text-decoration: none; display: block; padding: 10px; margin: 5px 0; border-radius: 5px; }
        .nav-sidebar a:hover { background: #495057; color: white; }
        .main-content { margin-left: 270px; padding: 20px; }
    </style>
</head>
<body>
    <div class="nav-sidebar">
        <h5 class="text-light mb-4">目次</h5>
        <a href="#section-0">エグゼクティブサマリー</a>
        <a href="#section-1">戦略パフォーマンス概要</a>
        <a href="#section-2">市場環境別分析</a>
        <a href="#section-3">統計的分析</a>
        <a href="#section-4">相関分析</a>
        <a href="#section-5">パフォーマンスランキング</a>
        <a href="#section-6">時系列分析</a>
        <a href="#section-7">リスク分析</a>
        <a href="#section-8">推奨事項</a>
    </div>
    
    <div class="main-content">
        <div class="header">
            <h1>{{TITLE}}</h1>
            <p>生成日時: {{TIMESTAMP}}</p>
        </div>
        
        {{CONTENT}}
    </div>
    
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
        """

    def _generate_html_charts(self, results: Dict[str, Any]) -> str:
        """HTMLチャートの生成"""
        try:
            charts_html = '<div class="section"><h2>チャート分析</h2>\n'
            
            # パフォーマンス比較チャート
            if 'performance_rankings' in results and 'overall' in results['performance_rankings']:
                rankings = results['performance_rankings']['overall']
                
                fig = go.Figure()
                strategies = list(rankings.keys())[:10]  # Top 10
                scores = [rankings[s] for s in strategies]
                
                fig.add_trace(go.Bar(
                    x=strategies,
                    y=scores,
                    name='パフォーマンススコア',
                    marker_color=self.color_palette[0]
                ))
                
                fig.update_layout(
                    title='戦略パフォーマンスランキング（Top 10）',
                    xaxis_title='戦略',
                    yaxis_title='スコア',
                    height=500
                )
                
                chart_html = pyo.plot(fig, output_type='div', include_plotlyjs=False)
                charts_html += f'<div class="chart-container">{chart_html}</div>\n'
            
            # 相関ヒートマップ
            if ('correlation_analysis' in results and 
                'strategy_correlation_matrix' in results['correlation_analysis']):
                
                corr_data = results['correlation_analysis']['strategy_correlation_matrix']
                if 'matrix' in corr_data and 'strategies' in corr_data:
                    
                    fig = go.Figure(data=go.Heatmap(
                        z=corr_data['matrix'],
                        x=corr_data['strategies'],
                        y=corr_data['strategies'],
                        colorscale='RdBu',
                        zmid=0
                    ))
                    
                    fig.update_layout(
                        title='戦略間相関ヒートマップ',
                        height=600
                    )
                    
                    chart_html = pyo.plot(fig, output_type='div', include_plotlyjs=False)
                    charts_html += f'<div class="chart-container">{chart_html}</div>\n'
            
            charts_html += '</div>\n'
            return charts_html
            
        except Exception as e:
            self.logger.warning(f"HTMLチャート生成失敗: {e}")
            return ""

    def _generate_json_report(self, results: Dict[str, Any]) -> Path:
        """JSONレポートの生成"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            json_path = self.output_dir / f"strategy_comparison_report_{timestamp}.json"
            
            # レポートデータの構築
            report_data = {
                'metadata': {
                    'report_title': '戦略比較レポート',
                    'generation_timestamp': datetime.now().isoformat(),
                    'report_config': asdict(self.config)
                },
                'sections': [],
                'raw_results': results
            }
            
            # セクションデータの追加
            for section in self.report_sections:
                section_data = {
                    'title': section.title,
                    'section_type': section.section_type,
                    'content': {}
                }
                
                if section.section_type == "table":
                    if isinstance(section.content, pd.DataFrame):
                        section_data['content'] = section.content.to_dict('records')
                    elif isinstance(section.content, dict):
                        for key, df in section.content.items():
                            if isinstance(df, pd.DataFrame):
                                section_data['content'][key] = df.to_dict('records')
                            else:
                                section_data['content'][key] = df
                
                elif section.section_type == "text":
                    section_data['content'] = section.content
                
                report_data['sections'].append(section_data)
            
            # JSON保存
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2, default=str)
            
            self.logger.info(f"JSONレポート生成完了: {json_path}")
            return json_path
            
        except Exception as e:
            self.logger.error(f"JSONレポート生成失敗: {e}")
            raise

    def _generate_interactive_dashboard(self, results: Dict[str, Any]) -> Path:
        """インタラクティブダッシュボードの生成"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            dashboard_path = self.output_dir / f"interactive_dashboard_{timestamp}.html"
            
            # ダッシュボード用のサブプロット作成
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=('パフォーマンスランキング', '相関分析', 'リスク・リターン分布', '時系列トレンド'),
                specs=[[{"type": "bar"}, {"type": "heatmap"}],
                       [{"type": "scatter"}, {"type": "scatter"}]]
            )
            
            # 1. パフォーマンスランキング
            if 'performance_rankings' in results and 'overall' in results['performance_rankings']:
                rankings = results['performance_rankings']['overall']
                strategies = list(rankings.keys())[:10]
                scores = [rankings[s] for s in strategies]
                
                fig.add_trace(
                    go.Bar(x=strategies, y=scores, name='スコア'),
                    row=1, col=1
                )
            
            # 2. 相関ヒートマップ
            if ('correlation_analysis' in results and 
                'strategy_correlation_matrix' in results['correlation_analysis']):
                corr_data = results['correlation_analysis']['strategy_correlation_matrix']
                if 'matrix' in corr_data and 'strategies' in corr_data:
                    fig.add_trace(
                        go.Heatmap(
                            z=corr_data['matrix'][:5][:5],  # 5x5に限定
                            x=corr_data['strategies'][:5],
                            y=corr_data['strategies'][:5],
                            colorscale='RdBu'
                        ),
                        row=1, col=2
                    )
            
            # 3. リスク・リターン分布
            risk_return_data = self._extract_risk_return_data(results)
            if risk_return_data:
                fig.add_trace(
                    go.Scatter(
                        x=risk_return_data['risk'],
                        y=risk_return_data['return'],
                        mode='markers+text',
                        text=risk_return_data['strategies'],
                        textposition='top center',
                        name='戦略'
                    ),
                    row=2, col=1
                )
            
            # レイアウトの設定
            fig.update_layout(
                height=800,
                title_text="戦略比較インタラクティブダッシュボード",
                showlegend=False
            )
            
            # HTMLファイルとして保存
            pyo.plot(fig, filename=str(dashboard_path), auto_open=False)
            
            self.logger.info(f"インタラクティブダッシュボード生成完了: {dashboard_path}")
            return dashboard_path
            
        except Exception as e:
            self.logger.error(f"インタラクティブダッシュボード生成失敗: {e}")
            raise

    def _extract_risk_return_data(self, results: Dict[str, Any]) -> Optional[Dict[str, List]]:
        """リスク・リターンデータの抽出"""
        try:
            strategy_market_performance = results.get('strategy_market_performance', {})
            
            if not strategy_market_performance:
                return None
            
            strategies = []
            returns = []
            risks = []
            
            for strategy, market_data in strategy_market_performance.items():
                strategy_returns = []
                strategy_volatilities = []
                
                for market_state, metrics in market_data.items():
                    if 'total_return' in metrics and 'mean' in metrics['total_return']:
                        strategy_returns.append(metrics['total_return']['mean'])
                    
                    if 'volatility' in metrics and 'mean' in metrics['volatility']:
                        strategy_volatilities.append(metrics['volatility']['mean'])
                
                if strategy_returns and strategy_volatilities:
                    strategies.append(strategy)
                    returns.append(np.mean(strategy_returns))
                    risks.append(np.mean(strategy_volatilities))
            
            if strategies:
                return {
                    'strategies': strategies,
                    'return': returns,
                    'risk': risks
                }
            
            return None
            
        except Exception as e:
            self.logger.warning(f"リスク・リターンデータ抽出失敗: {e}")
            return None

def create_report_config(
    output_formats: List[str] = None,
    include_charts: bool = True,
    include_interactive_dashboard: bool = True,
    **kwargs
) -> ReportConfig:
    """レポート設定の作成ヘルパー"""
    return ReportConfig(
        output_formats=output_formats,
        include_charts=include_charts,
        include_interactive_dashboard=include_interactive_dashboard,
        **kwargs
    )

def main():
    """メイン実行関数"""
    import argparse
    
    parser = argparse.ArgumentParser(description="戦略比較レポート生成システム")
    parser.add_argument("--input", required=True, help="集約結果ファイル")
    parser.add_argument("--output-dir", default="output/reports", help="出力ディレクトリ")
    parser.add_argument("--formats", nargs="+", default=["excel", "html", "json"], 
                       choices=["excel", "html", "json"], help="出力形式")
    parser.add_argument("--no-charts", action="store_true", help="チャートを含めない")
    parser.add_argument("--no-dashboard", action="store_true", help="ダッシュボードを含めない")
    
    args = parser.parse_args()
    
    try:
        # 入力ファイルの読み込み
        input_path = Path(args.input)
        if input_path.suffix == '.json':
            with open(input_path, 'r', encoding='utf-8') as f:
                aggregated_results = json.load(f)
        else:
            raise ValueError("JSONファイルを指定してください")
        
        # 設定の作成
        config = create_report_config(
            output_formats=args.formats,
            include_charts=not args.no_charts,
            include_interactive_dashboard=not args.no_dashboard,
            output_directory=args.output_dir
        )
        
        # レポート生成
        reporter = StrategyComparisonReporter(config)
        generated_files = reporter.generate_comprehensive_report(aggregated_results)
        
        print(f"\n=== 戦略比較レポート生成完了 ===")
        for format_type, file_path in generated_files.items():
            print(f"{format_type.upper()}: {file_path}")
        
    except Exception as e:
        print(f"レポート生成失敗: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
