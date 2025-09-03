"""
エクスポート管理モジュール

DSSMS改善プロジェクト Phase 3 Task 3.3
包括的レポートシステムのエクスポート管理コンポーネント

機能:
- マルチフォーマット出力（Excel/PDF/JSON）
- バッチエクスポート機能
- カスタムエクスポートオプション
- エクスポート履歴管理
"""

import os
import sys
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import pandas as pd
import numpy as np

# プロジェクトルートを追加
project_root = Path(__file__).parent.parent.parent.parent
sys.path.append(str(project_root))

from config.logger_config import setup_logger

# オプショナル依存関係の動的インポート
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from weasyprint import HTML, CSS
    PDF_AVAILABLE = True
except ImportError:
    try:
        import pdfkit
        PDF_AVAILABLE = True
    except ImportError:
        PDF_AVAILABLE = False


class ExportManager:
    """エクスポート管理クラス"""
    
    def __init__(self):
        """初期化"""
        self.logger = setup_logger(__name__)
        self.logger.info("=== ExportManager 初期化開始 ===")
        
        # エクスポート機能チェック
        self.available_formats = ['html', 'json']
        if EXCEL_AVAILABLE:
            self.available_formats.append('excel')
        if PDF_AVAILABLE:
            self.available_formats.append('pdf')
        
        self.logger.info(f"利用可能なエクスポート形式: {self.available_formats}")
        
        # エクスポート設定
        self.export_config = {
            'excel': {
                'include_charts': False,  # Excelチャートは複雑なため無効
                'include_formatting': True,
                'sheet_names': {
                    'summary': 'サマリー',
                    'data': 'データ',
                    'analysis': '分析',
                    'metadata': 'メタデータ'
                }
            },
            'pdf': {
                'page_size': 'A4',
                'orientation': 'portrait',
                'include_charts': True,
                'css_optimization': True
            },
            'json': {
                'indent': 2,
                'ensure_ascii': False,
                'include_raw_data': True
            }
        }
        
        # エクスポート履歴
        self.export_history = []
        
        self.logger.info("ExportManager 初期化完了")
    
    def export_report(
        self,
        report_data: Dict[str, Any],
        format_type: str,
        report_id: str,
        output_dir: Path,
        custom_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        レポートエクスポート
        
        Args:
            report_data: レポートデータ
            format_type: エクスポート形式
            report_id: レポートID
            output_dir: 出力ディレクトリ
            custom_options: カスタムオプション
            
        Returns:
            エクスポート結果
        """
        try:
            self.logger.info(f"レポートエクスポート開始: {format_type}")
            
            # 形式チェック
            if format_type not in self.available_formats:
                raise ValueError(f"サポートされていない形式: {format_type}")
            
            # 出力ディレクトリ作成
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # エクスポート実行
            export_result = None
            
            if format_type == 'excel':
                export_result = self._export_to_excel(
                    report_data, report_id, output_dir, custom_options
                )
            elif format_type == 'pdf':
                export_result = self._export_to_pdf(
                    report_data, report_id, output_dir, custom_options
                )
            elif format_type == 'json':
                export_result = self._export_to_json(
                    report_data, report_id, output_dir, custom_options
                )
            else:
                raise ValueError(f"未実装の形式: {format_type}")
            
            # エクスポート履歴追加
            self.export_history.append({
                'timestamp': datetime.now(),
                'report_id': report_id,
                'format_type': format_type,
                'export_path': export_result.get('export_path'),
                'success': export_result.get('success', False)
            })
            
            self.logger.info(f"エクスポート完了: {export_result.get('export_path')}")
            return export_result
            
        except Exception as e:
            self.logger.error(f"エクスポートエラー: {e}")
            return {
                'success': False,
                'error': str(e),
                'format_type': format_type,
                'report_id': report_id
            }
    
    def _export_to_excel(
        self,
        report_data: Dict[str, Any],
        report_id: str,
        output_dir: Path,
        custom_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Excelエクスポート"""
        try:
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxlが利用できません")
            
            self.logger.info("Excelエクスポート開始")
            
            # ファイルパス設定
            excel_path = output_dir / f"{report_id}.xlsx"
            
            # ワークブック作成
            workbook = openpyxl.Workbook()
            
            # デフォルトシート削除
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            # メタデータシート作成
            self._create_metadata_sheet(workbook, report_data)
            
            # サマリーシート作成
            self._create_summary_sheet(workbook, report_data)
            
            # データシート作成
            self._create_data_sheets(workbook, report_data, custom_options)
            
            # 分析シート作成
            self._create_analysis_sheet(workbook, report_data)
            
            # ファイル保存
            workbook.save(excel_path)
            
            self.logger.info(f"Excel保存完了: {excel_path}")
            
            return {
                'success': True,
                'export_path': str(excel_path),
                'format_type': 'excel',
                'file_size': excel_path.stat().st_size,
                'sheets_count': len(workbook.sheetnames)
            }
            
        except Exception as e:
            self.logger.error(f"Excelエクスポートエラー: {e}")
            return {
                'success': False,
                'error': str(e),
                'format_type': 'excel'
            }
    
    def _create_metadata_sheet(self, workbook: openpyxl.Workbook, report_data: Dict[str, Any]):
        """メタデータシート作成"""
        try:
            sheet = workbook.create_sheet(self.export_config['excel']['sheet_names']['metadata'])
            
            metadata = report_data.get('metadata', {})
            
            # ヘッダー設定
            headers = ['項目', '値']
            sheet.append(headers)
            
            # ヘッダーのスタイル設定
            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # メタデータ行追加
            for key, value in metadata.items():
                sheet.append([key, str(value)])
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 40
            
        except Exception as e:
            self.logger.error(f"メタデータシート作成エラー: {e}")
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, report_data: Dict[str, Any]):
        """サマリーシート作成"""
        try:
            sheet = workbook.create_sheet(self.export_config['excel']['sheet_names']['summary'])
            
            # タイトル
            sheet['A1'] = 'レポートサマリー'
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:D1')
            
            row = 3
            
            # サマリー統計
            summary_stats = report_data.get('data', {}).get('summary_statistics', {})
            if summary_stats:
                sheet[f'A{row}'] = 'サマリー統計'
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # データ概要
                data_overview = summary_stats.get('data_overview', {})
                for key, value in data_overview.items():
                    sheet[f'A{row}'] = key.replace('_', ' ').title()
                    sheet[f'B{row}'] = value
                    row += 1
                
                row += 1
            
            # 可視化情報
            visualizations = report_data.get('visualizations', {})
            if visualizations:
                sheet[f'A{row}'] = '可視化統計'
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                sheet[f'A{row}'] = 'チャート数'
                sheet[f'B{row}'] = visualizations.get('metadata', {}).get('total_charts', 0)
                row += 1
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 20
            
        except Exception as e:
            self.logger.error(f"サマリーシート作成エラー: {e}")
    
    def _create_data_sheets(
        self,
        workbook: openpyxl.Workbook,
        report_data: Dict[str, Any],
        custom_options: Optional[Dict[str, Any]] = None
    ):
        """データシート作成"""
        try:
            data = report_data.get('data', {})
            
            # DSSMSデータシート
            if 'dssms_data' in data and data['dssms_data']:
                self._create_dssms_data_sheet(workbook, data['dssms_data'])
            
            # 戦略データシート
            if 'strategy_data' in data and data['strategy_data']:
                self._create_strategy_data_sheet(workbook, data['strategy_data'])
            
            # パフォーマンスデータシート
            if 'performance_data' in data and data['performance_data']:
                self._create_performance_data_sheet(workbook, data['performance_data'])
            
        except Exception as e:
            self.logger.error(f"データシート作成エラー: {e}")
    
    def _create_dssms_data_sheet(self, workbook: openpyxl.Workbook, dssms_data: Dict[str, Any]):
        """DSSMSデータシート作成"""
        try:
            sheet = workbook.create_sheet('DSSMS データ')
            
            row = 1
            
            # ヘッダー
            headers = ['ファイル名', 'レコード数', '詳細']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            row += 1
            
            # データ行
            for file_name, file_data in dssms_data.items():
                sheet.cell(row=row, column=1, value=file_name)
                
                if isinstance(file_data, dict):
                    record_count = file_data.get('total_records', 0)
                    if record_count == 0 and 'full_data' in file_data:
                        if hasattr(file_data['full_data'], '__len__'):
                            record_count = len(file_data['full_data'])
                    
                    sheet.cell(row=row, column=2, value=record_count)
                    sheet.cell(row=row, column=3, value=f"データタイプ: {type(file_data).__name__}")
                else:
                    sheet.cell(row=row, column=2, value=0)
                    sheet.cell(row=row, column=3, value="データなし")
                
                row += 1
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 40
            
        except Exception as e:
            self.logger.error(f"DSSMSデータシート作成エラー: {e}")
    
    def _create_strategy_data_sheet(self, workbook: openpyxl.Workbook, strategy_data: Dict[str, Any]):
        """戦略データシート作成"""
        try:
            sheet = workbook.create_sheet('戦略データ')
            
            row = 1
            
            # ヘッダー
            headers = ['戦略名', 'コード行数', 'ファイルサイズ', '最終更新']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            
            row += 1
            
            # データ行
            for strategy_name, strategy_info in strategy_data.items():
                sheet.cell(row=row, column=1, value=strategy_name)
                
                if isinstance(strategy_info, dict):
                    sheet.cell(row=row, column=2, value=strategy_info.get('lines_count', 0))
                    sheet.cell(row=row, column=3, value=strategy_info.get('file_size', 0))
                    
                    last_modified = strategy_info.get('last_modified')
                    if last_modified:
                        sheet.cell(row=row, column=4, value=str(last_modified))
                    else:
                        sheet.cell(row=row, column=4, value="不明")
                else:
                    sheet.cell(row=row, column=2, value=0)
                    sheet.cell(row=row, column=3, value=0)
                    sheet.cell(row=row, column=4, value="不明")
                
                row += 1
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 20
            
        except Exception as e:
            self.logger.error(f"戦略データシート作成エラー: {e}")
    
    def _create_performance_data_sheet(self, workbook: openpyxl.Workbook, performance_data: Dict[str, Any]):
        """パフォーマンスデータシート作成"""
        try:
            sheet = workbook.create_sheet('パフォーマンス')
            
            row = 1
            
            # ヘッダー
            headers = ['ファイル名', 'レコード数', '状態']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
            
            row += 1
            
            # データ行
            for file_name, file_data in performance_data.items():
                sheet.cell(row=row, column=1, value=file_name)
                
                if isinstance(file_data, dict):
                    record_count = file_data.get('record_count', 0)
                    if record_count == 0 and 'data' in file_data:
                        if hasattr(file_data['data'], '__len__'):
                            record_count = len(file_data['data'])
                    
                    sheet.cell(row=row, column=2, value=record_count)
                    sheet.cell(row=row, column=3, value="データあり")
                else:
                    sheet.cell(row=row, column=2, value=0)
                    sheet.cell(row=row, column=3, value="データなし")
                
                row += 1
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 15
            
        except Exception as e:
            self.logger.error(f"パフォーマンスデータシート作成エラー: {e}")
    
    def _create_analysis_sheet(self, workbook: openpyxl.Workbook, report_data: Dict[str, Any]):
        """分析シート作成"""
        try:
            sheet = workbook.create_sheet(self.export_config['excel']['sheet_names']['analysis'])
            
            # タイトル
            sheet['A1'] = '詳細分析'
            sheet['A1'].font = Font(size=16, bold=True)
            sheet.merge_cells('A1:D1')
            
            row = 3
            
            # 生成情報
            metadata = report_data.get('metadata', {})
            if metadata:
                sheet[f'A{row}'] = '生成情報'
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for key, value in metadata.items():
                    sheet[f'A{row}'] = key
                    sheet[f'B{row}'] = str(value)
                    row += 1
                
                row += 1
            
            # データ品質分析
            data = report_data.get('data', {})
            if data:
                sheet[f'A{row}'] = 'データ品質分析'
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for category, category_data in data.items():
                    if isinstance(category_data, dict):
                        sheet[f'A{row}'] = category
                        sheet[f'B{row}'] = f"{len(category_data)} 項目"
                        row += 1
            
            # 列幅調整
            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 30
            
        except Exception as e:
            self.logger.error(f"分析シート作成エラー: {e}")
    
    def _export_to_pdf(
        self,
        report_data: Dict[str, Any],
        report_id: str,
        output_dir: Path,
        custom_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """PDFエクスポート"""
        try:
            if not PDF_AVAILABLE:
                raise ImportError("PDF生成ライブラリが利用できません")
            
            self.logger.info("PDFエクスポート開始")
            
            # HTMLレポートが必要（PDFはHTMLから生成）
            from src.reports.comprehensive.report_template_manager import ReportTemplateManager
            
            template_manager = ReportTemplateManager()
            html_content = template_manager.generate_html_report(
                report_data, 
                template_type="comprehensive", 
                level="detailed"
            )
            
            # ファイルパス設定
            pdf_path = output_dir / f"{report_id}.pdf"
            
            # PDF生成（WeasyPrintを優先）
            try:
                # CSS最適化
                pdf_css = CSS(string="""
                    @media print {
                        .chart-container { height: 300px !important; }
                        .table { font-size: 10px; }
                        .card { page-break-inside: avoid; }
                    }
                    body { font-size: 12px; }
                    .container-fluid { max-width: 100%; }
                """)
                
                HTML(string=html_content).write_pdf(pdf_path, stylesheets=[pdf_css])
                
            except Exception as e:
                # WeasyPrint失敗時はpdfkitを試行
                self.logger.warning(f"WeasyPrint失敗、pdfkitを試行: {e}")
                
                # 一時HTMLファイル作成
                temp_html_path = output_dir / f"{report_id}_temp.html"
                with open(temp_html_path, 'w', encoding='utf-8') as f:
                    f.write(html_content)
                
                # pdfkit使用
                pdfkit.from_file(str(temp_html_path), str(pdf_path))
                
                # 一時ファイル削除
                temp_html_path.unlink()
            
            self.logger.info(f"PDF保存完了: {pdf_path}")
            
            return {
                'success': True,
                'export_path': str(pdf_path),
                'format_type': 'pdf',
                'file_size': pdf_path.stat().st_size
            }
            
        except Exception as e:
            self.logger.error(f"PDFエクスポートエラー: {e}")
            return {
                'success': False,
                'error': str(e),
                'format_type': 'pdf'
            }
    
    def _export_to_json(
        self,
        report_data: Dict[str, Any],
        report_id: str,
        output_dir: Path,
        custom_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """JSONエクスポート"""
        try:
            self.logger.info("JSONエクスポート開始")
            
            # ファイルパス設定
            json_path = output_dir / f"{report_id}.json"
            
            # JSON用データ準備
            json_data = self._prepare_json_data(report_data, custom_options)
            
            # JSON保存
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(
                    json_data,
                    f,
                    indent=self.export_config['json']['indent'],
                    ensure_ascii=self.export_config['json']['ensure_ascii'],
                    default=self._json_serializer
                )
            
            self.logger.info(f"JSON保存完了: {json_path}")
            
            return {
                'success': True,
                'export_path': str(json_path),
                'format_type': 'json',
                'file_size': json_path.stat().st_size
            }
            
        except Exception as e:
            self.logger.error(f"JSONエクスポートエラー: {e}")
            return {
                'success': False,
                'error': str(e),
                'format_type': 'json'
            }
    
    def _prepare_json_data(
        self,
        report_data: Dict[str, Any],
        custom_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """JSON用データ準備"""
        try:
            json_data = {
                'export_info': {
                    'export_timestamp': datetime.now(),
                    'format': 'json',
                    'version': '1.0'
                }
            }
            
            # メタデータコピー
            if 'metadata' in report_data:
                json_data['metadata'] = report_data['metadata'].copy()
            
            # データ処理
            if 'data' in report_data:
                json_data['data'] = self._process_data_for_json(report_data['data'])
            
            # 可視化データ（JavaScript/HTMLは除外）
            if 'visualizations' in report_data:
                viz_data = report_data['visualizations'].copy()
                # JavaScriptとHTMLコードは除外（サイズとセキュリティのため）
                if 'javascript_code' in viz_data:
                    viz_data['javascript_code_count'] = len(viz_data['javascript_code'])
                    del viz_data['javascript_code']
                if 'html_snippets' in viz_data:
                    viz_data['html_snippets_count'] = len(viz_data['html_snippets'])
                    del viz_data['html_snippets']
                json_data['visualizations'] = viz_data
            
            # カスタムパラメータ
            if 'custom_params' in report_data:
                json_data['custom_params'] = report_data['custom_params']
            
            return json_data
            
        except Exception as e:
            self.logger.error(f"JSON用データ準備エラー: {e}")
            return {'error': str(e)}
    
    def _process_data_for_json(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """データをJSON用に処理"""
        processed_data = {}
        
        for key, value in data.items():
            if isinstance(value, pd.DataFrame):
                # DataFrameをJSONシリアライズ可能な形式に変換
                processed_data[key] = {
                    'type': 'DataFrame',
                    'shape': value.shape,
                    'columns': list(value.columns),
                    'data_sample': value.head().to_dict('records') if not value.empty else [],
                    'summary': value.describe().to_dict() if not value.empty else {}
                }
            elif isinstance(value, dict):
                processed_data[key] = self._process_data_for_json(value)
            elif isinstance(value, (list, tuple)):
                processed_data[key] = list(value)
            else:
                processed_data[key] = value
        
        return processed_data
    
    def _json_serializer(self, obj):
        """JSON シリアライザー"""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        elif hasattr(obj, '__dict__'):
            return obj.__dict__
        else:
            return str(obj)
    
    def batch_export(
        self,
        report_data: Dict[str, Any],
        report_id: str,
        output_dir: Path,
        formats: List[str],
        custom_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """バッチエクスポート"""
        try:
            self.logger.info(f"バッチエクスポート開始: {formats}")
            
            results = {
                'batch_id': f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'total_formats': len(formats),
                'successful_exports': 0,
                'failed_exports': 0,
                'results': {}
            }
            
            for format_type in formats:
                try:
                    export_result = self.export_report(
                        report_data, format_type, report_id, output_dir, custom_options
                    )
                    
                    results['results'][format_type] = export_result
                    
                    if export_result.get('success', False):
                        results['successful_exports'] += 1
                    else:
                        results['failed_exports'] += 1
                        
                except Exception as e:
                    self.logger.error(f"バッチエクスポートエラー {format_type}: {e}")
                    results['results'][format_type] = {
                        'success': False,
                        'error': str(e),
                        'format_type': format_type
                    }
                    results['failed_exports'] += 1
            
            self.logger.info(f"バッチエクスポート完了: {results['successful_exports']}/{results['total_formats']}")
            return results
            
        except Exception as e:
            self.logger.error(f"バッチエクスポートエラー: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_export_history(self, limit: int = 20) -> List[Dict[str, Any]]:
        """エクスポート履歴取得"""
        try:
            # 最新の履歴を返す
            return sorted(
                self.export_history,
                key=lambda x: x['timestamp'],
                reverse=True
            )[:limit]
            
        except Exception as e:
            self.logger.error(f"エクスポート履歴取得エラー: {e}")
            return []
    
    def get_supported_formats(self) -> List[str]:
        """サポートされている形式一覧取得"""
        return self.available_formats.copy()


if __name__ == "__main__":
    # デモ実行
    manager = ExportManager()
    
    print(f"サポートされている形式: {manager.get_supported_formats()}")
    
    # サンプルレポートデータ
    sample_data = {
        'metadata': {
            'report_id': 'test_export_001',
            'generation_timestamp': datetime.now(),
            'report_type': 'comprehensive'
        },
        'data': {
            'summary_statistics': {
                'data_overview': {
                    'total_files': 5,
                    'total_strategies': 3
                }
            }
        }
    }
    
    # JSONエクスポートテスト
    output_dir = Path("test_output")
    output_dir.mkdir(exist_ok=True)
    
    result = manager.export_report(
        sample_data,
        'json',
        'test_export_001',
        output_dir
    )
    
    print(f"エクスポート結果: {result}")
    
    # 履歴表示
    history = manager.get_export_history()
    print(f"エクスポート履歴: {len(history)} 件")
