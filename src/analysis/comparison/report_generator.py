"""
レポートジェネレーター
フェーズ4A3: バックテストvs実運用比較分析器
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
from datetime import datetime
import logging
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ReportGenerator:
    """レポートジェネレーター"""
    
    def __init__(self, config: Dict[str, Any], logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.output_settings = config.get('output_settings', {})
        self.reports_dir = self.output_settings.get('reports_dir', 'reports')
        
        # 出力ディレクトリ作成
        os.makedirs(self.reports_dir, exist_ok=True)
        
        # スタイル設定
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
    
    def generate_comprehensive_report(self, comparison_results: Dict[str, Any], 
                                    statistical_results: Dict[str, Any] = None,
                                    visualization_results: Dict[str, Any] = None) -> Dict[str, Any]:
        """包括的レポート生成"""
        try:
            self.logger.info("包括的レポート生成開始")
            
            report_results = {
                "timestamp": datetime.now(),
                "reports_generated": [],
                "output_directory": self.reports_dir
            }
            
            # 1. Excel詳細レポート
            if OPENPYXL_AVAILABLE:
                excel_report = self._generate_excel_report(
                    comparison_results, statistical_results, visualization_results
                )
                if excel_report:
                    report_results["reports_generated"].append(excel_report)
            
            # 2. JSONレポート
            json_report = self._generate_json_report(
                comparison_results, statistical_results, visualization_results
            )
            if json_report:
                report_results["reports_generated"].append(json_report)
            
            # 3. CSVエクスポート
            csv_reports = self._generate_csv_reports(comparison_results)
            report_results["reports_generated"].extend(csv_reports)
            
            # 4. テキストサマリーレポート
            text_report = self._generate_text_summary_report(
                comparison_results, statistical_results
            )
            if text_report:
                report_results["reports_generated"].append(text_report)
            
            # 5. エグゼクティブサマリー
            executive_summary = self._generate_executive_summary(comparison_results)
            if executive_summary:
                report_results["reports_generated"].append(executive_summary)
            
            self.logger.info(f"レポート生成完了 - ファイル数: {len(report_results['reports_generated'])}")
            return report_results
            
        except Exception as e:
            self.logger.error(f"包括的レポート生成エラー: {e}")
            return {}
    
    def _generate_excel_report(self, comparison_results: Dict[str, Any], 
                             statistical_results: Dict[str, Any] = None,
                             visualization_results: Dict[str, Any] = None) -> Optional[Dict[str, Any]]:
        """Excel詳細レポート生成"""
        try:
            if not OPENPYXL_AVAILABLE:
                self.logger.warning("openpyxlが利用できません")
                return None
            
            wb = Workbook()
            
            # デフォルトシート削除
            wb.remove(wb.active)
            
            # 1. サマリーシート
            self._create_summary_sheet(wb, comparison_results)
            
            # 2. 戦略比較シート
            self._create_strategy_comparison_sheet(wb, comparison_results)
            
            # 3. ポートフォリオ比較シート
            self._create_portfolio_comparison_sheet(wb, comparison_results)
            
            # 4. 統計分析シート
            if statistical_results:
                self._create_statistical_analysis_sheet(wb, statistical_results)
            
            # 5. パフォーマンスギャップシート
            self._create_performance_gap_sheet(wb, comparison_results)
            
            # ファイル保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backtest_vs_live_comparison_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            return {
                "report_type": "excel_detailed",
                "filename": filename,
                "filepath": filepath,
                "sheets_created": len(wb.sheetnames)
            }
            
        except Exception as e:
            self.logger.warning(f"Excelレポート生成エラー: {e}")
            return None
    
    def _create_summary_sheet(self, workbook: Any, comparison_results: Dict[str, Any]):
        """サマリーシート作成"""
        try:
            ws = workbook.create_sheet("サマリー", 0)
            
            # タイトル
            ws['A1'] = 'バックテスト vs 実運用 比較分析サマリー'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = self.center_alignment
            ws.merge_cells('A1:F1')
            
            # 基本情報
            row = 3
            ws[f'A{row}'] = '分析実行日時'
            ws[f'B{row}'] = comparison_results.get('timestamp', 'N/A')
            row += 1
            
            ws[f'A{row}'] = '分析タイプ'
            ws[f'B{row}'] = comparison_results.get('analysis_type', 'N/A')
            row += 1
            
            # サマリー情報
            summary = comparison_results.get('summary', {})
            
            ws[f'A{row}'] = '比較戦略数'
            ws[f'B{row}'] = summary.get('total_strategies_compared', 0)
            row += 1
            
            ws[f'A{row}'] = '全体パフォーマンスギャップ'
            ws[f'B{row}'] = summary.get('overall_performance_gap', 'unknown')
            row += 2
            
            # 主要発見事項
            ws[f'A{row}'] = '主要発見事項'
            ws[f'A{row}'].font = self.header_font
            row += 1
            
            key_findings = summary.get('key_findings', [])
            for i, finding in enumerate(key_findings, 1):
                ws[f'A{row}'] = f'{i}. {finding}'
                row += 1
            
            row += 1
            
            # 推奨事項
            ws[f'A{row}'] = '推奨事項'
            ws[f'A{row}'].font = self.header_font
            row += 1
            
            recommendations = summary.get('recommendations', [])
            for i, recommendation in enumerate(recommendations, 1):
                ws[f'A{row}'] = f'{i}. {recommendation}'
                row += 1
            
            # 列幅調整
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            self.logger.warning(f"サマリーシート作成エラー: {e}")
    
    def _create_strategy_comparison_sheet(self, workbook: Any, comparison_results: Dict[str, Any]):
        """戦略比較シート作成"""
        try:
            ws = workbook.create_sheet("戦略比較")
            
            strategy_comparisons = comparison_results.get('strategy_comparisons', {})
            if not strategy_comparisons:
                ws['A1'] = '戦略比較データがありません'
                return
            
            # ヘッダー作成
            headers = ['戦略名', 'メトリクス', 'バックテスト', '実運用', '絶対差分', '相対差分(%)', 'パフォーマンス判定']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
            
            # データ追加
            row = 2
            for strategy_name, comparison in strategy_comparisons.items():
                metrics_comparison = comparison.get('metrics_comparison', {})
                
                for metric_name, metric_data in metrics_comparison.items():
                    ws.cell(row=row, column=1, value=strategy_name)
                    ws.cell(row=row, column=2, value=metric_name)
                    ws.cell(row=row, column=3, value=metric_data.get('backtest', 0))
                    ws.cell(row=row, column=4, value=metric_data.get('live', 0))
                    ws.cell(row=row, column=5, value=metric_data.get('absolute_difference', 0))
                    ws.cell(row=row, column=6, value=metric_data.get('relative_difference', 0) * 100)
                    
                    # パフォーマンス判定
                    rel_diff = metric_data.get('relative_difference', 0)
                    if rel_diff > 0.1:
                        performance = "実運用優位"
                    elif rel_diff < -0.1:
                        performance = "バックテスト優位"
                    else:
                        performance = "同等"
                    
                    ws.cell(row=row, column=7, value=performance)
                    row += 1
            
            # 列幅調整
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['G'].width = 15
            
        except Exception as e:
            self.logger.warning(f"戦略比較シート作成エラー: {e}")
    
    def _create_portfolio_comparison_sheet(self, workbook: Any, comparison_results: Dict[str, Any]):
        """ポートフォリオ比較シート作成"""
        try:
            ws = workbook.create_sheet("ポートフォリオ比較")
            
            portfolio_comparison = comparison_results.get('portfolio_comparison', {})
            if not portfolio_comparison:
                ws['A1'] = 'ポートフォリオ比較データがありません'
                return
            
            # 集計メトリクス
            ws['A1'] = 'ポートフォリオ集計メトリクス比較'
            ws['A1'].font = self.title_font
            
            aggregate_metrics = portfolio_comparison.get('aggregate_metrics', {})
            
            # ヘッダー
            headers = ['メトリクス', 'バックテスト', '実運用', '差分', '相対差分(%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # データ
            row = 4
            for metric_name, metric_data in aggregate_metrics.items():
                ws.cell(row=row, column=1, value=metric_name)
                ws.cell(row=row, column=2, value=metric_data.get('backtest', 0))
                ws.cell(row=row, column=3, value=metric_data.get('live', 0))
                ws.cell(row=row, column=4, value=metric_data.get('difference', 0))
                ws.cell(row=row, column=5, value=metric_data.get('relative_difference', 0) * 100)
                row += 1
            
            # 分散化分析
            row += 2
            ws[f'A{row}'] = '分散化分析'
            ws[f'A{row}'].font = self.title_font
            row += 1
            
            diversification_analysis = portfolio_comparison.get('diversification_analysis', {})
            for analysis_name, analysis_data in diversification_analysis.items():
                ws.cell(row=row, column=1, value=analysis_name)
                ws.cell(row=row, column=2, value=str(analysis_data))
                row += 1
            
        except Exception as e:
            self.logger.warning(f"ポートフォリオ比較シート作成エラー: {e}")
    
    def _create_statistical_analysis_sheet(self, workbook: Any, statistical_results: Dict[str, Any]):
        """統計分析シート作成"""
        try:
            ws = workbook.create_sheet("統計分析")
            
            # タイトル
            ws['A1'] = '統計分析結果'
            ws['A1'].font = self.title_font
            
            statistical_tests = statistical_results.get('statistical_tests', {})
            
            # ヘッダー
            headers = ['戦略', 'テスト名', '統計量', 'p値', '有意性', '効果量', '解釈']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # データ
            row = 4
            for strategy_name, strategy_stats in statistical_tests.items():
                significance_tests = strategy_stats.get('significance_tests', {})
                
                for test_name, test_result in significance_tests.items():
                    ws.cell(row=row, column=1, value=strategy_name)
                    ws.cell(row=row, column=2, value=test_name)
                    ws.cell(row=row, column=3, value=test_result.get('statistic', 0))
                    ws.cell(row=row, column=4, value=test_result.get('p_value', 0))
                    ws.cell(row=row, column=5, value='有意' if test_result.get('is_significant', False) else '非有意')
                    ws.cell(row=row, column=6, value=test_result.get('effect', 'N/A'))
                    ws.cell(row=row, column=7, value=test_result.get('interpretation', 'N/A'))
                    row += 1
            
        except Exception as e:
            self.logger.warning(f"統計分析シート作成エラー: {e}")
    
    def _create_performance_gap_sheet(self, workbook: Any, comparison_results: Dict[str, Any]):
        """パフォーマンスギャップシート作成"""
        try:
            ws = workbook.create_sheet("パフォーマンスギャップ")
            
            strategy_comparisons = comparison_results.get('strategy_comparisons', {})
            
            # ヘッダー
            headers = ['戦略名', 'ギャップスコア', '全体ギャップ', '重要ギャップ数', 'ポジティブギャップ数']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # データ
            row = 2
            for strategy_name, comparison in strategy_comparisons.items():
                gap_analysis = comparison.get('performance_gap_analysis', {})
                
                ws.cell(row=row, column=1, value=strategy_name)
                ws.cell(row=row, column=2, value=gap_analysis.get('gap_score', 0))
                ws.cell(row=row, column=3, value=gap_analysis.get('overall_gap', 'neutral'))
                ws.cell(row=row, column=4, value=len(gap_analysis.get('critical_gaps', [])))
                ws.cell(row=row, column=5, value=len(gap_analysis.get('positive_gaps', [])))
                row += 1
            
        except Exception as e:
            self.logger.warning(f"パフォーマンスギャップシート作成エラー: {e}")
    
    def _generate_json_report(self, comparison_results: Dict[str, Any], 
                            statistical_results: Dict[str, Any] = None,
                            visualization_results: Dict[str, Any] = None) -> Optional[Dict[str, Any]]:
        """JSONレポート生成"""
        try:
            report_data = {
                "report_type": "comprehensive_json",
                "generation_timestamp": datetime.now().isoformat(),
                "comparison_analysis": comparison_results,
                "statistical_analysis": statistical_results or {},
                "visualization_analysis": visualization_results or {},
                "metadata": {
                    "config": self.config,
                    "version": "1.0.0"
                }
            }
            
            # ファイル保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"comparison_analysis_{timestamp}.json"
            filepath = os.path.join(self.reports_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2, default=str)
            
            return {
                "report_type": "json_comprehensive",
                "filename": filename,
                "filepath": filepath,
                "size_bytes": os.path.getsize(filepath)
            }
            
        except Exception as e:
            self.logger.warning(f"JSONレポート生成エラー: {e}")
            return None
    
    def _generate_csv_reports(self, comparison_results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """CSVレポート生成"""
        try:
            csv_reports = []
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 1. 戦略比較CSV
            strategy_csv = self._create_strategy_comparison_csv(comparison_results, timestamp)
            if strategy_csv:
                csv_reports.append(strategy_csv)
            
            # 2. ポートフォリオ比較CSV
            portfolio_csv = self._create_portfolio_comparison_csv(comparison_results, timestamp)
            if portfolio_csv:
                csv_reports.append(portfolio_csv)
            
            return csv_reports
            
        except Exception as e:
            self.logger.warning(f"CSVレポート生成エラー: {e}")
            return []
    
    def _create_strategy_comparison_csv(self, comparison_results: Dict[str, Any], timestamp: str) -> Optional[Dict[str, Any]]:
        """戦略比較CSV作成"""
        try:
            strategy_comparisons = comparison_results.get('strategy_comparisons', {})
            if not strategy_comparisons:
                return None
            
            # DataFrameで整理
            data = []
            for strategy_name, comparison in strategy_comparisons.items():
                metrics_comparison = comparison.get('metrics_comparison', {})
                
                for metric_name, metric_data in metrics_comparison.items():
                    data.append({
                        'strategy_name': strategy_name,
                        'metric': metric_name,
                        'backtest_value': metric_data.get('backtest', 0),
                        'live_value': metric_data.get('live', 0),
                        'absolute_difference': metric_data.get('absolute_difference', 0),
                        'relative_difference_pct': metric_data.get('relative_difference', 0) * 100,
                        'performance_ratio': metric_data.get('performance_ratio', 1)
                    })
            
            if not data:
                return None
            
            df = pd.DataFrame(data)
            
            # ファイル保存
            filename = f"strategy_comparison_{timestamp}.csv"
            filepath = os.path.join(self.reports_dir, filename)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            return {
                "report_type": "csv_strategy_comparison",
                "filename": filename,
                "filepath": filepath,
                "records_count": len(data)
            }
            
        except Exception as e:
            self.logger.warning(f"戦略比較CSV作成エラー: {e}")
            return None
    
    def _create_portfolio_comparison_csv(self, comparison_results: Dict[str, Any], timestamp: str) -> Optional[Dict[str, Any]]:
        """ポートフォリオ比較CSV作成"""
        try:
            portfolio_comparison = comparison_results.get('portfolio_comparison', {})
            aggregate_metrics = portfolio_comparison.get('aggregate_metrics', {})
            
            if not aggregate_metrics:
                return None
            
            # DataFrameで整理
            data = []
            for metric_name, metric_data in aggregate_metrics.items():
                data.append({
                    'metric': metric_name,
                    'backtest_value': metric_data.get('backtest', 0),
                    'live_value': metric_data.get('live', 0),
                    'difference': metric_data.get('difference', 0),
                    'relative_difference_pct': metric_data.get('relative_difference', 0) * 100
                })
            
            df = pd.DataFrame(data)
            
            # ファイル保存
            filename = f"portfolio_comparison_{timestamp}.csv"
            filepath = os.path.join(self.reports_dir, filename)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            return {
                "report_type": "csv_portfolio_comparison",
                "filename": filename,
                "filepath": filepath,
                "records_count": len(data)
            }
            
        except Exception as e:
            self.logger.warning(f"ポートフォリオ比較CSV作成エラー: {e}")
            return None
    
    def _generate_text_summary_report(self, comparison_results: Dict[str, Any], 
                                    statistical_results: Dict[str, Any] = None) -> Optional[Dict[str, Any]]:
        """テキストサマリーレポート生成"""
        try:
            lines = []
            lines.append("=" * 80)
            lines.append("バックテスト vs 実運用 比較分析 詳細レポート")
            lines.append("=" * 80)
            lines.append(f"生成日時: {datetime.now().strftime('%Y年%m月%d日 %H時%M分%S秒')}")
            lines.append("")
            
            # サマリーセクション
            summary = comparison_results.get('summary', {})
            lines.append("【エグゼクティブサマリー】")
            lines.append(f"比較戦略数: {summary.get('total_strategies_compared', 0)}")
            lines.append(f"全体パフォーマンスギャップ: {summary.get('overall_performance_gap', 'unknown')}")
            lines.append("")
            
            # 主要発見事項
            lines.append("【主要発見事項】")
            for i, finding in enumerate(summary.get('key_findings', []), 1):
                lines.append(f"{i}. {finding}")
            lines.append("")
            
            # 戦略別詳細
            lines.append("【戦略別詳細分析】")
            strategy_comparisons = comparison_results.get('strategy_comparisons', {})
            
            for strategy_name, comparison in strategy_comparisons.items():
                lines.append(f"\n■ {strategy_name}")
                lines.append("-" * 40)
                
                metrics_comparison = comparison.get('metrics_comparison', {})
                for metric_name, metric_data in metrics_comparison.items():
                    bt_val = metric_data.get('backtest', 0)
                    live_val = metric_data.get('live', 0)
                    rel_diff = metric_data.get('relative_difference', 0)
                    
                    lines.append(f"  {metric_name}:")
                    lines.append(f"    バックテスト: {bt_val:.6f}")
                    lines.append(f"    実運用: {live_val:.6f}")
                    lines.append(f"    相対差分: {rel_diff:.2%}")
                
                # パフォーマンスギャップ分析
                gap_analysis = comparison.get('performance_gap_analysis', {})
                overall_gap = gap_analysis.get('overall_gap', 'neutral')
                gap_score = gap_analysis.get('gap_score', 0)
                
                lines.append(f"  全体ギャップ: {overall_gap}")
                lines.append(f"  ギャップスコア: {gap_score:.4f}")
            
            # 統計分析結果
            if statistical_results:
                lines.append("\n\n【統計分析結果】")
                statistical_tests = statistical_results.get('statistical_tests', {})
                
                for strategy_name, strategy_stats in statistical_tests.items():
                    lines.append(f"\n■ {strategy_name} 統計テスト")
                    
                    significance_tests = strategy_stats.get('significance_tests', {})
                    for test_name, test_result in significance_tests.items():
                        lines.append(f"  {test_name}:")
                        lines.append(f"    統計量: {test_result.get('statistic', 0):.6f}")
                        lines.append(f"    p値: {test_result.get('p_value', 0):.6f}")
                        lines.append(f"    有意性: {'有意' if test_result.get('is_significant', False) else '非有意'}")
            
            # 推奨事項
            lines.append("\n\n【推奨事項】")
            for i, recommendation in enumerate(summary.get('recommendations', []), 1):
                lines.append(f"{i}. {recommendation}")
            
            lines.append("\n" + "=" * 80)
            lines.append("レポート終了")
            lines.append("=" * 80)
            
            # ファイル保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"detailed_analysis_report_{timestamp}.txt"
            filepath = os.path.join(self.reports_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            return {
                "report_type": "text_detailed",
                "filename": filename,
                "filepath": filepath,
                "lines_count": len(lines)
            }
            
        except Exception as e:
            self.logger.warning(f"テキストサマリーレポート生成エラー: {e}")
            return None
    
    def _generate_executive_summary(self, comparison_results: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """エグゼクティブサマリー生成"""
        try:
            summary = comparison_results.get('summary', {})
            
            lines = []
            lines.append("=" * 60)
            lines.append("エグゼクティブサマリー")
            lines.append("バックテスト vs 実運用 比較分析")
            lines.append("=" * 60)
            lines.append("")
            
            # 主要指標
            lines.append("【主要指標】")
            lines.append(f"• 分析対象戦略数: {summary.get('total_strategies_compared', 0)}")
            lines.append(f"• 全体パフォーマンス評価: {summary.get('overall_performance_gap', 'unknown')}")
            lines.append(f"• 分析実行日: {datetime.now().strftime('%Y年%m月%d日')}")
            lines.append("")
            
            # キーファインディング
            lines.append("【重要な発見】")
            key_findings = summary.get('key_findings', [])
            if key_findings:
                for finding in key_findings[:3]:  # 最大3つ
                    lines.append(f"• {finding}")
            else:
                lines.append("• 特筆すべき発見事項はありませんでした")
            lines.append("")
            
            # アクションアイテム
            lines.append("【推奨アクション】")
            recommendations = summary.get('recommendations', [])
            if recommendations:
                for recommendation in recommendations:
                    lines.append(f"• {recommendation}")
            else:
                lines.append("• 現状の運用を継続することを推奨します")
            lines.append("")
            
            # リスク評価
            lines.append("【リスク評価】")
            overall_gap = summary.get('overall_performance_gap', 'neutral')
            if overall_gap == 'negative':
                lines.append("• 高リスク: 実運用がバックテストを大幅に下回っています")
                lines.append("• 即座に原因分析と改善策の検討が必要です")
            elif overall_gap == 'positive':
                lines.append("• 低リスク: 実運用がバックテストを上回っています")
                lines.append("• 現在の戦略は効果的に機能しています")
            else:
                lines.append("• 中リスク: 実運用とバックテストの性能は同等です")
                lines.append("• 継続的なモニタリングを推奨します")
            
            lines.append("")
            lines.append("=" * 60)
            
            # ファイル保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"executive_summary_{timestamp}.txt"
            filepath = os.path.join(self.reports_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            return {
                "report_type": "executive_summary",
                "filename": filename,
                "filepath": filepath,
                "summary_points": len(key_findings) + len(recommendations)
            }
            
        except Exception as e:
            self.logger.warning(f"エグゼクティブサマリー生成エラー: {e}")
            return None
