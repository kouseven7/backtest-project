"""
バックテスト・実運用データ収集・統合システム
フェーズ4A3: バックテストvs実運用比較分析器
"""

import json
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, timedelta
import logging
import asyncio
import openpyxl
from openpyxl import load_workbook

class DataCollector:
    """データ収集・統合システム"""
    
    def __init__(self, config: Dict, logger):
        self.config = config
        self.logger = logger
        self.data_sources = config.get('data_sources', {})
    
    async def collect_all_data(self) -> Tuple[Dict, Dict]:
        """全データ収集"""
        try:
            self.logger.info("並行データ収集開始...")
            
            # 並行データ収集
            backtest_task = asyncio.create_task(self._collect_backtest_data())
            live_task = asyncio.create_task(self._collect_live_data())
            
            backtest_data, live_data = await asyncio.gather(backtest_task, live_task)
            
            self.logger.info(f"データ収集完了 - BT戦略数: {len(backtest_data.get('strategies', {}))}, Live戦略数: {len(live_data.get('strategies', {}))}")
            
            return backtest_data, live_data
            
        except Exception as e:
            self.logger.error(f"データ収集エラー: {e}")
            return {}, {}
    
    async def _collect_backtest_data(self) -> Dict:
        """バックテストデータ収集"""
        try:
            backtest_path = Path(self.data_sources.get('backtest_results_path', 'backtest_results/improved_results/'))
            
            if not backtest_path.exists():
                self.logger.warning(f"バックテスト結果パスが存在しません: {backtest_path}")
                return self._get_sample_backtest_data()
            
            strategies_data = {}
            
            # Excel結果ファイル処理
            excel_files = list(backtest_path.glob("*.xlsx"))
            if excel_files:
                self.logger.info(f"Excel結果ファイル {len(excel_files)}件を処理中...")
                
                for excel_file in excel_files[:5]:  # 最大5ファイルまで処理
                    try:
                        excel_data = self._load_excel_backtest_data(excel_file)
                        if excel_data:
                            strategies_data.update(excel_data)
                    except Exception as e:
                        self.logger.warning(f"Excel結果読み込みエラー [{excel_file}]: {e}")
            
            # JSONファイル処理
            json_files = list(backtest_path.glob("*.json"))
            for result_file in json_files:
                try:
                    with open(result_file, 'r', encoding='utf-8') as f:
                        strategy_result = json.load(f)
                    
                    strategy_name = result_file.stem
                    strategies_data[strategy_name] = self._normalize_backtest_data(strategy_result)
                    
                except Exception as e:
                    self.logger.warning(f"JSON結果読み込みエラー [{result_file}]: {e}")
            
            # データが空の場合はサンプルデータを生成
            if not strategies_data:
                self.logger.info("実データが見つからないため、サンプルデータを生成します")
                return self._get_sample_backtest_data()
            
            return {
                "type": "backtest",
                "timestamp": datetime.now(),
                "strategies": strategies_data,
                "data_quality": self._assess_data_quality(strategies_data),
                "date_range": self._extract_date_range(strategies_data)
            }
            
        except Exception as e:
            self.logger.error(f"バックテストデータ収集エラー: {e}")
            return self._get_sample_backtest_data()
    
    def _load_excel_backtest_data(self, excel_file: Path) -> Dict:
        """Excel形式のバックテスト結果読み込み"""
        try:
            # ファイル名から情報抽出
            filename = excel_file.stem
            parts = filename.split('_')
            
            if len(parts) >= 3:
                # improved_backtest_9101.T_20250731_113728.xlsx のような形式
                ticker = parts[2]
                strategy_name = f"Strategy_{ticker}"
            else:
                strategy_name = f"Strategy_{filename}"
            
            # Excelファイル読み込み
            workbook = load_workbook(excel_file, data_only=True)
            
            # 取引履歴シートから情報抽出
            trade_data = {}
            summary_data = {}
            
            if '取引履歴' in workbook.sheetnames:
                trades_df = pd.read_excel(excel_file, sheet_name='取引履歴')
                if not trades_df.empty:
                    # 基本メトリクス計算
                    total_trades = len(trades_df)
                    if '損益' in trades_df.columns:
                        pnl_values = trades_df['損益'].dropna()
                        winning_trades = len(pnl_values[pnl_values > 0])
                        total_pnl = pnl_values.sum()
                        win_rate = winning_trades / total_trades if total_trades > 0 else 0
                        
                        trade_data = {
                            "total_trades": total_trades,
                            "winning_trades": winning_trades,
                            "total_pnl": total_pnl,
                            "win_rate": win_rate,
                            "avg_trade": total_pnl / total_trades if total_trades > 0 else 0
                        }
            
            # パフォーマンスサマリーシートから情報抽出
            if 'パフォーマンスサマリー' in workbook.sheetnames:
                summary_df = pd.read_excel(excel_file, sheet_name='パフォーマンスサマリー')
                if not summary_df.empty and len(summary_df.columns) >= 2:
                    for _, row in summary_df.iterrows():
                        metric = row.iloc[0] if pd.notna(row.iloc[0]) else ""
                        value = row.iloc[1] if pd.notna(row.iloc[1]) else 0
                        
                        if isinstance(metric, str):
                            if "シャープレシオ" in metric:
                                summary_data["sharpe_ratio"] = float(value) if pd.notna(value) else 0
                            elif "最大ドローダウン" in metric:
                                summary_data["max_drawdown"] = float(value) if pd.notna(value) else 0
                            elif "総リターン" in metric:
                                summary_data["total_return"] = float(value) if pd.notna(value) else 0
            
            workbook.close()
            
            # データ統合
            if trade_data or summary_data:
                normalized_data = {
                    "basic_metrics": trade_data,
                    "risk_metrics": summary_data,
                    "source_file": str(excel_file),
                    "last_modified": datetime.fromtimestamp(excel_file.stat().st_mtime)
                }
                
                return {strategy_name: normalized_data}
            
            return {}
            
        except Exception as e:
            self.logger.warning(f"Excel読み込みエラー [{excel_file}]: {e}")
            return {}
    
    async def _collect_live_data(self) -> Dict:
        """実運用データ収集"""
        try:
            live_path = Path(self.data_sources.get('live_trading_path', 'logs/paper_trading/'))
            performance_path = Path(self.data_sources.get('performance_monitoring_path', 'logs/performance_monitoring/'))
            
            strategies_data = {}
            
            # ペーパートレードログ収集
            if live_path.exists() and any(live_path.iterdir()):
                self.logger.info("ペーパートレードログを収集中...")
                for log_file in live_path.glob("*.json"):
                    try:
                        with open(log_file, 'r', encoding='utf-8') as f:
                            trading_log = json.load(f)
                        
                        strategy_data = self._normalize_live_data(trading_log)
                        if strategy_data:
                            strategy_name = log_file.stem.replace('_trading_log', '')
                            strategies_data[strategy_name] = strategy_data
                            
                    except Exception as e:
                        self.logger.warning(f"ライブトレードログ読み込みエラー [{log_file}]: {e}")
            
            # パフォーマンス監視データ収集
            if performance_path.exists() and any(performance_path.iterdir()):
                self.logger.info("パフォーマンス監視データを収集中...")
                performance_data = self._collect_performance_monitoring_data(performance_path)
                if performance_data:
                    strategies_data.update(performance_data)
            
            # データが空の場合はサンプルデータを生成
            if not strategies_data:
                self.logger.info("実運用データが見つからないため、サンプルデータを生成します")
                strategies_data = self._generate_sample_live_data()
            
            return {
                "type": "live",
                "timestamp": datetime.now(),
                "strategies": strategies_data,
                "data_quality": self._assess_data_quality(strategies_data),
                "date_range": self._extract_date_range(strategies_data)
            }
            
        except Exception as e:
            self.logger.error(f"実運用データ収集エラー: {e}")
            return {
                "type": "live",
                "timestamp": datetime.now(),
                "strategies": self._generate_sample_live_data(),
                "data_quality": "sample",
                "date_range": {"start": datetime.now() - timedelta(days=30), "end": datetime.now()}
            }
    
    def _collect_performance_monitoring_data(self, performance_path: Path) -> Dict:
        """パフォーマンス監視データ収集"""
        try:
            strategies_data = {}
            
            # 最新のパフォーマンス分析ファイルを取得
            json_files = list(performance_path.glob("performance_analysis_*.json"))
            if json_files:
                # 最新ファイルを取得
                latest_file = max(json_files, key=lambda x: x.stat().st_mtime)
                
                with open(latest_file, 'r', encoding='utf-8') as f:
                    performance_data = json.load(f)
                
                # ポートフォリオ分析からの戦略データ抽出
                portfolio_analysis = performance_data.get('portfolio_analysis', {})
                strategy_performances = portfolio_analysis.get('strategy_performances', {})
                
                for strategy_name, performance in strategy_performances.items():
                    normalized_data = {
                        "basic_metrics": performance.get('basic_metrics', {}),
                        "risk_metrics": performance.get('risk_metrics', {}),
                        "performance_score": performance.get('performance_score', 0),
                        "source": "performance_monitor",
                        "timestamp": performance.get('timestamp', datetime.now())
                    }
                    strategies_data[strategy_name] = normalized_data
            
            return strategies_data
            
        except Exception as e:
            self.logger.warning(f"パフォーマンス監視データ収集エラー: {e}")
            return {}
    
    def collect_backtest_data_sync(self, strategy_name: str = None) -> Dict:
        """バックテストデータ同期収集"""
        try:
            backtest_path = Path(self.data_sources.get('backtest_results_path', 'backtest_results/improved_results/'))
            
            if not backtest_path.exists():
                return self._get_sample_backtest_data().get('strategies', {})
            
            strategies_data = {}
            
            if strategy_name:
                # 特定戦略のみ
                result_file = backtest_path / f"{strategy_name}.json"
                if result_file.exists():
                    with open(result_file, 'r', encoding='utf-8') as f:
                        strategy_result = json.load(f)
                    return {strategy_name: self._normalize_backtest_data(strategy_result)}
                else:
                    # Excel結果から探索
                    for excel_file in backtest_path.glob(f"*{strategy_name}*.xlsx"):
                        excel_data = self._load_excel_backtest_data(excel_file)
                        if excel_data:
                            return excel_data
                    return {}
            else:
                # 全戦略（制限付き）
                excel_files = list(backtest_path.glob("*.xlsx"))[:3]  # 最大3ファイル
                for excel_file in excel_files:
                    try:
                        excel_data = self._load_excel_backtest_data(excel_file)
                        if excel_data:
                            strategies_data.update(excel_data)
                    except Exception as e:
                        self.logger.warning(f"Excel結果読み込みエラー [{excel_file}]: {e}")
                
                # JSONファイルも処理
                for result_file in backtest_path.glob("*.json"):
                    try:
                        with open(result_file, 'r', encoding='utf-8') as f:
                            strategy_result = json.load(f)
                        
                        strategy_name = result_file.stem
                        strategies_data[strategy_name] = self._normalize_backtest_data(strategy_result)
                        
                    except Exception as e:
                        self.logger.warning(f"JSON結果読み込みエラー [{result_file}]: {e}")
                
                return strategies_data if strategies_data else self._get_sample_backtest_data().get('strategies', {})
                
        except Exception as e:
            self.logger.error(f"バックテストデータ同期収集エラー: {e}")
            return {}
    
    def collect_live_data_sync(self, strategy_name: str = None) -> Dict:
        """実運用データ同期収集"""
        try:
            live_path = Path(self.data_sources.get('live_trading_path', 'logs/paper_trading/'))
            
            if not live_path.exists() or not any(live_path.iterdir()):
                # サンプルデータ生成（テスト用）
                return self._generate_sample_live_data(strategy_name)
            
            strategies_data = {}
            
            if strategy_name:
                log_file = live_path / f"{strategy_name}_trading_log.json"
                if log_file.exists():
                    with open(log_file, 'r', encoding='utf-8') as f:
                        trading_log = json.load(f)
                    
                    strategy_data = self._normalize_live_data(trading_log)
                    if strategy_data:
                        strategies_data[strategy_name] = strategy_data
            else:
                for log_file in live_path.glob("*.json"):
                    try:
                        with open(log_file, 'r', encoding='utf-8') as f:
                            trading_log = json.load(f)
                        
                        strategy_data = self._normalize_live_data(trading_log)
                        if strategy_data:
                            strategy_name = log_file.stem.replace('_trading_log', '')
                            strategies_data[strategy_name] = strategy_data
                            
                    except Exception as e:
                        self.logger.warning(f"ライブトレードログ読み込みエラー [{log_file}]: {e}")
            
            return strategies_data if strategies_data else self._generate_sample_live_data()
            
        except Exception as e:
            self.logger.error(f"実運用データ同期収集エラー: {e}")
            return {}
    
    def _normalize_backtest_data(self, raw_data: Dict) -> Dict:
        """バックテストデータ正規化"""
        try:
            normalized = {
                "basic_metrics": {},
                "risk_metrics": {},
                "trades": [],
                "source": "backtest"
            }
            
            # 基本メトリクス抽出
            if isinstance(raw_data, dict):
                normalized["basic_metrics"] = {
                    "total_trades": raw_data.get("total_trades", 0),
                    "winning_trades": raw_data.get("winning_trades", 0),
                    "total_pnl": raw_data.get("total_pnl", 0),
                    "win_rate": raw_data.get("win_rate", 0)
                }
                
                normalized["risk_metrics"] = {
                    "max_drawdown": raw_data.get("max_drawdown", 0),
                    "sharpe_ratio": raw_data.get("sharpe_ratio", 0),
                    "volatility": raw_data.get("volatility", 0)
                }
            
            return normalized
            
        except Exception as e:
            self.logger.warning(f"バックテストデータ正規化エラー: {e}")
            return {"basic_metrics": {}, "risk_metrics": {}, "source": "backtest"}
    
    def _normalize_live_data(self, raw_data: Dict) -> Dict:
        """実運用データ正規化"""
        try:
            normalized = {
                "basic_metrics": {},
                "risk_metrics": {},
                "trades": [],
                "source": "live"
            }
            
            # 実運用データから基本メトリクス抽出
            if isinstance(raw_data, dict):
                trades = raw_data.get("trades", [])
                if trades:
                    total_trades = len(trades)
                    pnl_list = [trade.get("pnl", 0) for trade in trades if "pnl" in trade]
                    winning_trades = len([pnl for pnl in pnl_list if pnl > 0])
                    total_pnl = sum(pnl_list)
                    
                    normalized["basic_metrics"] = {
                        "total_trades": total_trades,
                        "winning_trades": winning_trades,
                        "total_pnl": total_pnl,
                        "win_rate": winning_trades / total_trades if total_trades > 0 else 0
                    }
                    
                    # リスクメトリクス計算
                    if pnl_list:
                        returns = np.array(pnl_list)
                        normalized["risk_metrics"] = {
                            "volatility": np.std(returns) if len(returns) > 1 else 0,
                            "max_drawdown": min(returns) if returns.size > 0 else 0,
                            "sharpe_ratio": np.mean(returns) / np.std(returns) if len(returns) > 1 and np.std(returns) > 0 else 0
                        }
            
            return normalized
            
        except Exception as e:
            self.logger.warning(f"実運用データ正規化エラー: {e}")
            return {"basic_metrics": {}, "risk_metrics": {}, "source": "live"}
    
    def _get_sample_backtest_data(self) -> Dict:
        """サンプルバックテストデータ生成"""
        import random
        
        sample_strategies = ["VWAP_Breakout", "Momentum_Investing", "Opening_Gap"]
        strategies_data = {}
        
        for strategy_name in sample_strategies:
            strategies_data[strategy_name] = {
                "basic_metrics": {
                    "total_trades": random.randint(20, 100),
                    "winning_trades": random.randint(10, 50),
                    "total_pnl": random.uniform(-5000, 15000),
                    "win_rate": random.uniform(0.3, 0.7)
                },
                "risk_metrics": {
                    "volatility": random.uniform(0.1, 0.4),
                    "max_drawdown": random.uniform(-0.3, -0.05),
                    "sharpe_ratio": random.uniform(-0.5, 2.5)
                },
                "source": "backtest_sample",
                "timestamp": datetime.now()
            }
        
        return {
            "type": "backtest",
            "timestamp": datetime.now(),
            "strategies": strategies_data,
            "data_quality": "sample",
            "date_range": {"start": datetime.now() - timedelta(days=90), "end": datetime.now()}
        }
    
    def _generate_sample_live_data(self, strategy_name: str = None) -> Dict:
        """サンプル実運用データ生成"""
        import random
        
        if strategy_name:
            sample_strategies = [strategy_name]
        else:
            sample_strategies = ["VWAP_Breakout", "Momentum_Investing", "Opening_Gap"]
        
        strategies_data = {}
        
        for name in sample_strategies:
            # バックテストと微妙に異なる結果を生成（実運用の現実性を模擬）
            bt_trades = random.randint(20, 100)
            live_trades = max(1, int(bt_trades * random.uniform(0.6, 0.9)))  # 実運用は取引数が少ない傾向
            
            strategies_data[name] = {
                "basic_metrics": {
                    "total_trades": live_trades,
                    "winning_trades": random.randint(1, live_trades),
                    "total_pnl": random.uniform(-3000, 8000),  # バックテストより控えめ
                    "win_rate": random.uniform(0.25, 0.65)  # やや低め
                },
                "risk_metrics": {
                    "volatility": random.uniform(0.15, 0.5),  # やや高め
                    "max_drawdown": random.uniform(-0.35, -0.08),  # やや深め
                    "sharpe_ratio": random.uniform(-0.8, 2.0)  # やや低め
                },
                "source": "live_sample",
                "timestamp": datetime.now()
            }
        
        return strategies_data
    
    def _assess_data_quality(self, strategies_data: Dict) -> str:
        """データ品質評価"""
        try:
            if not strategies_data:
                return "no_data"
            
            quality_scores = []
            for strategy_name, data in strategies_data.items():
                score = 0
                basic_metrics = data.get('basic_metrics', {})
                
                # 取引数チェック
                total_trades = basic_metrics.get('total_trades', 0)
                if total_trades >= 50:
                    score += 0.4
                elif total_trades >= 20:
                    score += 0.3
                elif total_trades >= 10:
                    score += 0.2
                
                # PnLデータの有無
                if basic_metrics.get('total_pnl') is not None:
                    score += 0.3
                
                # リスクメトリクスの有無
                risk_metrics = data.get('risk_metrics', {})
                if risk_metrics:
                    score += 0.3
                
                quality_scores.append(score)
            
            avg_score = np.mean(quality_scores) if quality_scores else 0
            
            if avg_score >= 0.8:
                return "high"
            elif avg_score >= 0.6:
                return "medium"
            elif avg_score >= 0.3:
                return "low"
            else:
                return "poor"
                
        except Exception as e:
            self.logger.warning(f"データ品質評価エラー: {e}")
            return "unknown"
    
    def _extract_date_range(self, strategies_data: Dict) -> Dict:
        """データ期間抽出"""
        try:
            timestamps = []
            for strategy_name, data in strategies_data.items():
                if 'timestamp' in data:
                    timestamps.append(data['timestamp'])
                elif 'last_modified' in data:
                    timestamps.append(data['last_modified'])
            
            if timestamps:
                return {
                    "start": min(timestamps),
                    "end": max(timestamps)
                }
            else:
                return {
                    "start": datetime.now() - timedelta(days=30),
                    "end": datetime.now()
                }
                
        except Exception as e:
            self.logger.warning(f"日付範囲抽出エラー: {e}")
            return {
                "start": datetime.now() - timedelta(days=30),
                "end": datetime.now()
            }
