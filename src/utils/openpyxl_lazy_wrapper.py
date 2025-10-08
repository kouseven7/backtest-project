#!/usr/bin/env python3
"""
openpyxl遅延インポートラッパー
TODO-PERF-001: Phase 1 Stage 3実装

openpyxlの初回Excel出力時のみ遅延を発生させ、
2回目以降は高速アクセスを提供する。
"""

import importlib.util
import sys
import time
from typing import Any, Optional, Dict

class OpenpyxlLazyWrapper:
    """openpyxl遅延インポートラッパー"""
    
    def __init__(self):
        self._openpyxl = None
        self._import_time = None
        self._first_access = True
        
    def _import_openpyxl(self) -> Any:
        """openpyxl実際のインポート（初回のみ）"""
        if self._openpyxl is None:
            start_time = time.perf_counter()
            
            try:
                import openpyxl
                self._openpyxl = openpyxl
                self._import_time = (time.perf_counter() - start_time) * 1000
                
                if self._first_access:
                    print(f"[INFO] openpyxl lazy import: {self._import_time:.1f}ms")
                    self._first_access = False
                    
            except ImportError as e:
                print(f"[ERROR] openpyxl import error: {e}")
                # SystemFallbackPolicy統合
                try:
                    from src.config.system_modes import SystemFallbackPolicy, ComponentType
                    fallback_policy = SystemFallbackPolicy()
                    return fallback_policy.handle_component_failure(
                        component_type=ComponentType.MULTI_STRATEGY,
                        component_name="openpyxl_lazy_wrapper",
                        error=e,
                        fallback_func=lambda: None
                    )
                except ImportError:
                    pass
                raise
                
        return self._openpyxl
    
    def __getattr__(self, name: str) -> Any:
        """openpyxlの属性・メソッドに透明アクセス"""
        openpyxl = self._import_openpyxl()
        return getattr(openpyxl, name)
    
    # よく使用されるクラス・関数の直接実装
    def Workbook(self, *args, **kwargs):
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: """openpyxl.Workbook()の遅延ラッパー"""
        openpyxl = self._import_openpyxl()
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: return openpyxl.Workbook(*args, **kwargs)
    
    def load_workbook(self, *args, **kwargs):
        """openpyxl.load_workbook()の遅延ラッパー"""
        openpyxl = self._import_openpyxl()
        return openpyxl.load_workbook(*args, **kwargs)
    
    def get_import_stats(self) -> Dict[str, Any]:
        """インポート統計取得"""
        return {
            'imported': self._openpyxl is not None,
            'import_time_ms': self._import_time,
            'first_access_completed': not self._first_access
        }

# グローバルインスタンス
_lazy_openpyxl = OpenpyxlLazyWrapper()

# openpyxlのAPIをエクスポート
def Workbook(*args, **kwargs):
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: return _lazy_openpyxl.Workbook(*args, **kwargs)

def load_workbook(*args, **kwargs):
    return _lazy_openpyxl.load_workbook(*args, **kwargs)

# 統計情報エクスポート
def get_openpyxl_import_stats():
    return _lazy_openpyxl.get_import_stats()

# 属性アクセス用
def __getattr__(name: str):
    return getattr(_lazy_openpyxl, name)
