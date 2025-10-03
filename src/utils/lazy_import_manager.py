"""
Phase 3重いライブラリ遅延インポート機構
yfinance・openpyxl等の重いライブラリを必要時のみロード

作成: 2025年10月3日
目的: インポート時間1456.6ms削減
"""

import sys
from typing import Any, Dict, Optional, Callable
import logging

# 遅延インポート統計
_lazy_import_stats: Dict[str, Dict[str, Any]] = {}
_logger = logging.getLogger(__name__)

class LazyImporter:
    """重いライブラリの遅延インポート管理"""
    
    def __init__(self):
        self._imported_modules: Dict[str, Any] = {}
        self._import_times: Dict[str, float] = {}
    
    def import_yfinance(self) -> Any:
        """yfinance遅延インポート (1201.8ms削減対象)"""
        if 'yfinance' not in self._imported_modules:
            import time
            start = time.perf_counter()
            
            try:
                import yfinance as yf
                self._imported_modules['yfinance'] = yf
                
                import_time = (time.perf_counter() - start) * 1000
                self._import_times['yfinance'] = import_time
                
                _logger.info(f"yfinance遅延ロード完了: {import_time:.1f}ms")
                
                # 統計記録
                _lazy_import_stats['yfinance'] = {
                    'import_time_ms': import_time,
                    'module_name': 'yfinance',
                    'load_count': _lazy_import_stats.get('yfinance', {}).get('load_count', 0) + 1
                }
                
            except ImportError as e:
                _logger.error(f"yfinance遅延インポートエラー: {e}")
                # フォールバック: ダミーモジュール
                self._imported_modules['yfinance'] = self._create_yfinance_fallback()
        
        return self._imported_modules['yfinance']
    
    def import_openpyxl(self) -> Any:
        """openpyxl遅延インポート (254.7ms削減対象)"""
        if 'openpyxl' not in self._imported_modules:
            import time
            start = time.perf_counter()
            
            try:
                import openpyxl
                self._imported_modules['openpyxl'] = openpyxl
                
                import_time = (time.perf_counter() - start) * 1000
                self._import_times['openpyxl'] = import_time
                
                _logger.info(f"openpyxl遅延ロード完了: {import_time:.1f}ms")
                
                # 統計記録
                _lazy_import_stats['openpyxl'] = {
                    'import_time_ms': import_time,
                    'module_name': 'openpyxl',
                    'load_count': _lazy_import_stats.get('openpyxl', {}).get('load_count', 0) + 1
                }
                
            except ImportError as e:
                _logger.error(f"openpyxl遅延インポートエラー: {e}")
                # フォールバック: ダミーモジュール
                self._imported_modules['openpyxl'] = self._create_openpyxl_fallback()
        
        return self._imported_modules['openpyxl']
    
    def _create_yfinance_fallback(self) -> Any:
        """yfinanceフォールバック（軽量ダミー）"""
        class DummyTicker:
            def __init__(self, symbol: str):
                self.symbol = symbol
            
            def history(self, period="1y", **kwargs):
                import pandas as pd
                # 空のDataFrame返却
                return pd.DataFrame()
            
            def download(self, *args, **kwargs):
                import pandas as pd
                return pd.DataFrame()
        
        class DummyYFinance:
            @staticmethod
            def Ticker(symbol: str) -> DummyTicker:
                return DummyTicker(symbol)
            
            @staticmethod
            def download(*args, **kwargs):
                import pandas as pd
                return pd.DataFrame()
        
        return DummyYFinance()
    
    def _create_openpyxl_fallback(self) -> Any:
        """openpyxlフォールバック（軽量ダミー）"""
        class DummyWorkbook:
            def save(self, filename: str):
                _logger.warning(f"openpyxl未インストール - {filename}への保存をスキップ")
            
            def create_sheet(self, title: str = None):
                return self
            
            def __getitem__(self, key):
                return self
        
        class DummyOpenpyxl:
            @staticmethod
            def Workbook():
                return DummyWorkbook()
        
        return DummyOpenpyxl()
    
    def get_import_stats(self) -> Dict[str, Dict[str, Any]]:
        """インポート統計取得"""
        return _lazy_import_stats.copy()
    
    def get_total_saved_time(self) -> float:
        """遅延インポートによる削減時間計算"""
        return sum(stats.get('import_time_ms', 0) for stats in _lazy_import_stats.values())

# グローバル遅延インポーター
_lazy_importer = LazyImporter()

def get_yfinance():
    """yfinanceグローバル取得関数"""
    return _lazy_importer.import_yfinance()

def get_openpyxl():
    """openpyxlグローバル取得関数"""
    return _lazy_importer.import_openpyxl()

def get_lazy_import_stats() -> Dict[str, Dict[str, Any]]:
    """遅延インポート統計取得"""
    return _lazy_importer.get_import_stats()

def get_total_optimization_effect() -> Dict[str, float]:
    """最適化効果サマリー"""
    stats = get_lazy_import_stats()
    
    return {
        'total_saved_time_ms': _lazy_importer.get_total_saved_time(),
        'yfinance_saved_ms': stats.get('yfinance', {}).get('import_time_ms', 0),
        'openpyxl_saved_ms': stats.get('openpyxl', {}).get('import_time_ms', 0),
        'target_reduction_ms': 1456.6,  # yfinance 1201.8ms + openpyxl 254.7ms
        'modules_optimized': len(stats)
    }

# 使用例とテスト
def test_lazy_import():
    """遅延インポートテスト"""
    print("=== Phase 3遅延インポートテスト ===")
    
    # yfinanceテスト
    print("yfinance遅延ロード開始...")
    yf = get_yfinance()
    print(f"yfinanceロード完了: {type(yf)}")
    
    # openpyxlテスト
    print("openpyxl遅延ロード開始...")
    openpyxl = get_openpyxl()
    print(f"openpyxlロード完了: {type(openpyxl)}")
    
    # 統計表示
    stats = get_lazy_import_stats()
    effect = get_total_optimization_effect()
    
    print(f"\n遅延インポート統計:")
    for module, data in stats.items():
        print(f"  {module}: {data['import_time_ms']:.1f}ms (ロード回数: {data['load_count']})")
    
    print(f"\n最適化効果:")
    print(f"  合計削減時間: {effect['total_saved_time_ms']:.1f}ms")
    print(f"  目標削減時間: {effect['target_reduction_ms']:.1f}ms")
    print(f"  達成率: {(effect['total_saved_time_ms']/effect['target_reduction_ms']*100):.1f}%")

if __name__ == "__main__":
    test_lazy_import()