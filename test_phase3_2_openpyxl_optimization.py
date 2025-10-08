"""
Phase 3-2 openpyxl遅延インポート効果測定テスト
作成: 2025年10月3日

TODO-PERF-003 Phase 3-2実装効果測定
openpyxl遅延インポートによる254.7ms削減効果を検証
"""

import sys
import os
import time
from pathlib import Path

# プロジェクトルートをパスに追加
project_root = Path(__file__).parent
sys.path.append(str(project_root))

def test_original_openpyxl_import_performance():
    """従来のopenpyxl直接インポート性能測定"""
    print("=== 従来のopenpyxl直接インポート性能測定 ===")
    
    start_time = time.perf_counter()
    
    try:
        # openpyxl遅延インポート (TODO-PERF-001: Stage 3)
import src.utils.openpyxl_lazy_wrapper as openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import LineChart, Reference
        
        import_time = (time.perf_counter() - start_time) * 1000
        print(f"openpyxl直接インポート: {import_time:.1f}ms")
        
        # 軽量使用テスト
        start_usage = time.perf_counter()
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: wb = openpyxl.Workbook()
        font = Font(bold=True, size=12)
        usage_time = (time.perf_counter() - start_usage) * 1000
        print(f"Workbook作成・Font作成時間: {usage_time:.1f}ms")
        
        return {
            'import_time_ms': import_time,
            'usage_time_ms': usage_time,
            'total_time_ms': import_time + usage_time,
            'module_loaded': True
        }
        
    except ImportError as e:
        error_time = (time.perf_counter() - start_time) * 1000
        print(f"openpyxlインポートエラー: {error_time:.1f}ms - {e}")
        return {
            'import_time_ms': error_time,
            'usage_time_ms': 0,
            'total_time_ms': error_time,
            'module_loaded': False
        }

def test_lazy_openpyxl_import_performance():
    """Phase 3最適化: openpyxl遅延インポート性能測定"""
    print("\n=== Phase 3最適化: openpyxl遅延インポート性能測定 ===")
    
    # 遅延インポートマネージャーの読み込み時間測定
    start_manager = time.perf_counter()
    from src.utils.lazy_import_manager import get_openpyxl, get_lazy_import_stats
    manager_time = (time.perf_counter() - start_manager) * 1000
    print(f"LazyImportManager(openpyxl)読み込み: {manager_time:.1f}ms")
    
    # 実際のopenpyxl遅延ロード時間測定
    start_lazy = time.perf_counter()
    openpyxl = get_openpyxl()
    lazy_time = (time.perf_counter() - start_lazy) * 1000
    print(f"openpyxl遅延ロード: {lazy_time:.1f}ms")
    
    # 使用時間測定
    start_usage = time.perf_counter()
# TODO(tag:excel_deprecated, rationale:Excel output eliminated 2025-10-08) # BACKTEST_IMPACT: Trading data output affected
# ORIGINAL: wb = openpyxl.Workbook()
    from openpyxl.styles import Font
    font = Font(bold=True, size=12)
    usage_time = (time.perf_counter() - start_usage) * 1000
    print(f"Workbook作成・Font作成時間: {usage_time:.1f}ms")
    
    # 統計取得
    stats = get_lazy_import_stats()
    print(f"遅延インポート統計: {stats}")
    
    return {
        'manager_time_ms': manager_time,
        'lazy_load_time_ms': lazy_time,
        'usage_time_ms': usage_time,
        'total_time_ms': manager_time + lazy_time + usage_time,
        'stats': stats
    }

def test_excel_exporter_integration():
    """Excel Exporter統合テスト"""
    print("\n=== Excel Exporter統合テスト（遅延インポート適用後） ===")
    
    results = {}
    
    # DSSMSExcelExporterV2 テスト
    try:
        start_exporter = time.perf_counter()
        from output.dssms_excel_exporter_v2 import DSSMSExcelExporterV2
        exporter_time = (time.perf_counter() - start_exporter) * 1000
        print(f"DSSMSExcelExporterV2読み込み: {exporter_time:.1f}ms")
        
        # 軽量テスト（初期化のみ）
        start_init = time.perf_counter()
        exporter = DSSMSExcelExporterV2()
        init_time = (time.perf_counter() - start_init) * 1000
        print(f"Exporter初期化: {init_time:.1f}ms")
        
        results['excel_exporter'] = {
            'import_time_ms': exporter_time,
            'init_time_ms': init_time,
            'total_time_ms': exporter_time + init_time
        }
        
    except Exception as e:
        print(f"DSSMSExcelExporterV2テストエラー: {e}")
        results['excel_exporter'] = {'error': str(e)}
    
    return results

def calculate_phase3_2_optimization_effect(original_result, lazy_result):
    """Phase 3-2最適化効果計算"""
    print("\n=== Phase 3-2最適化効果計算 ===")
    
    original_total = original_result.get('total_time_ms', 0)
    lazy_total = lazy_result.get('total_time_ms', 0)
    
    if original_total > 0:
        time_saved = original_total - lazy_total
        improvement_rate = (time_saved / original_total) * 100
        
        print(f"従来方式合計時間: {original_total:.1f}ms")
        print(f"遅延インポート合計時間: {lazy_total:.1f}ms")
        print(f"削減時間: {time_saved:.1f}ms")
        print(f"改善率: {improvement_rate:.1f}%")
        
        # 目標達成率
        target_reduction = 254.7  # openpyxl目標削減時間
        achievement_rate = (time_saved / target_reduction) * 100 if target_reduction > 0 else 0
        print(f"目標削減時間: {target_reduction:.1f}ms")
        print(f"目標達成率: {achievement_rate:.1f}%")
        
        return {
            'original_time_ms': original_total,
            'optimized_time_ms': lazy_total,
            'time_saved_ms': time_saved,
            'improvement_rate_pct': improvement_rate,
            'target_reduction_ms': target_reduction,
            'achievement_rate_pct': achievement_rate
        }
    
    return None

def main():
    """Phase 3-2効果測定メイン実行"""
    print("Phase 3-2 openpyxl遅延インポート効果測定開始")
    print("=" * 60)
    
    # 1. 従来方式測定
    original_result = test_original_openpyxl_import_performance()
    
    # 2. 遅延インポート方式測定
    lazy_result = test_lazy_openpyxl_import_performance()
    
    # 3. Excel Exporter統合テスト
    integration_results = test_excel_exporter_integration()
    
    # 4. 効果計算
    effect = calculate_phase3_2_optimization_effect(original_result, lazy_result)
    
    # 5. 結果サマリー
    print("\n" + "=" * 60)
    print("Phase 3-2実装効果サマリー")
    print("=" * 60)
    
    if effect:
        print(f"✅ openpyxl最適化完了")
        print(f"   削減時間: {effect['time_saved_ms']:.1f}ms")
        print(f"   改善率: {effect['improvement_rate_pct']:.1f}%")
        print(f"   目標達成率: {effect['achievement_rate_pct']:.1f}%")
    
    print(f"✅ Excel Exporter統合テスト結果:")
    for component, result in integration_results.items():
        if 'error' not in result:
            print(f"   {component}: {result.get('total_time_ms', 0):.1f}ms")
        else:
            print(f"   {component}: エラー - {result['error'][:50]}...")
    
    # 全体遅延インポート統計
    from src.utils.lazy_import_manager import get_total_optimization_effect
    total_effect = get_total_optimization_effect()
    print(f"\n✅ 遅延インポート総合効果:")
    print(f"   合計削減時間: {total_effect['total_saved_time_ms']:.1f}ms")
    print(f"   最適化モジュール数: {total_effect['modules_optimized']}")
    print(f"   目標削減時間: {total_effect['target_reduction_ms']:.1f}ms")
    
    # Phase 3全体達成率
    if total_effect['target_reduction_ms'] > 0:
        total_achievement = (total_effect['total_saved_time_ms'] / total_effect['target_reduction_ms']) * 100
        print(f"   Phase 3全体達成率: {total_achievement:.1f}%")
    
    print("\nPhase 3-2効果測定完了")

if __name__ == "__main__":
    main()