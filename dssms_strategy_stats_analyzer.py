#!/usr/bin/env python3
"""
DSSMS 戦略別統計問題分析スクリプト

問題:
1. 戦略別統計シートに7つの戦略が表示されない（DSSMSのみ）
2. 勝率が正しく計算されていない
3. マルチ戦略システムの統計が反映されていない

目的:
- Excelファイルの戦略別統計シートを詳細分析
- JSONデータの戦略統計を確認
- 統一出力システムの戦略別データ生成ロジックを検証
- 問題箇所を特定し修正方針を提示
"""

import pandas as pd
import json
import logging
from datetime import datetime
from pathlib import Path
import openpyxl
from typing import Dict, List, Any, Optional

def setup_logger():
    """ロガー設定"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

class DSSMSStrategyStatsAnalyzer:
    """DSSMS戦略別統計分析器"""
    
    def __init__(self):
        self.logger = setup_logger()
        self.issues = []
        self.analysis_results = {}
        
        # 期待する7つの戦略名
        self.expected_strategies = [
            'VWAPBreakoutStrategy',
            'MeanReversionStrategy', 
            'TrendFollowingStrategy',
            'MomentumStrategy',
            'ContrarianStrategy',
            'VolatilityBreakoutStrategy',
            'RSIStrategy'
        ]
        
    def analyze_excel_strategy_stats(self, excel_path: str) -> Dict[str, Any]:
        """Excelファイルの戦略別統計シートを分析"""
        self.logger.info(f"📊 Excel戦略別統計分析開始: {excel_path}")
        
        try:
            # Excelファイルを読み込み
            workbook = openpyxl.load_workbook(excel_path)
            
            # 全シート名を確認
            sheet_names = workbook.sheetnames
            self.logger.info(f"利用可能シート: {sheet_names}")
            
            # 戦略別統計シートを探す
            strategy_sheet_name = None
            possible_names = ['戦略別統計', 'Strategy Statistics', '戦略統計', 'ストラテジー統計']
            
            for name in possible_names:
                if name in sheet_names:
                    strategy_sheet_name = name
                    break
            
            if not strategy_sheet_name:
                self.logger.warning("戦略別統計シートが見つかりません")
                return {'error': '戦略別統計シートが存在しません', 'available_sheets': sheet_names}
            
            self.logger.info(f"戦略別統計シート発見: {strategy_sheet_name}")
            
            # 戦略別統計シートを分析
            sheet = workbook[strategy_sheet_name]
            
            # データを抽出
            strategy_data = []
            headers = []
            
            # ヘッダー行を取得
            for cell in sheet[1]:
                if cell.value:
                    headers.append(cell.value)
            
            self.logger.info(f"戦略統計ヘッダー: {headers}")
            
            # データ行を取得
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # 空行をスキップ
                    strategy_data.append(row)
            
            self.logger.info(f"戦略データ件数: {len(strategy_data)}件")
            
            # 戦略名を抽出
            strategy_names = []
            if strategy_data and len(strategy_data[0]) > 0:
                for row in strategy_data:
                    if row[0]:  # 最初の列が戦略名と仮定
                        strategy_names.append(str(row[0]))
            
            unique_strategies = list(set(strategy_names))
            
            analysis = {
                'sheet_found': True,
                'sheet_name': strategy_sheet_name,
                'headers': headers,
                'total_rows': len(strategy_data),
                'strategy_names': strategy_names,
                'unique_strategies': unique_strategies,
                'strategy_count': len(unique_strategies),
                'sample_data': strategy_data[:5] if strategy_data else [],
                'issues_found': []
            }
            
            # 問題チェック
            if len(unique_strategies) == 1 and 'DSSMS' in unique_strategies[0]:
                self.issues.append({
                    'category': 'single_strategy',
                    'severity': 'high',
                    'description': '戦略別統計に7つの戦略ではなくDSSMSのみが表示されている',
                    'details': {'found_strategies': unique_strategies, 'expected_count': 7}
                })
            
            if len(unique_strategies) < 7:
                self.issues.append({
                    'category': 'missing_strategies',
                    'severity': 'high', 
                    'description': f'期待される7つの戦略のうち{7 - len(unique_strategies)}つが不足',
                    'details': {
                        'found_strategies': unique_strategies,
                        'expected_strategies': self.expected_strategies,
                        'missing_count': 7 - len(unique_strategies)
                    }
                })
            
            # 勝率列をチェック
            win_rate_col = None
            for i, header in enumerate(headers):
                if '勝率' in str(header) or 'Win Rate' in str(header):
                    win_rate_col = i
                    break
            
            if win_rate_col is not None:
                win_rates = []
                for row in strategy_data:
                    if len(row) > win_rate_col and row[win_rate_col] is not None:
                        try:
                            rate = float(str(row[win_rate_col]).replace('%', ''))
                            win_rates.append(rate)
                        except (ValueError, TypeError):
                            pass
                
                analysis['win_rate_analysis'] = {
                    'win_rate_column': win_rate_col,
                    'win_rates': win_rates,
                    'unique_win_rates': list(set(win_rates)),
                    'all_zero': all(rate == 0 for rate in win_rates) if win_rates else True
                }
                
                if all(rate == 0 for rate in win_rates) if win_rates else True:
                    self.issues.append({
                        'category': 'zero_win_rates',
                        'severity': 'high',
                        'description': 'すべての戦略の勝率が0%になっている',
                        'details': analysis['win_rate_analysis']
                    })
            else:
                analysis['win_rate_analysis'] = {'error': '勝率列が見つかりません'}
                
            return analysis
            
        except Exception as e:
            self.logger.error(f"Excel分析エラー: {e}")
            return {'error': str(e)}
    
    def analyze_json_strategy_data(self, json_path: str) -> Dict[str, Any]:
        """JSONデータの戦略統計を分析"""
        self.logger.info(f"📦 JSON戦略データ分析開始: {json_path}")
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 戦略統計データを探す
            strategy_stats = data.get('strategy_statistics', {})
            trades_data = data.get('trades', [])
            
            analysis = {
                'strategy_statistics_exists': bool(strategy_stats),
                'strategy_statistics': strategy_stats,
                'trades_count': len(trades_data),
                'trades_sample': trades_data[:3] if trades_data else []
            }
            
            # 取引データから戦略を抽出
            if trades_data:
                strategies_in_trades = []
                for trade in trades_data:
                    if 'strategy' in trade:
                        strategies_in_trades.append(trade['strategy'])
                
                unique_strategies_in_trades = list(set(strategies_in_trades))
                analysis['strategies_in_trades'] = {
                    'unique_strategies': unique_strategies_in_trades,
                    'strategy_count': len(unique_strategies_in_trades),
                    'all_strategies': strategies_in_trades[:20]  # 最初の20件
                }
                
                # 戦略別統計を計算
                strategy_performance = {}
                for strategy in unique_strategies_in_trades:
                    strategy_trades = [t for t in trades_data if t.get('strategy') == strategy]
                    
                    if strategy_trades:
                        pnls = [t.get('pnl', 0) for t in strategy_trades]
                        winning_trades = len([p for p in pnls if p > 0])
                        total_trades = len(pnls)
                        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
                        
                        strategy_performance[strategy] = {
                            'total_trades': total_trades,
                            'winning_trades': winning_trades,
                            'win_rate': win_rate,
                            'total_pnl': sum(pnls),
                            'avg_pnl': sum(pnls) / total_trades if total_trades > 0 else 0
                        }
                
                analysis['calculated_strategy_performance'] = strategy_performance
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"JSON分析エラー: {e}")
            return {'error': str(e)}
    
    def analyze_text_report_strategies(self, text_path: str) -> Dict[str, Any]:
        """テキストレポートの戦略情報を分析"""
        self.logger.info(f"📄 テキストレポート戦略分析: {text_path}")
        
        try:
            with open(text_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 戦略関連の情報を抽出
            import re
            
            # 戦略名を抽出
            strategy_patterns = [
                r'戦略[：:]\s*([^\n]+)',
                r'Strategy[：:]\s*([^\n]+)',
                r'([A-Za-z]+Strategy)',
            ]
            
            found_strategies = []
            for pattern in strategy_patterns:
                matches = re.findall(pattern, content)
                found_strategies.extend(matches)
            
            # 勝率情報を抽出
            win_rate_pattern = r'勝率[：:]\s*([0-9.]+)%?'
            win_rate_matches = re.findall(win_rate_pattern, content)
            
            # 切替成功率を抽出
            switch_success_pattern = r'切替成功率[：:]\s*([0-9.]+)%?'
            switch_success_matches = re.findall(switch_success_pattern, content)
            
            analysis = {
                'found_strategies': found_strategies,
                'unique_strategies': list(set(found_strategies)),
                'win_rates_found': win_rate_matches,
                'switch_success_rates': switch_success_matches,
                'content_sample': content[:1000]
            }
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"テキスト分析エラー: {e}")
            return {'error': str(e)}
    
    def check_unified_output_engine(self) -> Dict[str, Any]:
        """統一出力エンジンのコードを確認"""
        self.logger.info("🔧 統一出力エンジンコード分析")
        
        try:
            engine_path = "src/dssms/unified_output_engine.py"
            with open(engine_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 戦略別統計生成部分を探す
            strategy_methods = []
            if '_generate_strategy_statistics' in content:
                strategy_methods.append('_generate_strategy_statistics')
            if 'strategy_statistics' in content:
                strategy_methods.append('strategy_statistics_処理')
            if '戦略別統計' in content:
                strategy_methods.append('戦略別統計_処理')
            
            # Excel生成部分を確認
            excel_generation_found = '_generate_excel_output' in content
            
            analysis = {
                'engine_file_exists': True,
                'strategy_methods_found': strategy_methods,
                'excel_generation_found': excel_generation_found,
                'file_size': len(content),
                'content_preview': content[:500]
            }
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"統一出力エンジン分析エラー: {e}")
            return {'error': str(e)}
    
    def generate_comprehensive_report(self, excel_path: str, json_path: str, text_path: str) -> str:
        """総合戦略別統計問題レポート生成"""
        self.logger.info("🔍 総合戦略別統計問題分析開始")
        
        # 各ファイルを分析
        excel_analysis = self.analyze_excel_strategy_stats(excel_path)
        json_analysis = self.analyze_json_strategy_data(json_path)
        text_analysis = self.analyze_text_report_strategies(text_path)
        engine_analysis = self.check_unified_output_engine()
        
        # 結果をまとめる
        self.analysis_results = {
            'excel_analysis': excel_analysis,
            'json_analysis': json_analysis,
            'text_analysis': text_analysis,
            'engine_analysis': engine_analysis,
            'issues_found': self.issues,
            'timestamp': datetime.now().isoformat()
        }
        
        # レポートファイルを生成
        report_path = f"dssms_strategy_stats_issue_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(self.analysis_results, f, ensure_ascii=False, indent=2, default=str)
        
        # サマリーレポートを生成
        summary = self._generate_strategy_summary_report()
        
        summary_path = f"dssms_strategy_stats_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary)
        
        self.logger.info(f"📋 詳細レポート: {report_path}")
        self.logger.info(f"📝 サマリーレポート: {summary_path}")
        
        return summary_path
    
    def _generate_strategy_summary_report(self) -> str:
        """戦略別統計サマリーレポート生成"""
        report = []
        report.append("=" * 70)
        report.append("DSSMS 戦略別統計問題分析レポート")
        report.append("=" * 70)
        report.append(f"実行日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}")
        report.append("")
        
        # 問題サマリー
        report.append("🚨 発見された問題:")
        report.append("-" * 40)
        if self.issues:
            for i, issue in enumerate(self.issues, 1):
                report.append(f"{i}. [{issue['severity'].upper()}] {issue['description']}")
                if 'details' in issue:
                    report.append(f"   詳細: {issue['details']}")
                report.append("")
        else:
            report.append("問題は発見されませんでした。")
        
        # Excel分析結果
        excel_data = self.analysis_results.get('excel_analysis', {})
        if excel_data and not excel_data.get('error'):
            report.append("📊 Excel戦略別統計分析:")
            report.append("-" * 40)
            report.append(f"戦略別統計シート: {'発見' if excel_data.get('sheet_found') else '未発見'}")
            if excel_data.get('sheet_found'):
                report.append(f"シート名: {excel_data.get('sheet_name')}")
                report.append(f"データ行数: {excel_data.get('total_rows', 0)}行")
                report.append(f"戦略数: {excel_data.get('strategy_count', 0)}種類")
                report.append(f"戦略名: {excel_data.get('unique_strategies', [])}")
                
                win_rate_data = excel_data.get('win_rate_analysis', {})
                if not win_rate_data.get('error'):
                    all_zero = win_rate_data.get('all_zero', True)
                    report.append(f"勝率状態: {'すべて0%' if all_zero else '正常'}")
            report.append("")
        
        # JSON分析結果
        json_data = self.analysis_results.get('json_analysis', {})
        if json_data and not json_data.get('error'):
            report.append("📦 JSON戦略データ分析:")
            report.append("-" * 40)
            
            trades_strategies = json_data.get('strategies_in_trades', {})
            if trades_strategies:
                report.append(f"取引データ内戦略数: {trades_strategies.get('strategy_count', 0)}種類")
                report.append(f"取引データ内戦略: {trades_strategies.get('unique_strategies', [])}")
            
            calc_performance = json_data.get('calculated_strategy_performance', {})
            if calc_performance:
                report.append("計算済み戦略パフォーマンス:")
                for strategy, perf in calc_performance.items():
                    report.append(f"  {strategy}: 勝率{perf['win_rate']:.1f}% (取引{perf['total_trades']}件)")
            
            report.append("")
        
        # 統一出力エンジン分析
        engine_data = self.analysis_results.get('engine_analysis', {})
        if engine_data and not engine_data.get('error'):
            report.append("🔧 統一出力エンジン分析:")
            report.append("-" * 40)
            report.append(f"エンジンファイル: {'存在' if engine_data.get('engine_file_exists') else '不在'}")
            report.append(f"戦略処理メソッド: {engine_data.get('strategy_methods_found', [])}")
            report.append(f"Excel生成機能: {'あり' if engine_data.get('excel_generation_found') else 'なし'}")
            report.append("")
        
        # 修正提案
        report.append("🔧 修正提案:")
        report.append("-" * 40)
        
        if any(issue['category'] in ['single_strategy', 'missing_strategies'] for issue in self.issues):
            report.append("1. 戦略別統計データ生成の修正")
            report.append("   - 統一出力エンジンで7つの戦略ごとの統計を生成")
            report.append("   - 取引データから戦略別にパフォーマンスを計算")
            report.append("   - Excel出力で戦略別統計シートを正しく生成")
            report.append("")
        
        if any(issue['category'] == 'zero_win_rates' for issue in self.issues):
            report.append("2. 勝率計算の修正")
            report.append("   - 戦略別の勝ち取引・負け取引を正確にカウント")
            report.append("   - 勝率 = (勝ち取引数 / 総取引数) × 100 の計算を実装")
            report.append("   - 各戦略の損益データを基に勝敗判定")
            report.append("")
        
        report.append("3. 統一出力エンジンの強化")
        report.append("   - 戦略別統計生成メソッドの追加/修正")
        report.append("   - Excel戦略別統計シートの自動生成")
        report.append("   - JSONデータとExcel出力の整合性確保")
        report.append("")
        
        return "\n".join(report)

def main():
    """メイン実行"""
    analyzer = DSSMSStrategyStatsAnalyzer()
    
    # 最新のファイルパスを指定
    excel_path = "backtest_results/dssms_results/dssms_unified_backtest_20250908_150951.xlsx"
    json_path = "backtest_results/dssms_results/dssms_unified_data_20250908_150951.json"
    text_path = "backtest_results/dssms_results/dssms_unified_report_20250908_150951.txt"
    
    print("🔍 DSSMS戦略別統計問題分析システム")
    print("=" * 60)
    print("対象ファイル:")
    print(f"Excel: {excel_path}")
    print(f"JSON: {json_path}")
    print(f"Text: {text_path}")
    print("=" * 60)
    
    # 総合分析実行
    summary_path = analyzer.generate_comprehensive_report(excel_path, json_path, text_path)
    
    print(f"\n✅ 分析完了")
    print(f"📋 サマリーレポート: {summary_path}")
    print("\n次のステップ:")
    print("1. サマリーレポートで問題を確認")
    print("2. 統一出力エンジンの戦略別統計生成機能を修正")
    print("3. Excelファイルで戦略別統計シートを確認")

if __name__ == "__main__":
    main()
