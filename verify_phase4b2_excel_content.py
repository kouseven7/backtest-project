#!/usr/bin/env python3
"""
Phase 4-B-2-1 Excel出力品質検証スクリプト
修正されたsimple_excel_exporter.pyによるN/A→実データ変換確認
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys

def verify_excel_content():
    """Phase 4-B-2-1修正後のExcel内容検証"""
    
    # 最新のExcelファイルを特定
    excel_dir = Path("backtest_results/improved_results")
    excel_files = list(excel_dir.glob("improved_backtest_*.xlsx"))
    
    if not excel_files:
        print("[ERROR] Excelファイルが見つかりません")
        return
    
    # 最新ファイルを取得
    latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
    print(f"[SEARCH] 検証対象: {latest_file}")
    
    try:
        # Excelファイル読み込み
        wb = openpyxl.load_workbook(latest_file)
        print(f"[CHART] シート数: {len(wb.sheetnames)}")
        print(f"[LIST] シート名: {wb.sheetnames}")
        
        # 各シートの内容確認
        for sheet_name in wb.sheetnames:
            print(f"\n=== シート: {sheet_name} ===")
            ws = wb[sheet_name]
            print(f"   最大行: {ws.max_row}, 最大列: {ws.max_column}")
            
            if ws.max_row > 1:  # ヘッダー以外にデータがある場合
                # 最初の数行のデータを確認
                print("   📝 サンプルデータ（最初の3行）:")
                for row in range(1, min(5, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(6, ws.max_column + 1)):  # 最初の5列
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is None:
                            cell_value = "None"
                        elif isinstance(cell_value, (int, float)):
                            cell_value = f"{cell_value:,.2f}" if isinstance(cell_value, float) else str(cell_value)
                        else:
                            cell_value = str(cell_value)[:20]  # 長い文字列は切り詰め
                        row_data.append(cell_value)
                    print(f"     Row {row}: {' | '.join(row_data)}")
            
            # 特にSummaryシートを詳細確認
            if 'Summary' in sheet_name or 'サマリー' in sheet_name:
                print("   [TARGET] Summary詳細分析:")
                for row in range(1, min(20, ws.max_row + 1)):
                    for col in range(1, min(3, ws.max_column + 1)):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            print(f"     [{row},{col}]: {cell.value}")
    
    except Exception as e:
        print(f"[ERROR] Excel読み込みエラー: {e}")
        return
    
    # pandasでも確認
    try:
        print("\n=== pandas読み込み確認 ===")
        # 最初のシートをpandasで読み込み
        df = pd.read_excel(latest_file, sheet_name=0)
        print(f"DataFrame形状: {df.shape}")
        print(f"列名: {list(df.columns)}")
        
        # N/A値の確認
        na_count = df.isna().sum().sum()
        print(f"N/A値の数: {na_count}")
        
        if na_count > 0:
            print("[WARNING]  N/A値が残存しています")
            print(df.isna().sum())
        else:
            print("[OK] N/A値は検出されませんでした")
        
        # 数値列の統計確認
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            print(f"\n[CHART] 数値列統計:")
            print(df[numeric_cols].describe())
    
    except Exception as e:
        print(f"[WARNING]  pandas読み込み警告: {e}")
    
    print(f"\n[OK] Phase 4-B-2-1 Excel品質検証完了")

if __name__ == "__main__":
    verify_excel_content()