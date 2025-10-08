#!/usr/bin/env python3
"""
Phase 4-B-2-3: 残存N/A値の特定・完全除去
"""

import pandas as pd
import openpyxl
from pathlib import Path

def find_remaining_na_values():
    """残存N/A値の詳細特定"""
    
    excel_dir = Path("backtest_results/improved_results")
    excel_files = list(excel_dir.glob("improved_backtest_*.xlsx"))
    
    if not excel_files:
        print("[ERROR] Excelファイルが見つかりません")
        return
    
    latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
    print(f"[SEARCH] 残存N/A値分析: {latest_file}")
    
    try:
        # openpyxlで詳細分析
        wb = openpyxl.load_workbook(latest_file)
        
        for sheet_name in wb.sheetnames:
            print(f"\n=== シート: {sheet_name} ===")
            ws = wb[sheet_name]
            
            na_count = 0
            na_locations = []
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value == 'N/A' or cell_value is None:
                        if cell_value == 'N/A':
                            na_count += 1
                            na_locations.append((row, col, cell_value))
            
            print(f"   N/A値の数: {na_count}")
            if na_locations:
                print("   N/A値の位置:")
                for row, col, value in na_locations:
                    print(f"     [{row},{col}]: {value}")
        
        # pandasでも確認
        print("\n=== pandas N/A分析 ===")
        df = pd.read_excel(latest_file, sheet_name=0)
        
        # 各列のN/A値をチェック
        for col_name in df.columns:
            na_count = df[col_name].isna().sum()
            if na_count > 0:
                print(f"   列 '{col_name}': {na_count}件のN/A値")
                na_indices = df[df[col_name].isna()].index.tolist()
                print(f"     行インデックス: {na_indices}")
                
                # N/A値の前後の値も表示
                for idx in na_indices:
                    if idx > 0:
                        prev_val = df.iloc[idx-1][col_name] if idx-1 < len(df) else "なし"
                    else:
                        prev_val = "なし"
                    
                    if idx < len(df) - 1:
                        next_val = df.iloc[idx+1][col_name] if idx+1 < len(df) else "なし"
                    else:
                        next_val = "なし"
                    
                    print(f"       行{idx}: 前={prev_val}, 現在=N/A, 次={next_val}")
    
    except Exception as e:
        print(f"[ERROR] 分析エラー: {e}")

if __name__ == "__main__":
    find_remaining_na_values()